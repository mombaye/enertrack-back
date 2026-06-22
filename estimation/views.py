# ─────────────────────────────────────────────────────────────────────────────
# estimation/views.py
# ─────────────────────────────────────────────────────────────────────────────

from rest_framework import viewsets, mixins, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils import timezone
from rest_framework.views import APIView
from .models import EstimationBatch, EstimationResult
from .serializers import EstimationBatchSerializer, EstimationResultSerializer
from .tasks import launch_estimation_batch
from rest_framework.parsers import MultiPartParser


class EstimationBatchViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    viewsets.GenericViewSet,
):
    """
    list   GET  /estimation/batches/
    detail GET  /estimation/batches/{id}/
    launch POST /estimation/batches/launch/
    status GET  /estimation/batches/{id}/status/
    """
    serializer_class   = EstimationBatchSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = EstimationBatch.objects.all().order_by("-year", "-month")
        year  = self.request.query_params.get("year")
        month = self.request.query_params.get("month")
        if year:  qs = qs.filter(year=int(year))
        if month: qs = qs.filter(month=int(month))
        return qs

    @action(detail=False, methods=["post"], url_path="launch")
    def launch(self, request):
        year  = request.data.get("year")
        month = request.data.get("month")

        if not year or not month:
            return Response(
                {"detail": "year et month sont requis."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            year, month = int(year), int(month)
        except (ValueError, TypeError):
            return Response(
                {"detail": "year et month doivent être des entiers."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        batch, created = EstimationBatch.objects.get_or_create(
            year=year, month=month,
            defaults={
                "created_by": request.user if request.user.is_authenticated else None,
                "status": EstimationBatch.Status.PENDING,
            },
        )

        if not created:
            if batch.status == EstimationBatch.Status.RUNNING:
                return Response(
                    {"detail": "Une estimation est déjà en cours pour ce mois.", "batch_id": batch.id},
                    status=status.HTTP_409_CONFLICT,
                )
            # Reset pour re-lancement
            batch.results.all().delete()
            batch.status      = EstimationBatch.Status.PENDING
            batch.finished_at = None
            batch.celery_task_id = None
            batch.total = 0
            batch.count_acm = 0
            batch.count_grid = 0
            batch.count_senelec = 0
            batch.count_target = 0
            batch.count_theorique = 0
            batch.count_histo = 0
            batch.count_nc = 0
            batch.count_hors_scope = 0
            batch.created_by = request.user if request.user.is_authenticated else None
            batch.save()

        task = launch_estimation_batch.delay(batch.id)
        batch.celery_task_id = task.id
        batch.save(update_fields=["celery_task_id"])

        return Response(
            {
                "batch_id":       batch.id,
                "celery_task_id": task.id,
                "status":         batch.status,
                "label":          batch.label,
                "detail":         "Estimation lancée." if created else "Estimation re-lancée.",
            },
            status=status.HTTP_202_ACCEPTED,
        )

    @action(methods=["get"], detail=True, url_path="status")
    def batch_status(self, request, pk=None):
        batch = self.get_object()
        return Response({
            "batch_id":     batch.id,
            "label":        batch.label,
            "status":       batch.status,
            "created_at":   batch.created_at,
            "finished_at":  batch.finished_at,
           "counters": {
                "total": batch.total,
                "count_acm": batch.count_acm,
                "count_grid": batch.count_grid,
                "count_senelec": batch.count_senelec,
                "count_target": batch.count_target,
                "count_theorique": batch.count_theorique,
                "count_histo": batch.count_histo,
                "count_nc": batch.count_nc,
                "count_hors_scope": batch.count_hors_scope,
            },
        })


class EstimationResultViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    viewsets.GenericViewSet,
):
    """
    list   GET /estimation/results/
    detail GET /estimation/results/{id}/

    Filtres : ?batch= ?source= ?site= ?fiabilite_grid=
    """
    serializer_class   = EstimationResultSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = EstimationResult.objects.select_related("batch", "site").order_by("-computed_at")

        batch          = self.request.query_params.get("batch")
        source         = self.request.query_params.get("source")
        site           = self.request.query_params.get("site")
        fiabilite_grid = self.request.query_params.get("fiabilite_grid")
        year           = self.request.query_params.get("year")
        month          = self.request.query_params.get("month")

        if batch:          qs = qs.filter(batch_id=batch)
        if source:         qs = qs.filter(source_utilisee=source.upper())
        if site:           qs = qs.filter(site__site_id=site)
        if fiabilite_grid: qs = qs.filter(fiabilite_grid=fiabilite_grid.upper())
        if year:           qs = qs.filter(batch__year=int(year))
        if month:          qs = qs.filter(batch__month=int(month))

        return qs


 
class EstimationHistoryImportView(APIView):
    """
    POST /estimation/batches/import-history/
    Importe le fichier Provisions_GRID_Conso.xlsx comme historique d'estimations.
 
    Format attendu — Feuil1 :
      site_ID | Site_Name | Conso_Kwh | Montant | Source | Mois
 
    Crée un EstimationBatch (status=DONE) par mois trouvé dans le fichier,
    et un EstimationResult par ligne avec le mapping source ci-dessous.
    """
    parser_classes     = [MultiPartParser]
    permission_classes = [IsAuthenticated]
 
    # ── Mapping source fichier → EstimationResult.Source ─────────────────────
    SOURCE_MAP = {
        # FMS / Grid
        "gfms":                         "GRID",
        "grid":                         "GRID",
        # Sénélec
        "estimation_senelec":           "HISTO",   # ancien label = historique Sénélec
        "estimation senelec":           "HISTO",
        "estmation coherence fuel":     "HISTO",
        "estimation qowisio":           "HISTO",
        # Target catalogue
        "estimation target":            "TARGET",
        # Théorique
        "estimation théorique":         "THEORIQUE",
        "estimation theorique":         "THEORIQUE",
        # Hors scope / statuts spéciaux
        "hors scope/ off-grid / sntl ": "HORS_SCOPE",
        "fournisseurs tiers":           "HORS_SCOPE",
        "site démonté":                 "HORS_SCOPE",
        "site demonte":                 "HORS_SCOPE",
        "closing board ":               "HORS_SCOPE",
        "new_site":                     "HORS_SCOPE",
        "en raccordement ":             "HORS_SCOPE",
        "site avec ge actuellement ":   "HORS_SCOPE",
        "integrate in nov invoice":     "HORS_SCOPE",
        "trouver la source ":           "NC",
    }
 
    def _map_source(self, raw: str) -> str:
        if not raw:
            return "NC"
        key = raw.strip().lower()
        return self.SOURCE_MAP.get(key, "NC")
 
    def post(self, request, *args, **kwargs):
        import io
        import openpyxl
        from decimal import Decimal, InvalidOperation
        from datetime import datetime
        from django.db import transaction
        from core.models import Site
        from billing.models import ContractSiteLink
 
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Aucun fichier fourni."}, status=400)
 
        file_bytes = f.read()
 
        # ── Lecture Excel ─────────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception as e:
            return Response({"detail": f"Impossible de lire le fichier : {e}"}, status=400)
 
        ws = wb["Feuil1"] if "Feuil1" in wb.sheetnames else wb.worksheets[0]
 
        # ── Préchargement des sites connus ────────────────────────────────────
        site_map: dict[str, object] = {
            s.site_id: s
            for s in Site.objects.all()
        }
        contract_map: dict[int, str] = dict(
            ContractSiteLink.objects
            .filter(site__in=site_map.values())
            .values_list("site_id", "numero_compte_contrat")
        )
 
        # ── Parse des lignes — groupage par (year, month) ─────────────────────
        # Structure : { (year, month): [ {site_id, conso, montant, source}, ... ] }
        from collections import defaultdict
        batches_data: dict[tuple, list] = defaultdict(list)
        skipped_site = 0
        skipped_date = 0
        total_parsed = 0
 
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
 
            site_id_raw = str(row[0]).strip()
            conso_raw   = row[2]
            montant_raw = row[3]
            source_raw  = str(row[4]).strip() if row[4] else ""
            date_raw    = row[5]
 
            # Parse date
            if date_raw is None:
                skipped_date += 1
                continue
            if hasattr(date_raw, "year"):
                year, month = date_raw.year, date_raw.month
            else:
                try:
                    dt = datetime.strptime(str(date_raw)[:7], "%Y-%m")
                    year, month = dt.year, dt.month
                except ValueError:
                    skipped_date += 1
                    continue
 
            # Parse valeurs numériques
            try:
                conso   = Decimal(str(conso_raw))   if conso_raw   is not None else Decimal("0")
                montant = Decimal(str(montant_raw)) if montant_raw is not None else Decimal("0")
            except (InvalidOperation, ValueError):
                conso   = Decimal("0")
                montant = Decimal("0")
 
            batches_data[(year, month)].append({
                "site_id":  site_id_raw,
                "conso":    conso,
                "montant":  montant,
                "source":   self._map_source(source_raw),
            })
            total_parsed += 1
 
        wb.close()
 
        if not batches_data:
            return Response({"detail": "Aucune ligne valide trouvée."}, status=400)
 
        # ── Upsert en base ────────────────────────────────────────────────────
        created_batches = 0
        updated_batches = 0
        created_results = 0
        updated_results = 0
        skipped_results = 0
 
        with transaction.atomic():
            for (year, month), rows in sorted(batches_data.items()):
                # Batch
                batch, b_created = EstimationBatch.objects.update_or_create(
                    year=year, month=month,
                    defaults={
                        "status":     EstimationBatch.Status.DONE,
                        "created_by": request.user,
                        "finished_at": __import__("django.utils.timezone",
                                       fromlist=["now"]).now(),
                    },
                )
                if b_created:
                    created_batches += 1
                else:
                    updated_batches += 1
 
                # Compteurs pour le batch
                counters = {
                    "ACM": 0, "GRID": 0, "HISTO": 0, "THEORIQUE": 0,
                    "TARGET": 0, "NC": 0, "HORS_SCOPE": 0, "SENELEC": 0,
                }
 
                for item in rows:
                    site = site_map.get(item["site_id"])
                    if not site:
                        skipped_results += 1
                        continue
 
                    contrat = contract_map.get(site.pk, "")
 
                    _, r_created = EstimationResult.objects.update_or_create(
                        batch=batch,
                        site=site,
                        defaults={
                            "numero_compte_contrat": contrat,
                            "source_utilisee":       item["source"],
                            "conso_estimee_kwh":     item["conso"] if item["conso"] > 0 else None,
                            "montant_estime":        item["montant"] if item["montant"] > 0 else None,
                            # Flags de disponibilité
                            "acm_disponible":        item["source"] == "ACM",
                            "grid_disponible":       item["source"] == "GRID",
                            "histo_disponible":      item["source"] in ("HISTO", "SENELEC"),
                        },
                    )
                    if r_created:
                        created_results += 1
                    else:
                        updated_results += 1
 
                    src = item["source"]
                    if src in counters:
                        counters[src] += 1
 
                # Mise à jour compteurs batch
                batch.total           = len(rows)
                batch.count_acm       = counters["ACM"]
                batch.count_grid      = counters["GRID"]
                batch.count_histo     = counters["HISTO"] + counters["SENELEC"]
                batch.count_nc        = counters["NC"]
                batch.count_hors_scope= counters["HORS_SCOPE"]
                batch.save(update_fields=[
                    "total", "count_acm", "count_grid",
                    "count_histo", "count_nc", "count_hors_scope",
                ])
 
        return Response(
            {
                "message":        "Import historique terminé.",
                "total_parsed":   total_parsed,
                "periods":        len(batches_data),
                "created_batches":created_batches,
                "updated_batches":updated_batches,
                "created_results":created_results,
                "updated_results":updated_results,
                "skipped_unknown_sites": skipped_results,
                "skipped_invalid_dates": skipped_date,
            },
            status=status.HTTP_201_CREATED,
        )
 