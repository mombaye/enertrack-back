# financial/views.py
from __future__ import annotations

import calendar
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

import pandas as pd
import zipfile
import xml.etree.ElementTree as ET

from django.shortcuts import get_object_or_404

from django.db.models import Q, Sum, Avg, Count, Max
from django.utils import timezone

from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from billing.models import ContractMonth, ContractSiteLink
from certification.models import CertificationResult
from core.models import Site

from .models import FinancialFeeRule, FinancialEvaluation, SiteMonthlyLoad
from billing.models import MonthlySynthesis as BillingMonthlySynthesis
from financial.services.conso_service import FinancialConsoService, ConsoMonthData


from django.db import transaction
import logging

from estimation.models import EstimationResult

logger = logging.getLogger(__name__)

DEC0 = Decimal("0")
Q3 = Decimal("0.001")


# ─────────────────────────────────────────────────────────────────────────────
# TARGET SOLAIRE
# x = 100% pour l'instant — sera paramétrable depuis la page "Suivi Conso"
# Formule : Min(x% × Load/1000 × 24 × NbrJrs ; Cap_kW × 24 × NbrJrs)
# ─────────────────────────────────────────────────────────────────────────────

_SOLAR_INVERTER_KW: dict[str, Decimal] = {
    "GG_S0": Decimal("0.736"),
    "GG_S1": Decimal("1.453"),
    "GG_S2": Decimal("2.180"),
    "GG_S3": Decimal("2.907"),
}
SOLAR_X_PCT = Decimal("1.0")   # 100% — TODO: page Suivi Conso


def _d(v) -> Decimal:
    if v is None:
        return DEC0
    return Decimal(str(v))

def _q3(v: Decimal) -> Decimal:
    return v.quantize(Q3, rounding=ROUND_HALF_UP)

    
def _map_load_source(raw: str | None, default: str = SiteMonthlyLoad.Source.PREVISIONNEL) -> str:
    """
    Mappe la colonne 'Source' du fichier vers le statut de load interne.
    Règle métier voulue :
      - aligné
      - non aligné (stocké comme PREVISIONNEL)
      - manuel seulement pour les edits inline
    """
    if raw is None:
        return default

    s = str(raw).strip().lower()
    if not s:
        return default

    # cas aligné / officiel
    if any(k in s for k in ["align", "aligné", "aligne", "sntl", "cmst", "officiel", "senelec"]):
        return SiteMonthlyLoad.Source.ALIGNE

    # tout le reste importé depuis ces fichiers = non aligné / prévisionnel
    return SiteMonthlyLoad.Source.PREVISIONNEL



def _is_blank(x) -> bool:
    if x is None:
        return True
    if isinstance(x, float):
        import math
        return math.isnan(x)
    return str(x).strip() in ("", "nan", "None", "N/A", "#N/A")

def _norm_config(v: str | None) -> str | None:
    """Normalise Indoor/Outdoor en majuscules."""
    if not v:
        return None
    s = str(v).strip().upper()
    if s in ("INDOOR", "IN DOOR", "IN-DOOR"):
        return "INDOOR"
    if s in ("OUTDOOR", "OUT DOOR", "OUT-DOOR"):
        return "OUTDOOR"
    return s

def _norm_typo(v: str | None) -> str | None:
    if not v:
        return None
    return str(v).strip()






def _compute_solar_target(
    typology: str | None, load_w: int | None, nb_jours: int
) -> Decimal | None:
    if not typology or not nb_jours:
        return None

    # Normalise les espaces → underscores pour matcher "GG_S3" dans "D2_GG S3"
    typo_normalized = typology.replace(" ", "_").upper()

    inverter_kw: Decimal | None = None
    for key, kw in _SOLAR_INVERTER_KW.items():
        if key in typo_normalized:
            inverter_kw = kw
            break

    if inverter_kw is None:
        return None

    cap_kwh = inverter_kw * 24 * nb_jours

    if load_w:
        load_kwh = SOLAR_X_PCT * Decimal(str(load_w)) / 1000 * 24 * nb_jours
        return min(load_kwh, cap_kwh).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

    return cap_kwh.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)


"""def _apply_period_filters(qs, request, *, year_field: str = "year", month_field: str = "month"):
    year = request.query_params.get("year")
    month = request.query_params.get("month")
    month_start = request.query_params.get("month_start")
    month_end = request.query_params.get("month_end")


    year_start  = int(request.query_params.get("year_start", year))
    year_end    = int(request.query_params.get("year_end",   year))

    filters = Q()

    if year:
        filters &= Q(**{year_field: int(year)})

    if month:
        filters &= Q(**{month_field: int(month)})
        return qs.filter(filters)

    ms = int(month_start) if month_start else None
    me = int(month_end) if month_end else None

    if ms and me and ms > me:
        ms, me = me, ms

    if ms:
        filters &= Q(**{f"{month_field}__gte": ms})
    if me:
        filters &= Q(**{f"{month_field}__lte": me})

    return qs.filter(filters)"""



# ─────────────────────────────────────────────────────────────────────────────
# PATCH 1/2 — Remplace _apply_period_filters
# Supporte :
#   - Ancien format :  ?year=2025&month_start=1&month_end=12
#   - Nouveau format : ?year_start=2024&month_start=10&year_end=2025&month_end=3
#   - Mois unique :    ?year=2025&month=6
# ─────────────────────────────────────────────────────────────────────────────

def _apply_period_filters(qs, request, *, year_field: str = "year", month_field: str = "month"):
    """
    Filtre un queryset sur une plage de périodes (year, month).
    Accepte deux syntaxes :
      • Mono-année (rétrocompat) : year + month_start / month_end
      • Inter-années (nouveau)   : year_start + month_start + year_end + month_end
    """
    year_raw        = request.query_params.get("year")
    month_raw       = request.query_params.get("month")
    month_start_raw = request.query_params.get("month_start")
    month_end_raw   = request.query_params.get("month_end")
    year_start_raw  = request.query_params.get("year_start")
    year_end_raw    = request.query_params.get("year_end")

    # ── Raccourci mois unique (usage legacy) ─────────────────────────────────
    if year_raw and month_raw:
        return qs.filter(**{year_field: int(year_raw), month_field: int(month_raw)})

    # ── Nouveau format inter-années ──────────────────────────────────────────
    if year_start_raw and year_end_raw:
        ys = int(year_start_raw)
        ye = int(year_end_raw)
        ms = int(month_start_raw) if month_start_raw else 1
        me = int(month_end_raw)   if month_end_raw   else 12

        # Normalise si l'utilisateur a inversé les bornes
        if (ys, ms) > (ye, me):
            ys, ms, ye, me = ye, me, ys, ms

        if ys == ye:
            # Même année → filtre simple
            return qs.filter(**{
                year_field:                ys,
                f"{month_field}__gte": ms,
                f"{month_field}__lte": me,
            })

        # Années différentes : trois segments
        #   • Année de début  : month >= ms
        #   • Années du milieu: toute l'année
        #   • Année de fin    : month <= me
        return qs.filter(
            Q(**{year_field: ys, f"{month_field}__gte": ms})
            | Q(**{f"{year_field}__gt": ys, f"{year_field}__lt": ye})
            | Q(**{year_field: ye, f"{month_field}__lte": me})
        )

    # ── Ancien format mono-année (rétrocompat) ───────────────────────────────
    if year_raw:
        filters = Q(**{year_field: int(year_raw)})
        if month_start_raw:
            filters &= Q(**{f"{month_field}__gte": int(month_start_raw)})
        if month_end_raw:
            filters &= Q(**{f"{month_field}__lte": int(month_end_raw)})
        return qs.filter(filters)

    # Aucun filtre temporel → retourne tout
    return qs


# ─────────────────────────────────────────────────────────────────────────────
# XLSX reader (format simple : col A=Typo, B=Load, C=Config, D=Redev, E=Cible, F=Cible/J)
# ─────────────────────────────────────────────────────────────────────────────

NS = {"ss": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _read_fee_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Parse le fichier Redevance_et_Cible.xlsx et retourne une liste de dicts :
    {typology, configuration, load_w, redevance, cible_kwh, cible_kwh_j}
    Lignes avec redevance=0 exclues (typologies hors scope).
    """
    import io

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
        # Shared strings
        ss_xml = z.read("xl/sharedStrings.xml")
        ss_root = ET.fromstring(ss_xml)
        shared = [
            "".join(t.text or "" for t in si.findall(".//ss:t", NS))
            for si in ss_root.findall(".//ss:si", NS)
        ]

        ws_xml = z.read("xl/worksheets/sheet1.xml")
        ws_root = ET.fromstring(ws_xml)

    records = []
    for row_el in ws_root.findall(".//ss:row", NS):
        rnum = int(row_el.get("r", 0))
        if rnum == 1:
            continue  # skip header row

        cells = {}
        for c in row_el.findall("ss:c", NS):
            ref  = c.get("r", "")
            t    = c.get("t", "")
            v_el = c.find("ss:v", NS)
            val  = v_el.text if v_el is not None else None
            if t == "s" and val is not None:
                val = shared[int(val)]
            elif val is not None:
                try:
                    val = float(val)
                except Exception:
                    pass
            col_letter = "".join(ch for ch in ref if ch.isalpha())
            cells[col_letter] = val

        typo  = _norm_typo(cells.get("A"))
        load  = cells.get("B")
        conf  = _norm_config(cells.get("C"))
        redev = cells.get("D")
        cible = cells.get("E")
        cible_j = cells.get("F")

        # Skip toute ligne dont les colonnes clés ne sont pas numériques
        # (lignes d'en-tête supplémentaires, lignes de titre fusionnées, etc.)
        if not typo or not conf:
            continue

        try:
            load_w = int(float(load))
        except (TypeError, ValueError):
            continue  # "Load Aligné", "Load (W)", lignes de titre → skip

        try:
            redev_f = float(redev)
        except (TypeError, ValueError):
            continue  # Cellule texte dans colonne Redevance → skip

        # Exclure les typologies hors scope (redevance = 0)
        if redev_f == 0:
            continue

        try:
            records.append({
                "typology":      typo,
                "configuration": conf,
                "load_w":        load_w,
                "redevance":     Decimal(str(redev_f)).quantize(Q3, rounding=ROUND_HALF_UP),
                "cible_kwh":     Decimal(str(float(cible))).quantize(Q3, rounding=ROUND_HALF_UP) if cible is not None else None,
                "cible_kwh_j":   Decimal(str(float(cible_j))) if cible_j is not None else None,
            })
        except (TypeError, ValueError, Exception):
            continue  # Ignore toute ligne malformée restante

    return records


# ─────────────────────────────────────────────────────────────────────────────
# Lookup redevance avec fallback load supérieur
# ─────────────────────────────────────────────────────────────────────────────

def _get_fee_rule(
    typology: str, configuration: str, load_w: int
) -> tuple[Optional[FinancialFeeRule], bool]:
    """
    Retourne (rule, hors_catalogue).
    hors_catalogue=True si le match est sur un load supérieur (fallback CDC).
    """
    # 1. Match exact
    rule = FinancialFeeRule.objects.filter(
        typology=typology,
        configuration=configuration,
        load_w=load_w,
    ).first()
    if rule:
        return rule, False

    # 2. Fallback : load supérieur le plus proche
    rule = (
        FinancialFeeRule.objects.filter(
            typology=typology,
            configuration=configuration,
            load_w__gt=load_w,
        )
        .order_by("load_w")
        .first()
    )
    return rule, (rule is not None)


# ─────────────────────────────────────────────────────────────────────────────
# Calcul récurrence NOK
# ─────────────────────────────────────────────────────────────────────────────

def _compute_recurrence(site_id: int, year: int, month: int) -> tuple[int, str | None]:
    """
    Compte les mois NOK consécutifs jusqu'à (year, month) inclus,
    en remontant dans l'historique de FinancialEvaluation.
    Retourne (nb_mois_nok, recurrence_type).
    """
    # On récupère l'historique trié desc
    history = list(
        FinancialEvaluation.objects.filter(site_id=site_id)
        .filter(
            Q(year__lt=year) | Q(year=year, month__lte=month)
        )
        .order_by("-year", "-month")
        .values_list("year", "month", "marge_statut")[:24]  # max 24 mois de lookback
    )

    consecutive = 0
    for y, m, st in history:
        if st == FinancialEvaluation.MargeStatut.NOK:
            consecutive += 1
        else:
            break

    if consecutive == 0:
        return 0, None
    if 3 <= consecutive < 6:
        # light seulement si M-4 était OK  (déjà garanti par le break ci-dessus)
        # On vérifie que le mois juste avant la séquence NOK était OK
        if len(history) > consecutive:
            _, _, prev_st = history[consecutive]
            if prev_st == FinancialEvaluation.MargeStatut.OK:
                return consecutive, FinancialEvaluation.RecurrenceType.LIGHT
        return consecutive, FinancialEvaluation.RecurrenceType.LIGHT
    if consecutive >= 6:
        return consecutive, FinancialEvaluation.RecurrenceType.CRITIQUE

    return consecutive, None


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 1 : Import catalogue FinancialFeeRule
# ─────────────────────────────────────────────────────────────────────────────

class FinancialFeeRuleImportView(APIView):
    """
    POST /financial/fee-rules/import/
    Fichier : Redevance_et_Cible_Akt.xlsx
    """
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Aucun fichier fourni."}, status=400)

        file_bytes = f.read()
        try:
            records = _read_fee_xlsx(file_bytes)
        except Exception as e:
            return Response({"detail": f"Erreur lecture fichier : {e}"}, status=400)

        if not records:
            return Response({"detail": "Aucune ligne valide trouvée (toutes redevance=0 ?)."}, status=400)

        created = updated = skipped = 0
        errors = []

        with transaction.atomic():
            for rec in records:
                try:
                    obj, was_created = FinancialFeeRule.objects.update_or_create(
                        typology=rec["typology"],
                        configuration=rec["configuration"],
                        load_w=rec["load_w"],
                        defaults={
                            "redevance":   rec["redevance"],
                            "cible_kwh":   rec["cible_kwh"],
                            "cible_kwh_j": rec["cible_kwh_j"],
                            "source_filename": f.name,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                except Exception as e:
                    skipped += 1
                    errors.append({"record": str(rec), "error": str(e)})

        return Response(
            {
                "message": "Import FinancialFeeRule terminé.",
                "total_parsed": len(records),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors_sample": errors[:20],
            },
            status=status.HTTP_201_CREATED,
        )


class FinancialFeeRuleListView(APIView):
    """
    GET /financial/fee-rules/?typology=A_Ax_GG&configuration=OUTDOOR
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = FinancialFeeRule.objects.all().order_by("typology", "configuration", "load_w")
        typo = request.query_params.get("typology")
        conf = request.query_params.get("configuration")
        if typo:
            qs = qs.filter(typology__icontains=typo)
        if conf:
            qs = qs.filter(configuration__iexact=conf)

        data = list(qs.values(
            "id", "typology", "configuration", "load_w",
            "redevance", "cible_kwh", "cible_kwh_j",
        ))
        return Response({"count": len(data), "results": data})


# ─────────────────────────────────────────────────────────────────────────────
# VIEW 2 : Import SiteMonthlyLoad (CSV)
# ─────────────────────────────────────────────────────────────────────────────

def _detect_load_format(df: pd.DataFrame, sheet_name: str = "") -> str:
    """
    Détecte le format du fichier de loads :
    - "alignement"   : Load_SNTL_CMST (Sheet "Alignement DC") — colonnes = mois, lignes = sites
    - "proposition"  : Proposition_de_load_MMYYYY — lignes = sites, col Average_Load + Month
    - "simple"       : Format générique Site_ID | Année | Mois | Load
    """
    cols = [str(c).strip().lower() for c in df.columns]
    # Proposition : a les colonnes Average_Load et Month/Site_ID
    if any("average_load" in c for c in cols) or any("average" in c for c in cols):
        return "proposition"
    # Alignement : les colonnes sont des dates (datetime), pas de header nommé Load
    if any(c in ("load sept 2024", "load oct 2024", "load aligné") for c in cols):
        return "alignement"
    if "site id" in cols or "site_id" in cols:
        return "simple"
    return "unknown"


def _parse_load_alignement(file_bytes: bytes) -> list[dict]:
    """
    Parse Load_SNTL_CMST_Sept24-Aout25 — feuille "Alignement DC".
    Format : row 6 = headers (Site ID, Site Name, …, LOAD SEPT 2024, LOAD OCT 2024, …)
    Les colonnes Load sont identifiées par les dates en row 1.
    Retourne : [{site_id, year, month, load_w}, …]
    """
    import io
    import openpyxl
    from datetime import datetime as _dt

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Try to find "Alignement DC" sheet or fallback to first sheet
    target_sheet = None
    for sname in wb.sheetnames:
        if "alignement dc" in sname.lower() or "alignement" in sname.lower():
            target_sheet = sname
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Row 1 (index 0): month dates in various columns
    # Row 6 (index 5): column headers (Site ID, Site Name, …, LOAD SEPT 2024, …)
    # Data from row 7 (index 6)
    month_row   = all_rows[0] if all_rows else []
    header_row  = None
    data_start  = 7  # default

    # Find header row (contains "Site ID" or "Site_ID")
    for i, row in enumerate(all_rows[:10]):
        if row and any(str(v).strip().lower() in ("site id", "site_id") for v in row if v):
            header_row = row
            data_start = i + 2  # data starts next row
            break

    if header_row is None:
        wb.close()
        return []

    # Build col_idx → (year, month) mapping from month_row
    month_col_map = {}
    for j, v in enumerate(month_row):
        if v and hasattr(v, 'year'):
            month_col_map[j] = (v.year, v.month)

    # Find site_id column index
    site_id_col = None
    for j, h in enumerate(header_row):
        if h and str(h).strip().lower() in ("site id", "site_id"):
            site_id_col = j
            break

    if site_id_col is None:
        wb.close()
        return []

    records = []
    for row in all_rows[data_start - 1:]:
        if not row or not row[site_id_col]:
            continue
        site_id = str(row[site_id_col]).strip()
        if not site_id or site_id.lower() in ("site id", "nan", "none", ""):
            continue

        for j, (yr, mo) in month_col_map.items():
            if j >= len(row):
                continue
            val = row[j]
            if val is None:
                continue
            try:
                load_w = int(float(str(val)))
                if load_w > 0:
                    records.append({"site_id": site_id, "year": yr, "month": mo, "load_w": load_w})
            except (TypeError, ValueError):
                continue

    wb.close()
    return records


def _parse_load_proposition(file_bytes: bytes) -> list[dict]:
    """
    Parse Proposition_de_load_MMYYYY files.
    Format : row 2 = headers, data from row 3.
    Colonnes clés : Site_ID (A), Month (H), Average_Load (R ou S selon le fichier).
    Retourne : [{site_id, year, month, load_w}, …]
    """
    import io
    import openpyxl
    from datetime import datetime as _dt

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Find header row (row 2, index 1) — contains "Site_ID"
    header_row = None
    data_start = 3
    for i, row in enumerate(all_rows[:5]):
        if row and any(str(v).strip().lower() in ("site_id", "site id") for v in row if v):
            header_row = row
            data_start = i + 2
            break

    if header_row is None:
        wb.close()
        return []

    # Map header names → col index
    col_map = {str(h).strip().lower(): j for j, h in enumerate(header_row) if h}

    # NOTE: use explicit "in" check — col index can be 0 which is falsy
    site_col  = col_map["site_id"]  if "site_id"  in col_map else col_map.get("site id")
    month_col = col_map["month"]    if "month"    in col_map else col_map.get("h")
    # Average_Load is the column to use as the "official" monthly load
    avg_col   = col_map.get("average_load")
    if avg_col is None:
        # fallback: look for any key containing "average"
        for k, v in col_map.items():
            if "average" in k:
                avg_col = v
                break

    if site_col is None or avg_col is None:
        wb.close()
        return []

    records = []
    for row in all_rows[data_start - 1:]:
        if not row or site_col >= len(row) or not row[site_col]:
            continue
        site_id = str(row[site_col]).strip()
        if not site_id or site_id.lower() in ("site_id", "nan", "none", ""):
            continue

        # Parse month/year
        month_val = row[month_col] if month_col is not None and month_col < len(row) else None
        year, month = None, None
        if month_val is not None:
            if hasattr(month_val, 'year'):
                year, month = month_val.year, month_val.month
            else:
                try:
                    m = int(float(str(month_val)))
                    if 1 <= m <= 12:
                        # Month only — year inferred from filename context; default 2025
                        year, month = 2025, m
                    else:
                        # Could be a year value — skip
                        continue
                except (TypeError, ValueError):
                    continue

        if not year or not month:
            continue

        # Load value
        load_val = row[avg_col] if avg_col < len(row) else None
        if load_val is None:
            continue
        try:
            load_w = int(float(str(load_val)))
            if load_w > 0:
                records.append({"site_id": site_id, "year": year, "month": month, "load_w": load_w})
        except (TypeError, ValueError):
            continue

    wb.close()
    return records


class SiteMonthlyLoadImportView(APIView):
    """
    POST /financial/loads/import/

    Accepte 3 formats :

    1. FORMAT ALIGNÉ (Load_SNTL_CMST_Sept24-Aout25.xlsx)
       Feuille "Alignement DC" — colonnes = mois (date en row 1), Site ID en col A.
       Couvre Sept 2024 → Août 2025. ~3160 sites.
       Colonne load : valeur directe dans la cellule.

    2. FORMAT PROPOSITION (Proposition_de_load_MMYYYY.xlsx)
       Une ligne par site, col A = Site_ID, col H = Month, col R/S = Average_Load.
       Couvre Sept 2025 → Déc 2025 (un fichier par mois).

    3. FORMAT SIMPLE (CSV ou Excel générique)
       Colonnes : Site_ID | Année | Mois | Load
    """
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    SIMPLE_ALIASES = {
        "site_id":   ["Site_ID", "Site ID", "Code site", "site_id"],
        "site_name": ["Site_Name", "Site Name", "Nom du site", "site_name"],
        "year":      ["Année", "Annee", "Year", "année"],
        "month":     ["Mois", "Month", "mois"],
        "load_w":    ["Load", "Load [W]", "Load_W", "load"],
        "source":    ["Source", "Status", "Statut", "source"],
    }

    def _pick(self, columns, key):
        for alias in self.SIMPLE_ALIASES[key]:
            if alias in columns:
                return alias
        return None

    def _upsert_records(self, records, user, source: str = "import"):
        """Upsert a list of {site_id, year, month, load_w} into SiteMonthlyLoad."""
        created = updated = skipped = 0
        errors = []

        # Preload all known site_ids for performance
        all_site_ids = set(Site.objects.values_list("site_id", flat=True))

        with transaction.atomic():
            for rec in records:
                sid = rec["site_id"]
                if sid not in all_site_ids:
                    skipped += 1
                    continue
                try:
                    site = Site.objects.get(site_id=sid)
                    obj, was_created = SiteMonthlyLoad.objects.update_or_create(
                        site=site, year=rec["year"], month=rec["month"],
                        defaults={
                            "load_w":      rec["load_w"],
                            "source": rec.get("source") or source or SiteMonthlyLoad.Source.PREVISIONNEL,
                            "imported_by": user,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                except Exception as e:
                    skipped += 1
                    errors.append({"site_id": sid, "error": str(e)})

        return created, updated, skipped, errors

    def post(self, request, *args, **kwargs):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Aucun fichier fourni."}, status=400)

        file_bytes = f.read()
        fname = f.name.lower()

        # ── Detect format and parse ───────────────────────────────────────────
        # Priorité 1 : paramètre explicite envoyé depuis le frontend
        force_format = request.data.get("format")  # "alignement" | "proposition" | "simple"

        def _detect_by_content(fb: bytes) -> str:
            """Détection par lecture des en-têtes du fichier — plus fiable que le nom."""
            try:
                import io as _io
                import openpyxl as _xl

                wb = _xl.load_workbook(_io.BytesIO(fb), read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(min_row=1, max_row=5, max_col=30, values_only=True))
                wb.close()

                normalized_rows = []
                for row in rows:
                    normalized_rows.append([str(v).strip().lower() for v in row if v is not None])

                # 1) proposition seulement si Average_Load est présent
                if any("average_load" in row for row in normalized_rows):
                    return "proposition"

                # 2) alignement si la première ligne contient des dates
                if rows and any(hasattr(v, "year") for v in rows[0] if v is not None):
                    return "alignement"

                # 3) simple si on voit les colonnes classiques
                flat = {v for row in normalized_rows for v in row}
                has_site = "site_id" in flat or "site id" in flat
                has_year = "année" in flat or "annee" in flat or "year" in flat
                has_month = "mois" in flat or "month" in flat
                has_load = "load" in flat or "load [w]" in flat or "load_w" in flat

                if has_site and has_year and has_month and has_load:
                    return "simple"

            except Exception:
                pass

            return "simple"

        if force_format:
            fmt_to_use = force_format
        else:
            # Détection par nom en premier (cas nominaux), puis par contenu si ambigu
            if "sntl" in fname or "cmst" in fname or "recap" in fname or "alignement" in fname:
                fmt_to_use = "alignement"
            elif "proposition" in fname:
                fmt_to_use = "proposition"
            else:
                fmt_to_use = _detect_by_content(file_bytes)

        # Format 1 : Alignement (Load_SNTL_CMST)
        if fmt_to_use == "alignement":
            try:
                records = _parse_load_alignement(file_bytes)
                detected = "alignement"
                load_source = SiteMonthlyLoad.Source.ALIGNE
            except Exception as e:
                return Response({"detail": f"Erreur lecture fichier aligné : {e}"}, status=400)

        # Format 2 : Proposition mensuelle
        elif fmt_to_use == "proposition":
            try:
                records = _parse_load_proposition(file_bytes)
                detected = "proposition"
                load_source = SiteMonthlyLoad.Source.PREVISIONNEL
            except Exception as e:
                return Response({"detail": f"Erreur lecture fichier proposition : {e}"}, status=400)

        # Format 3 : Simple CSV/Excel avec colonnes nommées
        else:
            try:
                import io
                if fname.endswith(".csv"):
                    df = pd.read_csv(io.BytesIO(file_bytes), dtype=object)
                else:
                    df = pd.read_excel(io.BytesIO(file_bytes), dtype=object)
            except Exception as e:
                return Response({"detail": f"Impossible de lire le fichier : {e}"}, status=400)

            df.columns = [str(c).strip() for c in df.columns]
            c_site   = self._pick(df.columns, "site_id")
            c_year   = self._pick(df.columns, "year")
            c_month  = self._pick(df.columns, "month")
            c_load   = self._pick(df.columns, "load_w")
            c_source = self._pick(df.columns, "source")  # optionnelle

            missing = [k for k, v in [("site_id", c_site), ("year", c_year), ("month", c_month), ("load_w", c_load)] if not v]
            if missing:
                return Response(
                    {
                        "detail": f"Colonnes manquantes : {missing}. Formats acceptés : alignement, proposition, ou CSV simple.",
                        "found": list(df.columns),
                    },
                    status=400,
                )

            records = []
            for _, row in df.iterrows():
                try:
                    sid   = str(row[c_site]).strip()
                    year  = int(float(str(row[c_year])))
                    month = int(float(str(row[c_month])))
                    load  = int(float(str(row[c_load])))

                    if sid and 1 <= month <= 12 and year >= 2020 and load > 0:
                        raw_source = row[c_source] if c_source else None
                        records.append({
                            "site_id": sid,
                            "year": year,
                            "month": month,
                            "load_w": load,
                            "source": _map_load_source(raw_source, default=SiteMonthlyLoad.Source.PREVISIONNEL),
                        })
                except Exception:
                    continue

            detected = "simple"
            load_source = None  # source gérée ligne par ligne

        if not records:
            return Response({"detail": "Aucune ligne valide trouvée dans le fichier."}, status=400)

        # Compute period coverage for info
        if records:
            periods = sorted(set(f"{r['year']}-{r['month']:02d}" for r in records))
            period_info = f"{periods[0]} → {periods[-1]}" if periods else "?"
        else:
            period_info = "?"

        created, updated, skipped, errors = self._upsert_records(records, request.user, source=load_source)

        return Response(
            {
                "message": "Import SiteMonthlyLoad terminé.",
                "format_detecte": detected,
                "source_assignee": load_source,
                "periode_couverte": period_info,
                "total_parsed": len(records),
                "created": created,
                "updated": updated,
                "skipped_sites_inconnus": skipped,
                "errors_sample": errors[:20],
            },
            status=status.HTTP_201_CREATED,
        )


class SiteMonthlyLoadListView(APIView):
    """
    GET /financial/loads/
    ?site=DKR_0026 &year=2026 &month=1 &source=import
    &search=DKR &page=1 &page_size=50
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = SiteMonthlyLoad.objects.select_related("site").order_by("-year", "-month", "site__site_id")

        site_code = request.query_params.get("site")
        year      = request.query_params.get("year")
        month     = request.query_params.get("month")
        source    = request.query_params.get("source")
        search    = request.query_params.get("search")

        if site_code:
            qs = qs.filter(site__site_id=site_code)
        if year:
            qs = qs.filter(year=int(year))
        if month:
            qs = qs.filter(month=int(month))
        if source:
            qs = qs.filter(source=source)
        if search:
            qs = qs.filter(
                Q(site__site_id__icontains=search) | Q(site__name__icontains=search)
            )

        total = qs.count()

        # Pagination
        try:
            page      = max(1, int(request.query_params.get("page", 1)))
            page_size = min(200, max(10, int(request.query_params.get("page_size", 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50

        offset = (page - 1) * page_size
        qs = qs[offset: offset + page_size]

        data = [
            {
                "id":        o.id,
                "site_id":   o.site.site_id,
                "site_name": o.site.name,
                "year":      o.year,
                "month":     o.month,
                "load_w":    o.load_w,
                "source":    o.source,
            }
            for o in qs
        ]
        return Response({
            "count":     total,
            "page":      page,
            "page_size": page_size,
            "pages":     (total + page_size - 1) // page_size if page_size else 1,
            "results":   data,
        })


class SiteMonthlyLoadUpdateView(APIView):
    """
    PATCH /financial/loads/<pk>/
    Body : { "load_w": 5500 }
    Met à jour le load d'un enregistrement existant (loads prévisionnels).
    """
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk: int, *args, **kwargs):
        try:
            obj = SiteMonthlyLoad.objects.select_related("site").get(pk=pk)
        except SiteMonthlyLoad.DoesNotExist:
            return Response({"detail": "Load introuvable."}, status=404)

        load_w = request.data.get("load_w")
        if load_w is None:
            return Response({"detail": "Champ load_w requis."}, status=400)
        try:
            load_w = int(float(str(load_w)))
            if load_w < 0:
                raise ValueError
        except (ValueError, TypeError):
            return Response({"detail": "load_w doit être un entier positif."}, status=400)

        obj.load_w = load_w
        obj.source = SiteMonthlyLoad.Source.MANUAL
        obj.imported_by = request.user
        obj.save(update_fields=["load_w", "source", "imported_by", "imported_at"])

        return Response({
            "id":        obj.id,
            "site_id":   obj.site.site_id,
            "site_name": obj.site.name,
            "year":      obj.year,
            "month":     obj.month,
            "load_w":    obj.load_w,
            "source":    obj.source,
        })


# ─────────────────────────────────────────────────────────────────────────────
# Calcul batch FinancialEvaluation pour un (year, month) donné.
# Extrait de FinancialEvaluateView.post() pour être réutilisable depuis la
# commande de management recompute_financial_evaluations (recalcul automatique
# au démarrage du conteneur, cf. docker-compose.yml).
# ─────────────────────────────────────────────────────────────────────────────

def _run_financial_evaluation(year: int, month: int) -> dict:
    from django.db.models import Min

    links = (
        ContractSiteLink.objects
        .select_related("site")
        .filter(
            site__invoice_payment__iexact="Aktivco",
            site__grid_fee=True,
        )
    )

    contract_ids = list(links.values_list("numero_compte_contrat", flat=True))
    site_pks = list(links.values_list("site_id", flat=True))

    # ─────────────────────────────────────────────────────────────
    # 1) Loads mensuels
    # ─────────────────────────────────────────────────────────────
    load_map: dict[int, int] = dict(
        SiteMonthlyLoad.objects
        .filter(year=year, month=month)
        .values_list("site_id", "load_w")
    )

    # ─────────────────────────────────────────────────────────────
    # 2) Factures PAID / UNPAID / BRUTES
    # ─────────────────────────────────────────────────────────────
    billing_qs = (
        BillingMonthlySynthesis.objects
        .filter(
            numero_compte_contrat__in=contract_ids,
            year=year,
            month=month,
        )
        .exclude(source__payment_status="OUT_OF_SCOPE")
        .values("numero_compte_contrat")
        .annotate(
            montant_paid=Sum(
                "montant_hors_tva",
                filter=Q(source__payment_status="PAID"),
            ),
            montant_unpaid=Sum(
                "montant_hors_tva",
                filter=Q(source__payment_status="UNPAID"),
            ),
            montant_raw=Sum(
                "montant_hors_tva",
                filter=Q(source__payment_status__in=["PAID", "UNPAID"]),
            ),
            paid_count=Count(
                "source_id",
                filter=Q(source__payment_status="PAID"),
                distinct=True,
            ),
            unpaid_count=Count(
                "source_id",
                filter=Q(source__payment_status="UNPAID"),
                distinct=True,
            ),
            raw_count=Count(
                "source_id",
                filter=Q(source__payment_status__in=["PAID", "UNPAID"]),
                distinct=True,
            ),
            first_period_start=Min("period_start"),
            last_period_end=Max("period_end"),
        )
    )

    billing_map = {
        r["numero_compte_contrat"]: r
        for r in billing_qs
    }

    # ─────────────────────────────────────────────────────────────
    # 3) Estimations du mois
    # ─────────────────────────────────────────────────────────────
    estimation_qs = (
        EstimationResult.objects
        .select_related("batch", "site")
        .filter(
            batch__year=year,
            batch__month=month,
            batch__status="DONE",
            site_id__in=site_pks,
        )
    )

    estimation_map = {
        e.site_id: e
        for e in estimation_qs
    }

    evaluations_to_upsert = []

    stats = {
        "processed": 0,
        "ok": 0,
        "nok": 0,
        "no_load": 0,
        "no_fee_rule": 0,
        "no_data": 0,
        "official_paid": 0,
        "provisional_raw_unpaid": 0,
        "provisional_estimation": 0,
        "raw_invoice_available": 0,
        "estimation_available": 0,
        "hors_catalogue": 0,
        "periode_courte": 0,
    }

    # ─────────────────────────────────────────────────────────────
    # 4) Calcul par site
    # ─────────────────────────────────────────────────────────────
    for link in links:
        site = link.site
        acc = link.numero_compte_contrat

        bill = billing_map.get(acc, {})
        estimation = estimation_map.get(site.pk)

        paid_count = int(bill.get("paid_count") or 0)
        unpaid_count = int(bill.get("unpaid_count") or 0)
        raw_count = int(bill.get("raw_count") or 0)

        montant_paid = _d(bill.get("montant_paid"))
        montant_unpaid = _d(bill.get("montant_unpaid"))
        montant_raw = _d(bill.get("montant_raw"))

        montant_estime = None
        estimation_source = None

        if estimation and estimation.montant_estime is not None:
            montant_estime = _d(estimation.montant_estime)
            estimation_source = estimation.source_utilisee
            stats["estimation_available"] += 1

        if raw_count > 0:
            stats["raw_invoice_available"] += 1

        # ── Load ─────────────────────────────────────────────────
        load_w = load_map.get(site.pk) or site.analysis_load
        if not load_w:
            stats["no_load"] += 1
            load_w = None

        # ── Règle redevance ─────────────────────────────────────
        typo = site.billing_typology or site.installed_typology or site.ordered_typology
        conf = _norm_config(site.site_type)

        rule, hors_cat = None, False
        if load_w and typo and conf:
            rule, hors_cat = _get_fee_rule(typo, conf, load_w)

        if not rule:
            stats["no_fee_rule"] += 1
            redevance = None
        else:
            redevance = rule.redevance
            if hors_cat:
                stats["hors_catalogue"] += 1

        # ── Nombre de jours facture ─────────────────────────────
        # Clamp à la fenêtre calendaire du mois : first_period_start/
        # last_period_end viennent d'un Min()/Max() sur toutes les
        # factures rattachées à (year, month) — si plusieurs factures
        # non calendaires (cycles Sénélec non alignés) sont rattachées
        # au même mois, l'écart brut peut dépasser 30j et gonfler la
        # cible de conso (cible_kwh_j × nb_jours) très au-delà du
        # mois réel. Même correctif que SiteMargeDetailView.
        p_start = bill.get("first_period_start")
        p_end = bill.get("last_period_end")

        if p_start and p_end:
            from datetime import date as _date
            _m_last = calendar.monthrange(year, month)[1]
            _month_s = _date(year, month, 1)
            _month_e = _date(year, month, _m_last)
            eff_start = max(p_start, _month_s)
            eff_end = min(p_end, _month_e)
            nb_jours = max(0, (eff_end - eff_start).days + 1)
        else:
            nb_jours = calendar.monthrange(year, month)[1]

        periode_courte = bool(raw_count > 0 and nb_jours < 15)
        if periode_courte:
            stats["periode_courte"] += 1

        # ── Marges comparatives ─────────────────────────────────
        marge_facture_brute = None
        marge_estimee = None

        if redevance is not None and raw_count > 0:
            marge_facture_brute = _q3(_d(redevance) - _d(montant_raw))

        if redevance is not None and montant_estime is not None:
            marge_estimee = _q3(_d(redevance) - _d(montant_estime))

        # ── Choix du montant principal de calcul ─────────────────
        montant_base_calcul = None
        calculation_source = FinancialEvaluation.CalculationSource.NO_DATA
        is_provisional = True
        warning = None

        if paid_count > 0:
            montant_base_calcul = montant_paid
            calculation_source = FinancialEvaluation.CalculationSource.PAID_INVOICE
            is_provisional = False
            stats["official_paid"] += 1

            if unpaid_count > 0:
                warning = (
                    "Certaines factures du mois restent non payées. "
                    "Le calcul officiel utilise uniquement les factures payées."
                )

        elif unpaid_count > 0:
            montant_base_calcul = montant_unpaid
            calculation_source = FinancialEvaluation.CalculationSource.RAW_UNPAID_INVOICE
            is_provisional = True
            stats["provisional_raw_unpaid"] += 1
            warning = (
                "Factures du mois non payées : calcul provisoire basé sur les factures brutes. "
                "Une comparaison avec l'estimation est disponible."
            )

        elif montant_estime is not None:
            montant_base_calcul = montant_estime
            calculation_source = FinancialEvaluation.CalculationSource.ESTIMATION_ONLY
            is_provisional = True
            stats["provisional_estimation"] += 1
            warning = (
                "Aucune facture disponible pour ce mois : calcul provisoire basé sur l'estimation."
            )

        else:
            stats["no_data"] += 1
            warning = (
                "Aucune facture payée, aucune facture brute et aucune estimation disponible."
            )

        # ── Calcul marge principale ──────────────────────────────
        if periode_courte and montant_base_calcul is not None and calculation_source in (
            FinancialEvaluation.CalculationSource.PAID_INVOICE,
            FinancialEvaluation.CalculationSource.RAW_UNPAID_INVOICE,
        ):
            marge = DEC0
            statut = FinancialEvaluation.MargeStatut.OK

        elif redevance is not None and montant_base_calcul is not None:
            marge = _q3(_d(redevance) - _d(montant_base_calcul))
            statut = (
                FinancialEvaluation.MargeStatut.OK
                if marge >= DEC0
                else FinancialEvaluation.MargeStatut.NOK
            )
        else:
            marge = None
            statut = None

        evaluations_to_upsert.append({
            "site": site,
            "year": year,
            "month": month,

            "load_w": load_w,
            "typology": typo,
            "configuration": conf,

            "fee_rule": rule,
            "fee_rule_load_w": rule.load_w if rule else None,
            "redevance": redevance,
            "hors_catalogue": hors_cat,

            # Compatibilité ancienne UI :
            # montant_htva_reel = montant réellement utilisé pour calculer la marge
            "montant_htva_reel": montant_base_calcul,

            # Nouvelle traçabilité
            "calculation_source": calculation_source,
            "is_provisional": is_provisional,
            "calculation_warning": warning,

            "montant_base_calcul": montant_base_calcul,
            "montant_factures_payees": montant_paid if paid_count > 0 else None,
            "montant_factures_brutes": montant_raw if raw_count > 0 else None,
            "montant_factures_non_payees": montant_unpaid if unpaid_count > 0 else None,
            "montant_estime": montant_estime,

            "marge_facture_brute": marge_facture_brute,
            "marge_estimee": marge_estimee,

            "paid_invoice_count": paid_count,
            "unpaid_invoice_count": unpaid_count,
            "raw_invoice_count": raw_count,

            "has_invoice_for_month": raw_count > 0,
            "has_unpaid_invoice": unpaid_count > 0,
            "estimation_source": estimation_source,

            "marge": marge,
            "marge_statut": statut,
            "periode_courte": periode_courte,
            "nb_jours_factures": nb_jours,
        })

        stats["processed"] += 1

        if statut == FinancialEvaluation.MargeStatut.OK:
            stats["ok"] += 1
        elif statut == FinancialEvaluation.MargeStatut.NOK:
            stats["nok"] += 1

    # ─────────────────────────────────────────────────────────────
    # 5) Upsert + recalcul récurrence
    # ─────────────────────────────────────────────────────────────
    with transaction.atomic():
        for ev in evaluations_to_upsert:
            site = ev.pop("site")

            obj, _ = FinancialEvaluation.objects.update_or_create(
                site=site,
                year=year,
                month=month,
                defaults=ev,
            )

            nb_nok, rec_type = _compute_recurrence(site.pk, year, month)

            if obj.recurrence_mois_nok != nb_nok or obj.recurrence_type != rec_type:
                obj.recurrence_mois_nok = nb_nok
                obj.recurrence_type = rec_type
                obj.save(update_fields=["recurrence_mois_nok", "recurrence_type"])

    return {
        "message": f"Évaluation financière {year}-{month:02d} terminée.",
        **stats,
    }


class FinancialEvaluateView(APIView):
    """
    POST /financial/evaluate/
    Body JSON : {"year": 2026, "month": 1}

    Règle métier :
      1) Si facture payée existe        → calcul officiel sur PAID
      2) Sinon si facture brute existe  → calcul provisoire sur UNPAID + comparaison estimation
      3) Sinon si estimation existe     → calcul provisoire sur estimation
      4) Sinon                          → non calculable
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        year = request.data.get("year")
        month = request.data.get("month")

        if not year or not month:
            return Response({"detail": "year et month requis."}, status=400)

        try:
            year, month = int(year), int(month)
            if not (1 <= month <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({"detail": "year/month invalides."}, status=400)

        result = _run_financial_evaluation(year, month)
        return Response(result, status=status.HTTP_200_OK)

# ─────────────────────────────────────────────────────────────────────────────
# VIEW 4b : Agrégats stats (pour les cards) — séparé de la liste paginée
# ─────────────────────────────────────────────────────────────────────────────


class FinancialEvaluationStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = FinancialEvaluation.objects.all()
        qs = _apply_period_filters(qs, request)

        agg = qs.aggregate(
            total_redevance=Sum("redevance"),
            total_facture=Sum("montant_htva_reel"),
            total_marge=Sum("marge"),
            count_ok=Count("id", filter=Q(marge_statut="OK")),
            count_nok=Count("id", filter=Q(marge_statut="NOK")),
            count_pc=Count("id", filter=Q(periode_courte=True)),
            count_hc=Count("id", filter=Q(hors_catalogue=True)),
            count_total=Count("id"),
            total_base_calcul=Sum("montant_base_calcul"),
            total_factures_payees=Sum("montant_factures_payees"),
            total_factures_brutes=Sum("montant_factures_brutes"),
            total_factures_non_payees=Sum("montant_factures_non_payees"),
            total_estime=Sum("montant_estime"),
            count_provisional=Count("id", filter=Q(is_provisional=True)),
            count_official=Count("id", filter=Q(is_provisional=False)),
            count_paid_source=Count("id", filter=Q(calculation_source="PAID_INVOICE")),
            count_raw_source=Count("id", filter=Q(calculation_source="RAW_UNPAID_INVOICE")),
            count_estimation_source=Count("id", filter=Q(calculation_source="ESTIMATION_ONLY")),
        )

        return Response({
            "total_redevance": str(_q3(_d(agg["total_redevance"]))),
            "total_facture": str(_q3(_d(agg["total_facture"]))),
            "total_marge": str(_q3(_d(agg["total_marge"]))),
            "count_ok": agg["count_ok"] or 0,
            "count_nok": agg["count_nok"] or 0,
            "count_pc": agg["count_pc"] or 0,
            "count_hc": agg["count_hc"] or 0,
            "count_total": agg["count_total"] or 0,

            "total_base_calcul": str(_q3(_d(agg["total_base_calcul"]))),
            "total_factures_payees": str(_q3(_d(agg["total_factures_payees"]))),
            "total_factures_brutes": str(_q3(_d(agg["total_factures_brutes"]))),
            "total_factures_non_payees": str(_q3(_d(agg["total_factures_non_payees"]))),
            "total_estime": str(_q3(_d(agg["total_estime"]))),

            "count_provisional": agg["count_provisional"] or 0,
            "count_official": agg["count_official"] or 0,
            "count_paid_source": agg["count_paid_source"] or 0,
            "count_raw_source": agg["count_raw_source"] or 0,
            "count_estimation_source": agg["count_estimation_source"] or 0,
        })

# ─────────────────────────────────────────────────────────────────────────────
# VIEW 4 : Liste des évaluations
# ─────────────────────────────────────────────────────────────────────────────

class FinancialEvaluationListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = FinancialEvaluation.objects.select_related("site", "fee_rule").order_by(
            "-year", "-month", "marge"
        )
        qs = _apply_period_filters(qs, request)

        statut = request.query_params.get("statut")
        hc = request.query_params.get("hors_catalogue")
        site = request.query_params.get("site")
        rec = request.query_params.get("recurrence_type")
        search = request.query_params.get("search")
        typo = request.query_params.get("typology")
        zone = request.query_params.get("zone")
        export = request.query_params.get("export")

        if statut:
            qs = qs.filter(marge_statut=statut.upper())
        if hc and hc.lower() == "true":
            qs = qs.filter(hors_catalogue=True)
        if site:
            qs = qs.filter(site__site_id=site)
        if rec:
            qs = qs.filter(recurrence_type=rec.lower())
        if search:
            qs = qs.filter(
                Q(site__site_id__icontains=search) | Q(site__name__icontains=search)
            )
        if typo:
            qs = qs.filter(typology__icontains=typo)
        if zone:
            qs = qs.filter(site__zone=zone.upper())

        if export == "csv":
            import csv
            import io as _io
            from django.http import HttpResponse

            year = request.query_params.get("year", "all")
            month = request.query_params.get("month")
            month_start = request.query_params.get("month_start")
            month_end = request.query_params.get("month_end")

            buf = _io.StringIO()
            w = csv.writer(buf)
            w.writerow([
                "Site ID", "Site Name", "Zone", "Année", "Mois",
                "Typologie", "Configuration", "Load (W)",
                "Redevance (FCFA)", "Montant HT (FCFA)", "Marge (FCFA)", "Statut",
                "Hors Catalogue", "Période courte", "Nb jours", "Récurrence",
            ])
            for o in qs:
                w.writerow([
                    o.site.site_id,
                    o.site.name or "",
                    o.site.zone or "",
                    o.year,
                    o.month,
                    o.typology or "",
                    o.configuration or "",
                    o.load_w or "",
                    o.redevance or "",
                    o.montant_htva_reel or "",
                    o.marge or "",
                    o.marge_statut or "",
                    "OUI" if o.hors_catalogue else "",
                    "OUI" if o.periode_courte else "",
                    o.nb_jours_factures or "",
                    o.recurrence_type or "",
                ])

            if month:
                suffix = f"{year}_{month}"
            elif month_start or month_end:
                suffix = f"{year}_{month_start or '1'}-{month_end or '12'}"
            else:
                suffix = str(year)

            resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8-sig")
            resp["Content-Disposition"] = f'attachment; filename="evaluations_{suffix}.csv"'
            return resp

        total = qs.count()
        try:
            page = max(1, int(request.query_params.get("page", 1)))
            page_size = min(500, max(25, int(request.query_params.get("page_size", 100))))
        except (ValueError, TypeError):
            page, page_size = 1, 100

        offset = (page - 1) * page_size
        qs_page = qs[offset: offset + page_size]

        data = [
            {
                "id": o.id,
                "site_id": o.site.site_id,
                "site_name": o.site.name,
                "zone": o.site.zone,
                "year": o.year,
                "month": o.month,
                "load_w": o.load_w,
                "typology": o.typology,
                "configuration": o.configuration,

                "redevance": str(o.redevance or ""),
                "montant_htva": str(o.montant_htva_reel or ""),
                "marge": str(o.marge or ""),
                "marge_statut": o.marge_statut,

                "calculation_source": o.calculation_source,
                "is_provisional": o.is_provisional,
                "calculation_warning": o.calculation_warning,

                "montant_base_calcul": str(o.montant_base_calcul or ""),
                "montant_factures_payees": str(o.montant_factures_payees or ""),
                "montant_factures_brutes": str(o.montant_factures_brutes or ""),
                "montant_factures_non_payees": str(o.montant_factures_non_payees or ""),
                "montant_estime": str(o.montant_estime or ""),

                "marge_facture_brute": str(o.marge_facture_brute or ""),
                "marge_estimee": str(o.marge_estimee or ""),

                "paid_invoice_count": o.paid_invoice_count,
                "unpaid_invoice_count": o.unpaid_invoice_count,
                "raw_invoice_count": o.raw_invoice_count,

                "has_invoice_for_month": o.has_invoice_for_month,
                "has_unpaid_invoice": o.has_unpaid_invoice,
                "estimation_source": o.estimation_source,

                "hors_catalogue": o.hors_catalogue,
                "fee_rule_load_w": o.fee_rule_load_w,
                "periode_courte": o.periode_courte,
                "nb_jours_factures": o.nb_jours_factures,
                "recurrence_mois_nok": o.recurrence_mois_nok,
                "recurrence_type": o.recurrence_type,
            }
            for o in qs_page
        ]
        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size,
            "results": data,
        })

# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARDS
# ─────────────────────────────────────────────────────────────────────────────

class DashboardFacturesRedevancesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = FinancialEvaluation.objects.all()
        qs = _apply_period_filters(qs, request)

        rows = (
            qs.values("year", "month")
            .annotate(
                total_redevance=Sum("redevance"),
                total_facture=Sum("montant_htva_reel"),
                total_marge=Sum("marge"),
                sites_ok=Count("id", filter=Q(marge_statut="OK")),
                sites_nok=Count("id", filter=Q(marge_statut="NOK")),
            )
            .order_by("year", "month")
        )

        return Response([
            {
                "period": f"{r['year']}-{r['month']:02d}",
                "total_redevance": str(_q3(_d(r["total_redevance"]))),
                "total_facture": str(_q3(_d(r["total_facture"]))),
                "total_marge": str(_q3(_d(r["total_marge"]))),
                "sites_ok": r["sites_ok"],
                "sites_nok": r["sites_nok"],
            }
            for r in rows
        ])


class DashboardMargeParSiteView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = FinancialEvaluation.objects.select_related("site").filter(marge__isnull=False)
        qs = _apply_period_filters(qs, request)
        limit = int(request.query_params.get("limit", 50))

        rows = (
            qs.values("site__site_id", "site__name")
            .annotate(
                marge_moyenne=Avg("marge"),
                nb_mois=Count("id"),
                nb_nok=Count("id", filter=Q(marge_statut="NOK")),
            )
            .order_by("marge_moyenne")[:limit]
        )

        return Response([
            {
                "site_id": r["site__site_id"],
                "site_name": r["site__name"],
                "marge_moyenne": str(_q3(_d(r["marge_moyenne"]))),
                "nb_mois": r["nb_mois"],
                "nb_nok": r["nb_nok"],
            }
            for r in rows
        ])


class DashboardSitesRecurrentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rec_type = request.query_params.get("recurrence_type")

        qs = FinancialEvaluation.objects.select_related("site").exclude(recurrence_type__isnull=True)
        qs = _apply_period_filters(qs, request)
        if rec_type:
            qs = qs.filter(recurrence_type=rec_type.lower())

        rows = (
            qs.values("site__site_id", "site__name", "recurrence_type")
            .annotate(
                marge_moyenne=Avg("marge"),
                recurrence_mois_max=Max("recurrence_mois_nok"),
            )
            .order_by("marge_moyenne")
        )

        return Response([
            {
                "site_id": r["site__site_id"],
                "site_name": r["site__name"],
                "recurrence_type": r["recurrence_type"],
                "mois_nok": r["recurrence_mois_max"] or 0,
                "marge_moyenne": str(_q3(_d(r["marge_moyenne"]))),
            }
            for r in rows
        ])




class SiteMargeDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, site_id: str):
        year        = int(request.query_params.get("year", timezone.now().year))
        month       = request.query_params.get("month")
        month_start = int(request.query_params.get("month_start", month or 1))
        month_end   = int(request.query_params.get("month_end",   month or 12))

        if month_start > month_end:
            month_start, month_end = month_end, month_start

        site = get_object_or_404(Site, site_id=site_id)

        evals = list(
            FinancialEvaluation.objects.filter(
                site=site,
                year=year,
                month__gte=month_start,
                month__lte=month_end,
            ).order_by("month")
        )

        history = [
            {
                "period":             f"{ev.year}-{ev.month:02d}",
                "year":               ev.year,
                "month":              ev.month,
                "redevance":          str(ev.redevance)         if ev.redevance         is not None else None,
                "montant_htva":       str(ev.montant_htva_reel) if ev.montant_htva_reel is not None else None,
                "marge":              str(ev.marge)             if ev.marge             is not None else None,
                "marge_statut":       ev.marge_statut,
                "load_w":             ev.load_w,
                "hors_catalogue":     ev.hors_catalogue,
                "periode_courte":     ev.periode_courte,
                "nb_jours":           ev.nb_jours_factures,
                "recurrence_type":    ev.recurrence_type,
                "recurrence_mois_nok": ev.recurrence_mois_nok,
            }
            for ev in evals
        ]

        current_eval = evals[-1] if evals else None

        link        = ContractSiteLink.objects.filter(site=site).first()
        contract_id = link.numero_compte_contrat if link else None

        # ── Billing rows (ContractMonth) ──────────────────────────────────────
        billing_rows   = []
        total_ht       = DEC0
        total_cosphi   = DEC0
        total_penalite = DEC0

        if contract_id:
            cms = ContractMonth.objects.filter(
                numero_compte_contrat=contract_id,
                year=year,
                month__gte=month_start,
                month__lte=month_end,
            ).order_by("month")

            for cm in cms:
                # APRÈS — clamp à la fenêtre calendaire du mois
                if cm.first_period_start and cm.last_period_end:
                    from datetime import date as _date
                    _m_last   = calendar.monthrange(cm.year, cm.month)[1]
                    _month_s  = _date(cm.year, cm.month, 1)
                    _month_e  = _date(cm.year, cm.month, _m_last)
                    eff_start = max(cm.first_period_start, _month_s)
                    eff_end   = min(cm.last_period_end,   _month_e)
                    nb_jours  = max(0, (eff_end - eff_start).days + 1)
                else:
                    nb_jours = None
                total_ht       += _d(cm.montant_hors_tva)
                total_cosphi   += _d(cm.montant_cosinus_phi)
                total_penalite += _d(cm.penalite_abonnement_calculee)

                billing_rows.append({
                    "period":             f"{cm.year}-{cm.month:02d}",
                    "year":               cm.year,
                    "month":              cm.month,
                    "montant_hors_tva":   str(cm.montant_hors_tva)              if cm.montant_hors_tva              is not None else "0",
                    "energie":            str(cm.energie_calculee)               if cm.energie_calculee              is not None else "0",
                    "abonnement":         str(cm.abonnement_calcule)             if cm.abonnement_calcule            is not None else "0",
                    "montant_cosinus_phi":str(cm.montant_cosinus_phi)            if cm.montant_cosinus_phi           is not None else "0",
                    "penalite_abonnement":str(cm.penalite_abonnement_calculee)   if cm.penalite_abonnement_calculee  is not None else "0",
                    "puissance_souscrite":str(link.puissance_contractuelle)      if link and link.puissance_contractuelle is not None else None,
                    "puissance_max":      None,
                    "cosphi":             str(cm.valeur_cosinus_phi)             if cm.valeur_cosinus_phi            is not None else None,
                    "nb_jours":           nb_jours,
                })
            conso_kwh_by_month: dict[int, Decimal] = {}
            if contract_id:
                ms_conso = (
                    BillingMonthlySynthesis.objects
                    .filter(
                        numero_compte_contrat=contract_id,
                        year=year,
                        month__gte=month_start,
                        month__lte=month_end,
                        source__payment_status__in=["PAID"],
                    )
                    .values("month")
                    .annotate(total_conso=Sum("conso"))
                )
                for ms in ms_conso:
                    if ms["total_conso"] is not None:
                        conso_kwh_by_month[ms["month"]] = _d(ms["total_conso"])

            for row in billing_rows:
                conso_kwh = conso_kwh_by_month.get(row["month"])
                row["conso_kwh"] = str(conso_kwh) if conso_kwh else None

        # ── Étape A : conso_target depuis le catalogue (cible_kwh_j × nb_jours) ──
        eval_by_month = {ev.month: ev for ev in evals}

        for row in billing_rows:
            ev   = eval_by_month.get(row["month"])
            rule = ev.fee_rule if ev else None
            nb_j = row.get("nb_jours") or 30

            if rule and rule.cible_kwh_j:
                row["conso_target_kwh"] = str(
                    (Decimal(str(rule.cible_kwh_j)) * Decimal(str(nb_j)))
                    .quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
                )
                row["cible_kwh_j"] = str(rule.cible_kwh_j)
                row["cible_kwh"]   = str(rule.cible_kwh) if rule.cible_kwh else None
            elif rule and rule.cible_kwh:
                row["conso_target_kwh"] = str(rule.cible_kwh)
                row["cible_kwh_j"]      = None
                row["cible_kwh"]        = str(rule.cible_kwh)
            else:
                row["conso_target_kwh"] = None
                row["cible_kwh_j"]      = None
                row["cible_kwh"]        = None

        
        # Étape A2 : target solaire
        for row in billing_rows:
            ev   = eval_by_month.get(row["month"])
            nb_j = row.get("nb_jours") or 30

            # Fallback : si l'éval n'a pas la typo/load, utiliser les données site directement
            typo_for_solar = (
                (ev.typology if ev else None)
                or site.billing_typology
                or site.installed_typology
                or site.ordered_typology
            )
            load_for_solar = (ev.load_w if ev else None) or site.analysis_load

            solar_tgt = _compute_solar_target(typo_for_solar, load_for_solar, nb_j)
            row["solar_target_kwh"] = str(solar_tgt) if solar_tgt is not None else None

        # ── Étape B : conso FMS/ACM + Solar via FinancialConsoService ────────
        # Stratégie : Local (CertificationResult) → Remote eFMS → Solar SQL2
        

        conso_svc = FinancialConsoService(
            site_id=site_id,
            year=year,
            month_start=month_start,
            month_end=month_end,
            billing_rows=billing_rows,
        )
        conso_map = conso_svc.fetch()   # dict[int, ConsoMonthData]

        for row in billing_rows:
            data: ConsoMonthData = conso_map.get(
                row["month"],
                ConsoMonthData(month=row["month"], period_label=row["period"])
            )
            row["conso_fms_kwh"]  = str(data.conso_fms_kwh)  if data.conso_fms_kwh  else None
            row["conso_acm_kwh"]  = str(data.conso_acm_kwh)  if data.conso_acm_kwh  else None
            row["ratio_fms"]      = str(data.ratio_fms)       if data.ratio_fms      else None
            row["fms_available"]  = data.fms_available
            row["acm_available"]  = data.acm_available
            row["fms_source"]     = data.fms_source   # "local" | "remote:exact" | "remote:extrapol" | "none"
            row["acm_source"]     = data.acm_source
            row["solar_kwh"]      = data.solar_kwh
            row["unavail_hours"]  = data.unavail_hours

        # ── Certification (résumé + détail — reste intact) ────────────────────
        cert_rows = []
        cert_status_counts = {
            "CERTIFIED_FMS":      0,
            "CERTIFIED_SENELEC":  0,
            "NEEDS_REVIEW":       0,
            "FMS_UNAVAILABLE":    0,
            "UNKNOWN_CONTRACT":   0,
        }

        cert_qs = (
            CertificationResult.objects
            .filter(
                site=site,
                cert_batch__echeance_year=year,
                cert_batch__echeance_month__gte=month_start,
                cert_batch__echeance_month__lte=month_end,
            )
            .select_related("cert_batch", "invoice")
            .order_by("cert_batch__echeance_month", "computed_at")
        )

        for c in cert_qs:
            cert_status_counts[c.status] = cert_status_counts.get(c.status, 0) + 1
            cert_rows.append({
                "period":             f"{c.cert_batch.echeance_year}-{c.cert_batch.echeance_month:02d}",
                "status":             c.status,
                "certified_by_rule":  c.certified_by_rule,
                "ratio_fms":          str(c.ratio_fms_periode)      if c.ratio_fms_periode      is not None else None,
                "ratio_fms_30j":      str(c.ratio_fms_30j)          if c.ratio_fms_30j          is not None else None,
                "energie_fms":        str(c.conso_fms_periode)       if c.conso_fms_periode      is not None else None,
                "energie_fms_30j":    str(c.conso_fms_30j)           if c.conso_fms_30j          is not None else None,
                "energie_senelec":    str(c.conso_facturee_periode)  if c.conso_facturee_periode is not None else None,
                "energie_senelec_30j":str(c.conso_facturee_30j)      if c.conso_facturee_30j    is not None else None,
                "montant_coherent":   c.montant_coherent,
                "variation_montant":  str(c.variation_montant_pct)   if c.variation_montant_pct  is not None else None,
                "fms_available":      c.fms_available,
                "acm_available":      c.acm_available,
                "fms_error":          c.fms_error,
                "acm_error":          c.acm_error,
            })

        # ── Agrégats ──────────────────────────────────────────────────────────
        total_marge    = sum((_d(ev.marge) for ev in evals), DEC0)
        count_nok      = sum(1 for ev in evals if ev.marge_statut == FinancialEvaluation.MargeStatut.NOK)
        count_ok       = sum(1 for ev in evals if ev.marge_statut == FinancialEvaluation.MargeStatut.OK)
        hc_count       = sum(1 for ev in evals if ev.hors_catalogue)
        pc_count       = sum(1 for ev in evals if ev.periode_courte)
        max_recurrence = max((ev.recurrence_mois_nok or 0 for ev in evals), default=0)

        # ── Diagnostics ───────────────────────────────────────────────────────
        diagnostics = []

        if total_marge < 0:
            diagnostics.append({
                "type":     "MARGE",
                "severity": "CRITICAL" if total_marge <= Decimal("-500000") else "HIGH",
                "message":  f"Marge cumulée négative : {int(total_marge):,} FCFA",
                "detail":   "Le total facturé HT dépasse la redevance attendue sur la période.",
            })

        if total_cosphi > 0:
            diagnostics.append({
                "type":     "COSPHI",
                "severity": "HIGH" if total_cosphi > Decimal("500000") else "MEDIUM",
                "message":  f"Pénalité cos φ cumulée : {int(total_cosphi):,} FCFA",
                "detail":   "Le facteur de puissance dégrade la performance financière du site.",
            })

        if total_penalite > 0:
            diagnostics.append({
                "type":     "PUISSANCE",
                "severity": "HIGH" if total_penalite > Decimal("300000") else "MEDIUM",
                "message":  f"Pénalité dépassement puissance : {int(total_penalite):,} FCFA",
                "detail":   "La puissance contractuelle semble inadaptée ou régulièrement dépassée.",
            })

        if hc_count > 0:
            diagnostics.append({
                "type":     "CATALOGUE",
                "severity": "MEDIUM",
                "message":  f"Hors catalogue sur {hc_count} mois",
                "detail":   "Le load réel n'entre pas exactement dans la grille de redevance et utilise un fallback.",
            })

        if cert_status_counts.get("NEEDS_REVIEW", 0) > 0 or cert_status_counts.get("FMS_UNAVAILABLE", 0) > 0:
            diagnostics.append({
                "type":     "CERTIFICATION",
                "severity": "HIGH",
                "message":  (
                    f"{cert_status_counts.get('NEEDS_REVIEW', 0)} mois à revoir · "
                    f"{cert_status_counts.get('FMS_UNAVAILABLE', 0)} mois FMS indisponible"
                ),
                "detail":   "Les données de certification montrent des incohérences ou une indisponibilité FMS.",
            })

        if max_recurrence >= 6:
            diagnostics.append({
                "type":     "RECURRENCE",
                "severity": "CRITICAL",
                "message":  f"Récurrence critique : {max_recurrence} mois NOK",
                "detail":   "Le site présente une dérive structurelle sur plusieurs mois.",
            })
        elif max_recurrence >= 3:
            diagnostics.append({
                "type":     "RECURRENCE",
                "severity": "HIGH",
                "message":  f"Récurrence light : {max_recurrence} mois NOK",
                "detail":   "Une tendance négative commence à s'installer et doit être suivie.",
            })

        # ── Current snapshot ──────────────────────────────────────────────────
        current = None
        if current_eval:
            current = {
                "period":           f"{current_eval.year}-{current_eval.month:02d}",
                "typology":         current_eval.typology,
                "configuration":    current_eval.configuration,
                "load_w":           current_eval.load_w,
                "redevance":        str(current_eval.redevance)         if current_eval.redevance         is not None else None,
                "montant_htva":     str(current_eval.montant_htva_reel) if current_eval.montant_htva_reel is not None else None,
                "marge":            str(current_eval.marge)             if current_eval.marge             is not None else None,
                "marge_statut":     current_eval.marge_statut,
                "hors_catalogue":   current_eval.hors_catalogue,
                "recurrence_type":  current_eval.recurrence_type,
                "recurrence_mois":  current_eval.recurrence_mois_nok,
                "periode_courte":   current_eval.periode_courte,
                "nb_jours_factures":current_eval.nb_jours_factures,
            }

        # ── Response ──────────────────────────────────────────────────────────
        return Response({
            "site": {
                "site_id":        site.site_id,
                "name":           site.name,
                "zone":           site.zone,
                "typology":       site.billing_typology or site.installed_typology or site.ordered_typology,
                "configuration":  site.configuration or site.site_type,
                "grid_fee":       site.grid_fee,
                "invoice_payment":site.invoice_payment,
            },
            "period": {
                "year":        year,
                "month_start": month_start,
                "month_end":   month_end,
            },
            "contract_id": contract_id,
            "current":     current,
            "summary": {
                "total_marge":           str(total_marge),
                "count_ok":              count_ok,
                "count_nok":             count_nok,
                "count_hors_catalogue":  hc_count,
                "count_periode_courte":  pc_count,
                "billing_total_ht":      str(total_ht),
                "billing_total_cosphi":  str(total_cosphi),
                "billing_total_penalite":str(total_penalite),
                "cert_count":            len(cert_rows),
                "cert_status_counts":    cert_status_counts,
            },
            "history": history,
            "billing": {
                "summary": {
                    "total_ht":       str(total_ht),
                    "total_cosphi":   str(total_cosphi),
                    "total_penalite": str(total_penalite),
                },
                "rows": billing_rows,
            },
            "certification": {
                "summary": cert_status_counts,
                "rows":    cert_rows,
            },
            # ✅ NOUVEAU — comparaison consommations multi-sources
            "conso_comparison": {
                "rows": [
                    {
                        "period":           row["period"],
                        "month":            row["month"],
                        "nb_jours":         row.get("nb_jours"),
                        # Sénélec facturée
                        "conso_facturee":   row.get("conso_kwh"),
                        "montant_htva":     row.get("montant_hors_tva"),    # ← NEW
                        # FMS / ACM
                        "conso_fms":        row.get("conso_fms_kwh"),
                        "conso_acm":        row.get("conso_acm_kwh"),
                        "fms_available":    row.get("fms_available", False),
                        "acm_available":    row.get("acm_available", False),
                        "fms_source":       row.get("fms_source", "none"),
                        "acm_source":       row.get("acm_source", "none"),
                        "ratio_fms":        row.get("ratio_fms"),
                        # Cible catalogue
                        "conso_target":     row.get("conso_target_kwh"),
                        "cible_kwh_j":      row.get("cible_kwh_j"),
                        "cible_kwh":        row.get("cible_kwh"),
                        # Solar
                        "solar_kwh":           row.get("solar_kwh"),
                        "solar_target_kwh":    row.get("solar_target_kwh"),  # ← NEW
                        "unavail_hours":       row.get("unavail_hours"),
                    }
                    for row in billing_rows
                ]
            },
            "diagnostics": diagnostics,
        })








# ─────────────────────────────────────────────────────────────────────────────
# PATCH 2/2 — Remplace SuiviConsoView.get()
# Supporte year_start / year_end + month_start / month_end
# ─────────────────────────────────────────────────────────────────────────────

class SuiviConsoView(APIView):
    """
    GET /financial/suivi-conso/
    Agrège les consommations par site × mois sur une période.

    Paramètres acceptés (les deux syntaxes sont supportées) :
      Nouveau : ?year_start=2024&month_start=10&year_end=2025&month_end=3
      Ancien  : ?year=2025&month_start=1&month_end=12   (rétrocompat)

    Filtres additionnels : zone, search, typology, statut
    Export               : &export=csv
    """
    permission_classes = [IsAuthenticated]

    # ── Helpers internes ──────────────────────────────────────────────────────

    @staticmethod
    def _parse_period(request):
        """
        Retourne (year_start, month_start, year_end, month_end) normalisés.
        Compatible avec les deux syntaxes frontend.
        """
        now = timezone.now()

        ys_raw = request.query_params.get("year_start")
        ye_raw = request.query_params.get("year_end")
        ms_raw = request.query_params.get("month_start")
        me_raw = request.query_params.get("month_end")
        y_raw  = request.query_params.get("year")

        if ys_raw and ye_raw:
            ys = int(ys_raw)
            ye = int(ye_raw)
            ms = int(ms_raw) if ms_raw else 1
            me = int(me_raw) if me_raw else 12
        else:
            # Rétrocompat : ?year=2025&month_start=1&month_end=12
            ys = ye = int(y_raw) if y_raw else now.year
            ms = int(ms_raw) if ms_raw else 1
            me = int(me_raw) if me_raw else 12

        # Normalise si bornes inversées
        if (ys, ms) > (ye, me):
            ys, ms, ye, me = ye, me, ys, ms

        return ys, ms, ye, me

    @staticmethod
    def _period_q(ys: int, ms: int, ye: int, me: int,
                  year_field: str = "year", month_field: str = "month") -> Q:
        """Construit le Q() inter-années pour BillingMonthlySynthesis / FinancialEvaluation."""
        if ys == ye:
            return Q(**{
                year_field:                ys,
                f"{month_field}__gte": ms,
                f"{month_field}__lte": me,
            })
        return (
            Q(**{year_field: ys, f"{month_field}__gte": ms})
            | Q(**{f"{year_field}__gt": ys, f"{year_field}__lt": ye})
            | Q(**{year_field: ye, f"{month_field}__lte": me})
        )


    # NOTE : l'ancienne copie locale de fetch_bulk_for_list (requêtes Snowflake live,
    # dupliquée avec financial/services/conso_service.py) a été supprimée — cette vue
    # utilise désormais exclusivement FinancialConsoService.fetch_bulk_for_list, qui
    # lit FinancialConsoMonthly (Postgres) au lieu d'interroger Snowflake en direct.

    # ── Handler principal ─────────────────────────────────────────────────────

    def get(self, request):
        ys, ms, ye, me = self._parse_period(request)
        period_q = self._period_q(ys, ms, ye, me)

        zone = request.query_params.get("zone")
        search = request.query_params.get("search")
        typo = request.query_params.get("typology")
        statut = request.query_params.get("statut")
        export = request.query_params.get("export")

        def fmt(v):
            if v is None:
                return None
            return str(_q3(_d(v)))

        def period_key(y: int, m: int) -> int:
            return y * 100 + m

        # ─────────────────────────────────────────────────────────────
        # 1) Conso facturée Sénélec : BillingMonthlySynthesis
        # ─────────────────────────────────────────────────────────────
        billing_qs = (
            BillingMonthlySynthesis.objects
            .select_related("source__site")
            .filter(
                period_q,
                source__site__isnull=False,
                source__site__invoice_payment__iexact="Aktivco",
                source__site__grid_fee=True,
                source__payment_status__in=["PAID", "UNPAID"],
            )
        )

        if zone:
            billing_qs = billing_qs.filter(source__site__zone=zone.upper())

        if search:
            billing_qs = billing_qs.filter(
                Q(source__site__site_id__icontains=search)
                | Q(source__site__name__icontains=search)
            )

        if typo:
            billing_qs = billing_qs.filter(
                Q(source__site__billing_typology__icontains=typo)
                | Q(source__site__installed_typology__icontains=typo)
                | Q(source__site__ordered_typology__icontains=typo)
            )

        billing_rows = (
            billing_qs
            .values(
                "source__site__site_id",
                "source__site__name",
                "source__site__zone",
                "year",
                "month",
            )
            .annotate(
                conso_kwh=Sum("conso"),
                montant_ht=Sum("montant_hors_tva"),
                montant_energie=Sum("montant_energie"),
            )
            .order_by("source__site__site_id", "year", "month")
        )

        billing_index = {}
        site_meta = {}
        all_keys = set()

        for r in billing_rows:
            sid = r["source__site__site_id"]
            y = r["year"]
            m = r["month"]
            key = (sid, y, m)

            billing_index[key] = r
            all_keys.add(key)

            site_meta[sid] = {
                "site_id": sid,
                "site_name": r["source__site__name"],
                "zone": r["source__site__zone"],
            }

        # ─────────────────────────────────────────────────────────────
        # 2) Conso estimée : EstimationResult
        #    Important : cette source permet d’afficher les mois sans facture
        # ─────────────────────────────────────────────────────────────
        estimation_period_q = self._period_q(
            ys, ms, ye, me,
            year_field="batch__year",
            month_field="batch__month",
        )

        estimation_qs = (
            EstimationResult.objects
            .select_related("batch", "site")
            .filter(
                estimation_period_q,
                batch__status="DONE",
                site__isnull=False,
                site__invoice_payment__iexact="Aktivco",
                site__grid_fee=True,
            )
        )

        if zone:
            estimation_qs = estimation_qs.filter(site__zone=zone.upper())

        if search:
            estimation_qs = estimation_qs.filter(
                Q(site__site_id__icontains=search)
                | Q(site__name__icontains=search)
            )

        if typo:
            estimation_qs = estimation_qs.filter(
                Q(site__billing_typology__icontains=typo)
                | Q(site__installed_typology__icontains=typo)
                | Q(site__ordered_typology__icontains=typo)
            )

        estimation_rows = (
            estimation_qs
            .values(
                "site__site_id",
                "site__name",
                "site__zone",
                "batch__year",
                "batch__month",
                "source_utilisee",
            )
            .annotate(
                conso_estimee_kwh=Sum("conso_estimee_kwh"),
                montant_estime=Sum("montant_estime"),

                acm_conso_kwh=Sum("acm_conso_kwh"),
                grid_conso_kwh=Sum("grid_conso_kwh"),
                histo_conso_30j=Sum("histo_conso_30j"),
                target_conso_kwh=Sum("target_conso_kwh"),
                theorique_conso_kwh=Sum("theorique_conso_kwh"),

                nb_jours_mois=Max("nb_jours_mois"),
            )
            .order_by("site__site_id", "batch__year", "batch__month")
        )

        estimation_index = {}

        for r in estimation_rows:
            sid = r["site__site_id"]
            y = r["batch__year"]
            m = r["batch__month"]
            key = (sid, y, m)

            estimation_index[key] = r
            all_keys.add(key)

            site_meta[sid] = {
                "site_id": sid,
                "site_name": r["site__name"],
                "zone": r["site__zone"],
            }

        site_ids_in_result = {sid for sid, _, _ in all_keys}

        # ─────────────────────────────────────────────────────────────
        # 3) Jointure FinancialEvaluation : marge / typologie / cible
        # ─────────────────────────────────────────────────────────────
        eval_index = {}

        if site_ids_in_result:
            fe_qs = (
                FinancialEvaluation.objects
                .filter(
                    period_q,
                    site__site_id__in=site_ids_in_result,
                )
                .select_related("fee_rule", "site")
                .values(
                    "site__site_id",
                    "year",
                    "month",
                    "typology",
                    "load_w",
                    "redevance",
                    "marge",
                    "marge_statut",
                    "recurrence_type",
                    "hors_catalogue",
                    "nb_jours_factures",
                    "fee_rule__cible_kwh_j",
                    "fee_rule__cible_kwh",
                )
            )

            eval_index = {
                (fe["site__site_id"], fe["year"], fe["month"]): fe
                for fe in fe_qs
            }

        # ─────────────────────────────────────────────────────────────
        # 4) eFMS / ACM / Grid / Solar
        # ─────────────────────────────────────────────────────────────
        efms_index = {}

        if site_ids_in_result:
            try:
                efms_index = FinancialConsoService.fetch_bulk_for_list(
                    site_ids=list(site_ids_in_result),
                    year_start=ys,
                    month_start=ms,
                    year_end=ye,
                    month_end=me,
                )
            except Exception as e:
                logger.warning("[SuiviConsoView] fetch_bulk_for_list échoué : %s", e)

        # ─────────────────────────────────────────────────────────────
        # 5) Construction des lignes : union Billing + Estimation
        # ─────────────────────────────────────────────────────────────
        result_rows = []

        for sid, year, month in sorted(all_keys, key=lambda x: (x[1], x[2], x[0])):
            key = (sid, year, month)

            meta = site_meta.get(sid, {})
            bill = billing_index.get(key, {})
            est = estimation_index.get(key, {})
            ev = eval_index.get(key, {})
            efms = efms_index.get(key, {})

            if statut:
                if not ev or ev.get("marge_statut") != statut.upper():
                    continue

            # nb_jours = nombre de jours calendaires du mois considéré, point.
            # Plus aucune dépendance aux données de facturation (period_start/
            # end, days_covered, nb_jours_factures...) : c'était la source de
            # tous les bugs précédents (doublement, factures dupliquées, etc.).
            nb_jours = calendar.monthrange(year, month)[1]

            # Cible catalogue : conso_target = nb_jours × cible_kwh_j
            cible_j = ev.get("fee_rule__cible_kwh_j")
            cible_fix = ev.get("fee_rule__cible_kwh")

            if cible_j is not None:
                try:
                    conso_target = str(
                        (_d(cible_j) * _d(nb_jours)).quantize(
                            Decimal("0.001"),
                            rounding=ROUND_HALF_UP,
                        )
                    )
                except Exception:
                    conso_target = None
            elif cible_fix is not None:
                conso_target = fmt(cible_fix)
            else:
                conso_target = None

            # Target solaire
            typo_val = ev.get("typology") or ""
            load_val = ev.get("load_w")
            solar_target = _compute_solar_target(typo_val, load_val, int(nb_jours or 30))

            # Coût moyen énergie facturée
            conso_facturee = bill.get("conso_kwh")
            montant_ht = bill.get("montant_ht")

            if conso_facturee and montant_ht and _d(conso_facturee) > 0:
                cout_nrj = _d(montant_ht) / _d(conso_facturee)
            else:
                cout_nrj = None

            # Priorité eFMS :
            # - si remote SQL disponible : on l’utilise
            # - sinon fallback sur les champs stockés dans EstimationResult
            remote_grid = efms.get("fms_grid_kwh")
            remote_acm = efms.get("fms_acm_kwh")

            est_grid = est.get("grid_conso_kwh")
            est_acm = est.get("acm_conso_kwh")

            fms_grid = remote_grid if remote_grid is not None else est_grid
            fms_acm = remote_acm if remote_acm is not None else est_acm

            result_rows.append({
                "site_id": meta.get("site_id") or sid,
                "site_name": meta.get("site_name"),
                "zone": meta.get("zone"),

                "year": year,
                "month": month,
                "period": f"{year}-{month:02d}",
                "nb_jours": nb_jours,

                # Sénélec facturée
                "conso_kwh": fmt(conso_facturee),
                "conso_facturee_kwh": fmt(conso_facturee),
                "montant_ht": fmt(montant_ht),
                "montant_energie": fmt(bill.get("montant_energie")),
                "cout_moyen_kwh": fmt(cout_nrj),

                # Estimation
                "conso_estimee_kwh": fmt(est.get("conso_estimee_kwh")),
                "montant_estime": fmt(est.get("montant_estime")),
                "source_estimation": est.get("source_utilisee"),
                "estimation_available": est.get("conso_estimee_kwh") is not None,

                # Détails estimation stockés
                "est_acm_kwh": fmt(est.get("acm_conso_kwh")),
                "est_grid_kwh": fmt(est.get("grid_conso_kwh")),
                "est_histo_kwh": fmt(est.get("histo_conso_30j")),
                "est_target_kwh": fmt(est.get("target_conso_kwh")),
                "est_theorique_kwh": fmt(est.get("theorique_conso_kwh")),

                # eFMS / ACM / Grid affichage principal
                "fms_grid_kwh": fmt(fms_grid),
                "fms_acm_kwh": fmt(fms_acm),
                "fms_grid_src": "remote" if remote_grid is not None else ("estimation" if est_grid is not None else None),
                "fms_acm_src": "remote" if remote_acm is not None else ("estimation" if est_acm is not None else None),

                # Solaire
                "solar_kwh": fmt(efms.get("solar_kwh")),
                "solar_target": fmt(solar_target),
                "unavail_hours": efms.get("unavail_hours"),

                # Cible catalogue / marge
                "conso_target": conso_target,
                "typology": ev.get("typology"),
                "load_w": ev.get("load_w"),
                "redevance": fmt(ev.get("redevance")),
                "marge": fmt(ev.get("marge")),
                "marge_statut": ev.get("marge_statut"),
                "recurrence_type": ev.get("recurrence_type"),
                "hors_catalogue": bool(ev.get("hors_catalogue", False)),
            })

        # ─────────────────────────────────────────────────────────────
        # 6) Export CSV
        # ─────────────────────────────────────────────────────────────
        if export == "csv":
            import csv
            import io as _io
            from django.http import HttpResponse

            buf = _io.StringIO()
            w = csv.writer(buf)

            w.writerow([
                "Site ID", "Nom", "Zone", "Année", "Mois", "Nb Jours",
                "Conso facturée kWh",
                "Conso estimée kWh",
                "Source estimation",
                "FMS Grid kWh",
                "FMS ACM kWh",
                "Solar kWh",
                "Solar Target kWh",
                "Montant HT",
                "Montant estimé",
                "Statut",
                "Récurrence",
                "Hors Cat.",
                "Typologie",
                "Load (W)",
            ])

            for rr in result_rows:
                w.writerow([
                    rr["site_id"],
                    rr["site_name"] or "",
                    rr["zone"] or "",
                    rr["year"],
                    rr["month"],
                    rr["nb_jours"],
                    rr["conso_facturee_kwh"] or "",
                    rr["conso_estimee_kwh"] or "",
                    rr["source_estimation"] or "",
                    rr["fms_grid_kwh"] or "",
                    rr["fms_acm_kwh"] or "",
                    rr["solar_kwh"] or "",
                    rr["solar_target"] or "",
                    rr["montant_ht"] or "",
                    rr["montant_estime"] or "",
                    rr["marge_statut"] or "",
                    rr["recurrence_type"] or "",
                    "OUI" if rr["hors_catalogue"] else "",
                    rr["typology"] or "",
                    rr["load_w"] or "",
                ])

            suffix = f"{ys}{ms:02d}-{ye}{me:02d}"
            resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8-sig")
            resp["Content-Disposition"] = f'attachment; filename="suivi_conso_{suffix}.csv"'
            return resp

        # ─────────────────────────────────────────────────────────────
        # 7) Pagination
        # ─────────────────────────────────────────────────────────────
        total = len(result_rows)

        try:
            page = max(1, int(request.query_params.get("page", 1)))
            # Plafond élevé (20000) pour permettre au frontend de récupérer
            # tout le jeu de résultats en 1 seul appel quand il construit les
            # graphiques d'évolution — result_rows est déjà entièrement
            # recalculé à chaque requête (voir plus haut), donc paginer par
            # petits lots (ex: 500) pour tout récupérer forçait des dizaines
            # d'appels qui refaisaient chacun tout le calcul depuis zéro.
            page_size = min(20000, max(25, int(request.query_params.get("page_size", 100))))
        except (ValueError, TypeError):
            page, page_size = 1, 100

        offset = (page - 1) * page_size

        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size if page_size else 1,
            "results": result_rows[offset: offset + page_size],
        })


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard Marge Grid — Focus Sites en Marge Négative (CDC v1.0)
# ─────────────────────────────────────────────────────────────────────────────

class MargeDashboardDataView(APIView):
    """
    Données brutes du Dashboard Marge Grid — une ligne par site, les deux
    bases de marge (estimée/réelle) déjà calculées. Tout le filtrage/tri/
    agrégation (sélecteur de périmètre, base de marge, filtres transverses)
    se fait côté client, comme spécifié dans le CDC.

    Le filtre "période" ne s'applique qu'à la marge réelle (FinancialEvaluation) :
    la marge estimée vient de BOMarginSnapshot, un snapshot figé Mai→Juin, non
    paramétrable par année.
    """
    permission_classes = [IsAuthenticated]

    REAL_INVOICE_SOURCES = [
        FinancialEvaluation.CalculationSource.PAID_INVOICE,
        FinancialEvaluation.CalculationSource.RAW_UNPAID_INVOICE,
    ]

    def get(self, request):
        from financial.services.marge_dashboard_service import build_marge_dashboard_rows

        # Périodes (année, mois) pour lesquelles une vraie facture Sénélec est
        # rapprochée quelque part dans le portefeuille — souvent très peu
        # nombreuses (le rapprochement facture est un processus lent), donc on
        # expose la liste exacte plutôt qu'un simple sélecteur d'année : un
        # sélecteur d'année seul peut retomber sur un mois non représentatif
        # (ex. mars 2026 : 0 site NOK sur 1370, contre 752/1745 en janvier 2026).
        available_periods = list(
            FinancialEvaluation.objects
            .filter(calculation_source__in=self.REAL_INVOICE_SOURCES)
            .order_by("-year", "-month")
            .values_list("year", "month")
            .distinct()
        )
        available_years = sorted({y for y, _m in available_periods}, reverse=True)

        req_year = request.query_params.get("year")
        req_month = request.query_params.get("month")

        if req_year and req_month:
            year, month = int(req_year), int(req_month)
        elif available_periods:
            # Par défaut : la période la plus récente de l'année demandée si elle
            # existe, sinon la période la plus récente disponible tout court.
            target_year = int(req_year) if req_year else timezone.localdate().year
            match = next((p for p in available_periods if p[0] == target_year), None)
            year, month = match or available_periods[0]
        else:
            year, month = timezone.localdate().year, 6

        data = build_marge_dashboard_rows(year, month)
        data["meta"]["available_years"] = available_years or [year]
        data["meta"]["available_periods"] = [{"year": y, "month": m} for y, m in available_periods] or [{"year": year, "month": month}]
        data["meta"]["reelle_year"] = year
        data["meta"]["reelle_month"] = month
        return Response(data)