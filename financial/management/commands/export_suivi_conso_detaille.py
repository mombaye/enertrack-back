# financial/management/commands/export_suivi_conso_detaille.py
from __future__ import annotations

import calendar
import time
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand
from django.db.models import Max, Q, Sum
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from billing.models import MonthlySynthesis as BillingMonthlySynthesis
from core.models import Site
from estimation.models import EstimationResult
from financial.models import FinancialEvaluation
from financial.services.conso_service import FinancialConsoService


DEC0 = Decimal("0")
Q3 = Decimal("0.001")
MONTHS_FR = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jun", "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]

_SOLAR_INVERTER_KW: dict[str, Decimal] = {
    "GG_S0": Decimal("0.736"),
    "GG_S1": Decimal("1.453"),
    "GG_S2": Decimal("2.180"),
    "GG_S3": Decimal("2.907"),
}
SOLAR_X_PCT = Decimal("1.0")


def _period_q(
    ys: int,
    ms: int,
    ye: int,
    me: int,
    *,
    year_field: str = "year",
    month_field: str = "month",
) -> Q:
    """Construit un filtre Q() fiable sur une plage inter-années."""
    if ys == ye:
        return Q(**{year_field: ys, f"{month_field}__gte": ms, f"{month_field}__lte": me})

    return (
        Q(**{year_field: ys, f"{month_field}__gte": ms})
        | Q(**{f"{year_field}__gt": ys, f"{year_field}__lt": ye})
        | Q(**{year_field: ye, f"{month_field}__lte": me})
    )


def _months_between(ys: int, ms: int, ye: int, me: int) -> list[tuple[int, int]]:
    months: list[tuple[int, int]] = []
    y, m = ys, ms
    while (y, m) <= (ye, me):
        months.append((y, m))
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1
    return months


def _year_bounds(year: int, all_months: list[tuple[int, int]]) -> tuple[int, int]:
    months = [m for y, m in all_months if y == year]
    return min(months), max(months)


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Q3, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _to_float(value: Any) -> float | None:
    d = _to_decimal(value)
    return float(d) if d is not None else None


def _num(value: Any) -> float:
    v = _to_float(value)
    return v if v is not None else 0.0


def _blank_or_float(value: Any) -> float | None:
    return _to_float(value)


def _first_not_empty(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def _delta_pct(actual: Any, ref: Any) -> float | None:
    a = _to_float(actual)
    b = _to_float(ref)
    if a is None or b is None or b == 0:
        return None
    return (a / b - 1) * 100


def _reference_conso(facturee: Any, fms_grid: Any, fms_acm: Any, estimation: Any) -> tuple[float | None, str]:
    fact = _to_float(facturee)
    fms_g = _to_float(fms_grid)
    fms_a = _to_float(fms_acm)
    est = _to_float(estimation)

    if fact is not None and fact > 0:
        return fact, "FACTUREE"
    if fms_g is not None and fms_g > 0:
        return fms_g, "FMS"
    if fms_a is not None and fms_a > 0:
        return fms_a, "FMS"
    if est is not None and est > 0:
        return est, "ESTIMATION"
    return None, "NONE"


def _target_status(reference: float | None, target: Any) -> tuple[str, float | None]:
    t = _to_float(target)
    if t is None or t <= 0:
        return "NO_TARGET", None
    if reference is None:
        return "NO_DATA", None
    gap = (reference / t - 1) * 100
    return ("OK" if reference <= t else "NOK"), gap


def _compute_solar_target(typology: str | None, load_w: int | None, nb_jours: int | None) -> Decimal | None:
    """
    Même logique que la page Suivi Conso : cible solaire selon typologie GG_Sx,
    load et nombre de jours. Si la typologie n'est pas solaire, retourne None.
    """
    if not typology or not nb_jours:
        return None

    typo_normalized = str(typology).replace(" ", "_").upper()
    inverter_kw: Decimal | None = None

    for key, kw in _SOLAR_INVERTER_KW.items():
        if key in typo_normalized:
            inverter_kw = kw
            break

    if inverter_kw is None:
        return None

    cap_kwh = inverter_kw * Decimal("24") * Decimal(str(nb_jours))

    if load_w:
        load_kwh = SOLAR_X_PCT * Decimal(str(load_w)) / Decimal("1000") * Decimal("24") * Decimal(str(nb_jours))
        return min(load_kwh, cap_kwh).quantize(Q3, rounding=ROUND_HALF_UP)

    return cap_kwh.quantize(Q3, rounding=ROUND_HALF_UP)


def _format_sheet_title(year: int) -> str:
    return str(year)


class Command(BaseCommand):
    help = (
        "Exporte le suivi consommation au format détaillé de la page SuiviConso "
        "avec feuilles par année, synthèse et traitement SQL Server safe."
    )

    def add_arguments(self, parser):
        now = timezone.now()

        parser.add_argument("--start-year", type=int, default=2024)
        parser.add_argument("--start-month", type=int, default=9)
        parser.add_argument("--end-year", type=int, default=now.year)
        parser.add_argument("--end-month", type=int, default=now.month)

        parser.add_argument(
            "--output",
            type=str,
            default="exports/suivi_conso_detaille.xlsx",
            help="Chemin du fichier XLSX à générer.",
        )

        parser.add_argument("--zone", type=str, default=None, help="Filtrer par zone.")
        parser.add_argument("--search", type=str, default=None, help="Filtrer par site_id ou nom.")
        parser.add_argument("--typology", type=str, default=None, help="Filtrer par typologie contenant cette valeur.")

        parser.add_argument(
            "--remote-mode",
            choices=["none", "safe"],
            default="safe",
            help="none = pas de SQL Server distant ; safe = lots limités + pause.",
        )
        parser.add_argument(
            "--remote-chunk-size",
            type=int,
            default=120,
            help="Nombre de sites par lot pour SQL Server distant.",
        )
        parser.add_argument(
            "--remote-sleep",
            type=float,
            default=0.5,
            help="Pause en secondes entre deux lots SQL Server.",
        )
        parser.add_argument(
            "--only-sites-with-data",
            action="store_true",
            help="Si activé, exclut les sites sans aucune donnée sur la période.",
        )

    def handle(self, *args, **opts):
        ys = int(opts["start_year"])
        ms = int(opts["start_month"])
        ye = int(opts["end_year"])
        me = int(opts["end_month"])

        if (ys, ms) > (ye, me):
            ys, ms, ye, me = ye, me, ys, ms

        if not (1 <= ms <= 12 and 1 <= me <= 12):
            raise ValueError("Les mois doivent être compris entre 1 et 12.")

        output = Path(opts["output"])
        output.parent.mkdir(parents=True, exist_ok=True)

        zone = opts.get("zone")
        search = opts.get("search")
        typology_filter = opts.get("typology")
        remote_mode = opts["remote_mode"]
        remote_chunk_size = max(20, int(opts["remote_chunk_size"]))
        remote_sleep = max(0.0, float(opts["remote_sleep"]))
        only_sites_with_data = bool(opts["only_sites_with_data"])

        all_months = _months_between(ys, ms, ye, me)
        years = sorted({y for y, _ in all_months})
        period_label = f"{ys}-{ms:02d} → {ye}-{me:02d}"

        self.stdout.write(self.style.NOTICE(f"Export suivi conso détaillé : {period_label}"))

        # ─────────────────────────────────────────────────────────────
        # 1) Sites éligibles
        # ─────────────────────────────────────────────────────────────
        site_qs = (
            Site.objects
            .filter(invoice_payment__iexact="Aktivco", grid_fee=True)
            .values(
                "site_id",
                "name",
                "zone",
                "billing_typology",
                "installed_typology",
                "ordered_typology",
                "analysis_load",
            )
            .order_by("zone", "site_id")
        )

        if zone:
            site_qs = site_qs.filter(zone__iexact=zone)

        if search:
            site_qs = site_qs.filter(Q(site_id__icontains=search) | Q(name__icontains=search))

        site_meta: dict[str, dict[str, Any]] = {}
        for s in site_qs.iterator(chunk_size=2000):
            typo = _first_not_empty(
                s.get("billing_typology"),
                s.get("installed_typology"),
                s.get("ordered_typology"),
            )
            if typology_filter and typology_filter.lower() not in str(typo or "").lower():
                continue

            sid = s["site_id"]
            site_meta[sid] = {
                "site_id": sid,
                "site_name": s.get("name"),
                "zone": s.get("zone"),
                "typology": typo,
                "load_w": s.get("analysis_load"),
            }

        site_ids = list(site_meta.keys())
        if not site_ids:
            self.stdout.write(self.style.WARNING("Aucun site trouvé pour les filtres fournis."))
            return

        self.stdout.write(f"Sites éligibles : {len(site_ids)}")
        period = _period_q(ys, ms, ye, me)

        # ─────────────────────────────────────────────────────────────
        # 2) Données locales : facturation Sénélec
        # ─────────────────────────────────────────────────────────────
        billing_index: dict[tuple[str, int, int], dict[str, Any]] = {}

        billing_rows = (
            BillingMonthlySynthesis.objects
            .filter(
                period,
                source__site__site_id__in=site_ids,
                source__site__invoice_payment__iexact="Aktivco",
                source__site__grid_fee=True,
                source__payment_status__in=["PAID", "UNPAID"],
            )
            .values("source__site__site_id", "year", "month")
            .annotate(
                conso_kwh=Sum("conso"),
                montant_ht=Sum("montant_hors_tva"),
                montant_energie=Sum("montant_energie"),
                nb_jours=Sum("days_covered"),
            )
            .order_by("source__site__site_id", "year", "month")
        )

        for r in billing_rows.iterator(chunk_size=5000):
            key = (r["source__site__site_id"], int(r["year"]), int(r["month"]))
            billing_index[key] = r

        self.stdout.write(f"Lignes facturation agrégées : {len(billing_index)}")

        # ─────────────────────────────────────────────────────────────
        # 3) Données locales : estimations
        # ─────────────────────────────────────────────────────────────
        estimation_index: dict[tuple[str, int, int], dict[str, Any]] = {}
        estimation_period = _period_q(ys, ms, ye, me, year_field="batch__year", month_field="batch__month")

        estimation_rows = (
            EstimationResult.objects
            .filter(
                estimation_period,
                batch__status="DONE",
                site__site_id__in=site_ids,
                site__invoice_payment__iexact="Aktivco",
                site__grid_fee=True,
            )
            .values("site__site_id", "batch__year", "batch__month", "source_utilisee")
            .annotate(
                conso_estimee_kwh=Sum("conso_estimee_kwh"),
                montant_estime=Sum("montant_estime"),
                acm_conso_kwh=Sum("acm_conso_kwh"),
                grid_conso_kwh=Sum("grid_conso_kwh"),
                target_conso_kwh=Sum("target_conso_kwh"),
                theorique_conso_kwh=Sum("theorique_conso_kwh"),
                nb_jours_mois=Max("nb_jours_mois"),
            )
            .order_by("site__site_id", "batch__year", "batch__month")
        )

        for r in estimation_rows.iterator(chunk_size=5000):
            key = (r["site__site_id"], int(r["batch__year"]), int(r["batch__month"]))
            estimation_index[key] = r

        self.stdout.write(f"Lignes estimation agrégées : {len(estimation_index)}")

        # ─────────────────────────────────────────────────────────────
        # 4) Évaluation financière : target catalogue / typologie / load
        # ─────────────────────────────────────────────────────────────
        eval_index: dict[tuple[str, int, int], dict[str, Any]] = {}

        eval_rows = (
            FinancialEvaluation.objects
            .filter(period, site__site_id__in=site_ids)
            .values(
                "site__site_id",
                "year",
                "month",
                "typology",
                "load_w",
                "hors_catalogue",
                "fee_rule__cible_kwh",
                "fee_rule__cible_kwh_j",
                "nb_jours_factures",
            )
            .order_by("site__site_id", "year", "month")
        )

        for r in eval_rows.iterator(chunk_size=5000):
            key = (r["site__site_id"], int(r["year"]), int(r["month"]))
            eval_index[key] = r
            meta = site_meta.get(r["site__site_id"])
            if meta:
                meta["typology"] = _first_not_empty(r.get("typology"), meta.get("typology"))
                meta["load_w"] = _first_not_empty(r.get("load_w"), meta.get("load_w"))

        self.stdout.write(f"Lignes évaluation financière : {len(eval_index)}")

        # ─────────────────────────────────────────────────────────────
        # 5) Remote SQL Server : eFMS/ACM/Solar en safe mode
        # ─────────────────────────────────────────────────────────────
        remote_index: dict[tuple[str, int, int], dict[str, Any]] = {}

        if remote_mode == "safe":
            self.stdout.write(
                self.style.NOTICE(
                    f"Remote SQL Server activé en mode safe : chunk={remote_chunk_size}, sleep={remote_sleep}s"
                )
            )
            for year in years:
                y_ms, y_me = _year_bounds(year, all_months)

                for start in range(0, len(site_ids), remote_chunk_size):
                    chunk = site_ids[start:start + remote_chunk_size]
                    try:
                        data = FinancialConsoService.fetch_bulk_for_list(
                            site_ids=chunk,
                            year_start=year,
                            month_start=y_ms,
                            year_end=year,
                            month_end=y_me,
                        )
                        remote_index.update(data)
                    except Exception as exc:
                        self.stdout.write(
                            self.style.WARNING(
                                f"Remote ignoré pour année={year}, lot={start // remote_chunk_size + 1} : {exc}"
                            )
                        )

                    if remote_sleep:
                        time.sleep(remote_sleep)

                self.stdout.write(f"Remote année {year} terminé.")
        else:
            self.stdout.write(self.style.WARNING("Remote SQL Server désactivé : export basé sur les données locales."))

        # ─────────────────────────────────────────────────────────────
        # 6) Construction des lignes au format SuiviConsoPage
        # ─────────────────────────────────────────────────────────────
        rows_by_year: dict[int, list[dict[str, Any]]] = defaultdict(list)
        site_summary_by_year: dict[int, dict[str, dict[str, Any]]] = defaultdict(dict)
        summary_by_year: dict[int, dict[str, Any]] = {}

        for year in years:
            months = [m for y, m in all_months if y == year]
            year_stats = {
                "rows": 0,
                "facturee": 0.0,
                "fms": 0.0,
                "solar": 0.0,
                "solar_target": 0.0,
                "estimation": 0.0,
                "target": 0.0,
                "ok": 0,
                "nok": 0,
                "no_target": 0,
                "no_data": 0,
                "suivi_total": 0.0,
                "suivi_points": len(site_ids) * len(months),
                "points_with_data": 0,
            }

            for sid in site_ids:
                meta = site_meta[sid]
                site_values_for_average: list[float] = []
                site_points_with_data = 0

                for month in months:
                    key = (sid, year, month)
                    bill = billing_index.get(key, {})
                    est = estimation_index.get(key, {})
                    ev = eval_index.get(key, {})
                    rem = remote_index.get(key, {})

                    nb_jours = (
                        bill.get("nb_jours")
                        or ev.get("nb_jours_factures")
                        or est.get("nb_jours_mois")
                        or calendar.monthrange(year, month)[1]
                    )
                    try:
                        nb_jours_int = int(nb_jours) if nb_jours is not None else calendar.monthrange(year, month)[1]
                    except Exception:
                        nb_jours_int = calendar.monthrange(year, month)[1]

                    typology = _first_not_empty(ev.get("typology"), meta.get("typology"))
                    load_w = _first_not_empty(ev.get("load_w"), meta.get("load_w"))
                    try:
                        load_w_int = int(load_w) if load_w not in (None, "") else None
                    except Exception:
                        load_w_int = None

                    facturee = _blank_or_float(bill.get("conso_kwh"))
                    montant_ht = _blank_or_float(bill.get("montant_ht"))

                    remote_grid = rem.get("fms_grid_kwh")
                    remote_acm = rem.get("fms_acm_kwh")
                    est_grid = est.get("grid_conso_kwh")
                    est_acm = est.get("acm_conso_kwh")

                    fms_grid = _blank_or_float(remote_grid if remote_grid is not None else est_grid)
                    fms_acm = _blank_or_float(remote_acm if remote_acm is not None else est_acm)
                    fms_grid_src = "remote" if remote_grid is not None else ("estimation" if est_grid is not None else None)
                    fms_acm_src = "remote" if remote_acm is not None else ("estimation" if est_acm is not None else None)

                    solar_kwh = _blank_or_float(rem.get("solar_kwh"))
                    unavail_hours = _blank_or_float(rem.get("unavail_hours"))
                    solar_target = _compute_solar_target(typology, load_w_int, nb_jours_int)
                    solar_target_float = _blank_or_float(solar_target)

                    estimation = _blank_or_float(est.get("conso_estimee_kwh"))
                    source_estimation = est.get("source_utilisee")

                    # Conso target catalogue : cible_jour × nb_jours si disponible, sinon cible_kwh.
                    conso_target = None
                    cible_kwh_j = ev.get("fee_rule__cible_kwh_j")
                    cible_kwh = ev.get("fee_rule__cible_kwh")
                    if cible_kwh_j is not None:
                        try:
                            conso_target = (
                                Decimal(str(cible_kwh_j)) * Decimal(str(nb_jours_int))
                            ).quantize(Q3, rounding=ROUND_HALF_UP)
                        except Exception:
                            conso_target = None
                    if conso_target is None and cible_kwh is not None:
                        conso_target = cible_kwh
                    conso_target_float = _blank_or_float(conso_target)

                    ref_value, ref_source = _reference_conso(facturee, fms_grid, fms_acm, estimation)
                    statut_target, delta_vs_target = _target_status(ref_value, conso_target_float)

                    fms_main = fms_grid if fms_grid is not None else fms_acm
                    delta_fms_fact = _delta_pct(fms_main, facturee)
                    delta_sol_cible = _delta_pct(solar_kwh, solar_target_float)

                    # Suivi conso demandé : moyenne/somme avec missing=0.
                    suivi_val = 0.0
                    suivi_source = "MISSING"
                    if facturee is not None and facturee > 0:
                        suivi_val, suivi_source = facturee, "FACTUREE"
                    elif estimation is not None and estimation > 0:
                        suivi_val, suivi_source = estimation, "ESTIMATION"
                    elif fms_acm is not None and fms_acm > 0:
                        suivi_val, suivi_source = fms_acm, "ACM"
                    elif fms_grid is not None and fms_grid > 0:
                        suivi_val, suivi_source = fms_grid, "GRID"

                    site_values_for_average.append(suivi_val)
                    if suivi_source != "MISSING":
                        site_points_with_data += 1

                    if only_sites_with_data and suivi_source == "MISSING":
                        continue

                    row = {
                        "Site": sid,
                        "Nom site": meta.get("site_name"),
                        "Zone": meta.get("zone"),
                        "Période": f"{year}-{month:02d}",
                        "Jours": nb_jours_int,
                        "Conso facturée kWh": facturee,
                        "Réf.": "Référence" if ref_source == "FACTUREE" else "",
                        "FMS Grid kWh": fms_grid,
                        "FMS ACM kWh": fms_acm,
                        "Δ FMS/Fact %": delta_fms_fact,
                        "Source FMS": fms_grid_src or fms_acm_src or "",
                        "Solar kWh": solar_kwh,
                        "Solar Target kWh": solar_target_float,
                        "Δ Sol/Cible %": delta_sol_cible,
                        "Conso estimée kWh": estimation,
                        "Source estim.": source_estimation or "",
                        "Typologie": typology,
                        "Load (W)": load_w_int,
                        "Conso Target kWh": conso_target_float,
                        "Statut target": statut_target,
                        "Référence target": ref_source,
                        "Δ vs Target %": delta_vs_target,
                        "Suivi conso retenu kWh": suivi_val,
                        "Source suivi": suivi_source,
                        "Monitoring unavailability h": unavail_hours,
                    }
                    rows_by_year[year].append(row)

                    year_stats["rows"] += 1
                    year_stats["facturee"] += facturee or 0.0
                    year_stats["fms"] += fms_main or 0.0
                    year_stats["solar"] += solar_kwh or 0.0
                    year_stats["solar_target"] += solar_target_float or 0.0
                    year_stats["estimation"] += estimation or 0.0
                    year_stats["target"] += conso_target_float or 0.0
                    year_stats["suivi_total"] += suivi_val
                    if suivi_source != "MISSING":
                        year_stats["points_with_data"] += 1
                    if statut_target == "OK":
                        year_stats["ok"] += 1
                    elif statut_target == "NOK":
                        year_stats["nok"] += 1
                    elif statut_target == "NO_TARGET":
                        year_stats["no_target"] += 1
                    elif statut_target == "NO_DATA":
                        year_stats["no_data"] += 1

                # synthèse site par année : missing = 0 dans la moyenne.
                if site_values_for_average:
                    total_site = sum(site_values_for_average)
                    site_summary_by_year[year][sid] = {
                        "Zone": meta.get("zone"),
                        "Site": sid,
                        "Nom site": meta.get("site_name"),
                        "Typologie": meta.get("typology"),
                        "Load (W)": meta.get("load_w"),
                        "Total suivi conso kWh": total_site,
                        "Moyenne suivi conso kWh": total_site / len(site_values_for_average),
                        "Mois avec donnée": site_points_with_data,
                        "Mois manquants": len(site_values_for_average) - site_points_with_data,
                        "Taux couverture": site_points_with_data / len(site_values_for_average),
                    }

            summary_by_year[year] = year_stats

        # ─────────────────────────────────────────────────────────────
        # 7) Génération Excel
        # ─────────────────────────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(wb, summary_by_year, period_label, remote_mode, remote_chunk_size, remote_sleep)
        self._create_site_summary_sheet(wb, site_summary_by_year)

        for year in years:
            self._create_detail_sheet(wb, year, rows_by_year[year])

        wb.save(output)
        self.stdout.write(self.style.SUCCESS(f"Export généré : {output}"))

    # ─────────────────────────────────────────────────────────────
    # Excel helpers
    # ─────────────────────────────────────────────────────────────

    def _styles(self):
        thin = Side(style="thin", color="CBD5E1")
        return {
            "navy_fill": PatternFill("solid", fgColor="010E2A"),
            "blue_dark": PatternFill("solid", fgColor="032566"),
            "blue": PatternFill("solid", fgColor="0A3D96"),
            "teal": PatternFill("solid", fgColor="0E7490"),
            "solar": PatternFill("solid", fgColor="B45309"),
            "estim": PatternFill("solid", fgColor="5B21B6"),
            "green": PatternFill("solid", fgColor="DCFCE7"),
            "red": PatternFill("solid", fgColor="FEE2E2"),
            "orange": PatternFill("solid", fgColor="FEF3C7"),
            "slate": PatternFill("solid", fgColor="F1F5F9"),
            "white_font": Font(color="FFFFFF", bold=True),
            "title_font": Font(color="FFFFFF", bold=True, size=16),
            "header_font": Font(color="FFFFFF", bold=True, size=10),
            "bold": Font(bold=True),
            "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        }

    def _create_summary_sheet(
        self,
        wb: Workbook,
        summary_by_year: dict[int, dict[str, Any]],
        period_label: str,
        remote_mode: str,
        remote_chunk_size: int,
        remote_sleep: float,
    ) -> None:
        st = self._styles()
        ws = wb.create_sheet("Synthèse")

        ws.merge_cells("A1:N1")
        ws["A1"] = "Synthèse — Suivi des consommations"
        ws["A1"].fill = st["navy_fill"]
        ws["A1"].font = st["title_font"]
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws["A2"] = "Période"
        ws["B2"] = period_label
        ws["D2"] = "Généré le"
        ws["E2"] = timezone.now().strftime("%Y-%m-%d %H:%M")
        ws["A3"] = "Remote SQL Server"
        ws["B3"] = f"{remote_mode} | chunk={remote_chunk_size} | sleep={remote_sleep}s"
        ws["D3"] = "Règle suivi conso"
        ws["E3"] = "Facturée > Estimation > ACM > Grid > 0"

        headers = [
            "Année",
            "Nb lignes",
            "Facturée kWh",
            "eFMS kWh",
            "Solaire kWh",
            "Solar Target kWh",
            "Estimée kWh",
            "Target kWh",
            "Target OK",
            "Target NOK",
            "Sans target",
            "Sans donnée",
            "Total suivi conso kWh",
            "Couverture suivi",
        ]
        header_row = 5
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(header_row, col_idx, h)
            c.fill = st["blue_dark"]
            c.font = st["white_font"]
            c.border = st["border"]
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for year in sorted(summary_by_year):
            s = summary_by_year[year]
            possible = s["suivi_points"] or 1
            ws.append([
                year,
                s["rows"],
                round(s["facturee"], 3),
                round(s["fms"], 3),
                round(s["solar"], 3),
                round(s["solar_target"], 3),
                round(s["estimation"], 3),
                round(s["target"], 3),
                s["ok"],
                s["nok"],
                s["no_target"],
                s["no_data"],
                round(s["suivi_total"], 3),
                s["points_with_data"] / possible,
            ])

        last_row = ws.max_row
        if last_row >= 6:
            for row in ws.iter_rows(min_row=6, max_row=last_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = st["border"]
                    cell.alignment = Alignment(vertical="center")
            for col_idx in [3, 4, 5, 6, 7, 8, 13]:
                for row_idx in range(6, last_row + 1):
                    ws.cell(row_idx, col_idx).number_format = '#,##0.000'
            for row_idx in range(6, last_row + 1):
                ws.cell(row_idx, 14).number_format = "0.00%"

            tab = Table(displayName="TableSyntheseSuiviConso", ref=f"A5:N{last_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tab)

            chart = BarChart()
            chart.title = "Total suivi conso par année"
            chart.y_axis.title = "kWh"
            chart.x_axis.title = "Année"
            data = Reference(ws, min_col=13, min_row=5, max_row=last_row)
            cats = Reference(ws, min_col=1, min_row=6, max_row=last_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, "P5")

        for idx, width in enumerate([12, 12, 16, 16, 16, 18, 16, 16, 12, 12, 14, 14, 22, 16], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A6"

    def _create_site_summary_sheet(self, wb: Workbook, site_summary_by_year: dict[int, dict[str, dict[str, Any]]]) -> None:
        st = self._styles()
        ws = wb.create_sheet("Synthèse sites")

        headers = [
            "Année",
            "Zone",
            "Site",
            "Nom site",
            "Typologie",
            "Load (W)",
            "Total suivi conso kWh",
            "Moyenne suivi conso kWh",
            "Mois avec donnée",
            "Mois manquants",
            "Taux couverture",
        ]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"] = "Synthèse par site — moyenne et somme avec mois manquants = 0"
        ws["A1"].fill = st["navy_fill"]
        ws["A1"].font = st["title_font"]
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 28

        header_row = 3
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(header_row, col_idx, h)
            c.fill = st["blue_dark"]
            c.font = st["white_font"]
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = st["border"]

        for year in sorted(site_summary_by_year):
            for sid, row in sorted(site_summary_by_year[year].items(), key=lambda x: (x[1].get("Zone") or "", x[0])):
                ws.append([
                    year,
                    row.get("Zone") or "",
                    sid,
                    row.get("Nom site") or "",
                    row.get("Typologie") or "",
                    row.get("Load (W)") or "",
                    round(row.get("Total suivi conso kWh") or 0, 3),
                    round(row.get("Moyenne suivi conso kWh") or 0, 3),
                    row.get("Mois avec donnée") or 0,
                    row.get("Mois manquants") or 0,
                    row.get("Taux couverture") or 0,
                ])

        last_row = ws.max_row
        for row in ws.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = st["border"]
                cell.alignment = Alignment(vertical="center")

        if last_row >= 4:
            for row_idx in range(4, last_row + 1):
                ws.cell(row_idx, 7).number_format = '#,##0.000'
                ws.cell(row_idx, 8).number_format = '#,##0.000'
                ws.cell(row_idx, 11).number_format = "0.00%"
            tab = Table(displayName="TableSyntheseSitesConso", ref=f"A3:K{last_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tab)

        widths = [10, 12, 14, 28, 22, 12, 22, 24, 18, 16, 16]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A4"

    def _create_detail_sheet(self, wb: Workbook, year: int, rows: list[dict[str, Any]]) -> None:
        st = self._styles()
        ws = wb.create_sheet(_format_sheet_title(year))

        headers = [
            "Site",
            "Nom site",
            "Zone",
            "Période",
            "Jours",
            "Conso facturée kWh",
            "Réf.",
            "FMS Grid kWh",
            "FMS ACM kWh",
            "Δ FMS/Fact %",
            "Source FMS",
            "Solar kWh",
            "Solar Target kWh",
            "Δ Sol/Cible %",
            "Conso estimée kWh",
            "Source estim.",
            "Typologie",
            "Load (W)",
            "Conso Target kWh",
            "Statut target",
            "Référence target",
            "Δ vs Target %",
            "Suivi conso retenu kWh",
            "Source suivi",
            "Monitoring unavailability h",
        ]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"] = f"Tableau détaillé — Suivi consommation {year}"
        ws["A1"].fill = st["navy_fill"]
        ws["A1"].font = st["title_font"]
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Groupes de colonnes comme la page front.
        groups = [
            ("Site", 1, 4, st["blue_dark"]),
            ("Facturée", 5, 7, st["blue"]),
            ("eFMS / ACM", 8, 11, st["teal"]),
            ("Solaire", 12, 14, st["solar"]),
            ("Estimation", 15, 16, st["estim"]),
            ("Target", 17, 22, st["blue"]),
            ("Suivi conso", 23, 25, st["blue_dark"]),
        ]
        group_row = 3
        for title, start_col, end_col, fill in groups:
            ws.merge_cells(start_row=group_row, start_column=start_col, end_row=group_row, end_column=end_col)
            c = ws.cell(group_row, start_col, title)
            c.fill = fill
            c.font = st["white_font"]
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = st["border"]

        header_row = 4
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(header_row, col_idx, h)
            c.fill = st["blue"]
            c.font = st["header_font"]
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = st["border"]

        data_start = 5
        for row_idx, row in enumerate(rows, data_start):
            values = [row.get(h) for h in headers]
            for col_idx, value in enumerate(values, 1):
                c = ws.cell(row_idx, col_idx, value)
                c.border = st["border"]
                c.alignment = Alignment(vertical="center")

            status = row.get("Statut target")
            if status == "NOK":
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row_idx, col_idx).fill = st["red"]
            elif status == "OK":
                ws.cell(row_idx, 20).fill = st["green"]
            elif status in ("NO_TARGET", "NO_DATA"):
                ws.cell(row_idx, 20).fill = st["orange"] if status == "NO_DATA" else st["slate"]

            if row.get("Source suivi") == "MISSING":
                ws.cell(row_idx, 23).fill = st["slate"]

        last_row = max(ws.max_row, data_start)

        # Formats numériques
        numeric_cols = [6, 8, 9, 12, 13, 15, 19, 23, 25]
        pct_cols = [10, 14, 22]
        for row_idx in range(data_start, last_row + 1):
            for col_idx in numeric_cols:
                ws.cell(row_idx, col_idx).number_format = '#,##0.000'
            for col_idx in pct_cols:
                ws.cell(row_idx, col_idx).number_format = '0.00%'
                value = ws.cell(row_idx, col_idx).value
                if isinstance(value, (int, float)):
                    ws.cell(row_idx, col_idx).value = value / 100

        # Mise en forme conditionnelle des deltas.
        for col_idx in pct_cols:
            letter = get_column_letter(col_idx)
            rng = f"{letter}{data_start}:{letter}{last_row}"
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f"ABS({letter}{data_start})>0.2"], fill=st["red"]))
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f"ABS({letter}{data_start})<=0.1"], fill=st["green"]))

        # Table Excel
        if rows:
            last_col = get_column_letter(len(headers))
            tab = Table(displayName=f"TableSuiviConso{year}", ref=f"A4:{last_col}{last_row}")
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)

        widths = {
            1: 14, 2: 28, 3: 12, 4: 12, 5: 10, 6: 18, 7: 12,
            8: 16, 9: 16, 10: 14, 11: 14,
            12: 14, 13: 18, 14: 14,
            15: 18, 16: 15,
            17: 24, 18: 12, 19: 18, 20: 16, 21: 16, 22: 14,
            23: 22, 24: 14, 25: 20,
        }
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = "D5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
