# financial/management/commands/export_suivi_conso_annuel.py
from __future__ import annotations

import calendar
import time
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand
from django.db.models import Max, Q, Sum
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from billing.models import MonthlySynthesis as BillingMonthlySynthesis
from core.models import Site
from estimation.models import EstimationResult
from financial.models import FinancialEvaluation
from financial.services.conso_service import FinancialConsoService


DEC0 = Decimal("0")
Q3 = Decimal("0.001")


def _period_q(
    ys: int,
    ms: int,
    ye: int,
    me: int,
    *,
    year_field: str = "year",
    month_field: str = "month",
) -> Q:
    """Filtre une plage inter-années sur des champs year/month."""
    if ys == ye:
        return Q(**{year_field: ys, f"{month_field}__gte": ms, f"{month_field}__lte": me})

    return (
        Q(**{year_field: ys, f"{month_field}__gte": ms})
        | Q(**{f"{year_field}__gt": ys, f"{year_field}__lt": ye})
        | Q(**{year_field: ye, f"{month_field}__lte": me})
    )


def _months_between(ys: int, ms: int, ye: int, me: int) -> list[tuple[int, int]]:
    months: list[tuple[int, int]] = []
    y, m = ys, ms
    while (y, m) <= (ye, me):
        months.append((y, m))
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1
    return months


def _year_bounds(year: int, all_months: list[tuple[int, int]]) -> tuple[int, int]:
    months = [m for y, m in all_months if y == year]
    return min(months), max(months)


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value)).quantize(Q3)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _to_float(value: Any) -> float:
    d = _to_decimal(value)
    return float(d) if d is not None else 0.0


def _first_not_empty(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


class Command(BaseCommand):
    help = (
        "Exporte le suivi des consommations par année dans un fichier Excel moderne. "
        "Par défaut : période depuis septembre 2024 jusqu'au mois courant."
    )

    def add_arguments(self, parser):
        now = timezone.now()

        parser.add_argument("--start-year", type=int, default=2024)
        parser.add_argument("--start-month", type=int, default=9)
        parser.add_argument("--end-year", type=int, default=now.year)
        parser.add_argument("--end-month", type=int, default=now.month)

        parser.add_argument(
            "--output",
            type=str,
            default="exports/suivi_conso_annuel.xlsx",
            help="Chemin du fichier XLSX à générer.",
        )

        parser.add_argument("--zone", type=str, default=None, help="Filtrer par zone.")
        parser.add_argument("--search", type=str, default=None, help="Filtrer par site_id ou nom.")

        parser.add_argument(
            "--remote-mode",
            choices=["none", "safe"],
            default="safe",
            help=(
                "none = pas de requête SQL Server distante. "
                "safe = requêtes eFMS/Solar en lots limités avec pause."
            ),
        )
        parser.add_argument(
            "--remote-chunk-size",
            type=int,
            default=120,
            help="Nombre de sites par lot pour SQL Server distant.",
        )
        parser.add_argument(
            "--remote-sleep",
            type=float,
            default=0.5,
            help="Pause en secondes entre deux lots SQL Server.",
        )
        parser.add_argument(
            "--only-sites-with-data",
            action="store_true",
            help="Si activé, exclut les sites sans aucune donnée sur la période.",
        )

    def handle(self, *args, **opts):
        ys = int(opts["start_year"])
        ms = int(opts["start_month"])
        ye = int(opts["end_year"])
        me = int(opts["end_month"])

        if (ys, ms) > (ye, me):
            ys, ms, ye, me = ye, me, ys, ms

        if not (1 <= ms <= 12 and 1 <= me <= 12):
            raise ValueError("Les mois doivent être compris entre 1 et 12.")

        output = Path(opts["output"])
        output.parent.mkdir(parents=True, exist_ok=True)

        zone = opts.get("zone")
        search = opts.get("search")
        remote_mode = opts["remote_mode"]
        remote_chunk_size = max(20, int(opts["remote_chunk_size"]))
        remote_sleep = max(0.0, float(opts["remote_sleep"]))
        only_sites_with_data = bool(opts["only_sites_with_data"])

        all_months = _months_between(ys, ms, ye, me)
        years = sorted({y for y, _ in all_months})
        period_label = f"{ys}-{ms:02d} → {ye}-{me:02d}"

        self.stdout.write(self.style.NOTICE(f"Export suivi conso : {period_label}"))

        # ─────────────────────────────────────────────────────────────
        # 1) Sites éligibles
        # ─────────────────────────────────────────────────────────────
        site_qs = (
            Site.objects
            .filter(invoice_payment__iexact="Aktivco", grid_fee=True)
            .values(
                "site_id",
                "name",
                "zone",
                "billing_typology",
                "installed_typology",
                "ordered_typology",
                "analysis_load",
            )
            .order_by("zone", "site_id")
        )

        if zone:
            site_qs = site_qs.filter(zone=zone.upper())

        if search:
            site_qs = site_qs.filter(Q(site_id__icontains=search) | Q(name__icontains=search))

        site_meta: dict[str, dict[str, Any]] = {}
        for s in site_qs.iterator(chunk_size=2000):
            sid = s["site_id"]
            site_meta[sid] = {
                "site_id": sid,
                "site_name": s.get("name"),
                "zone": s.get("zone"),
                "typology": _first_not_empty(
                    s.get("billing_typology"),
                    s.get("installed_typology"),
                    s.get("ordered_typology"),
                ),
                "load_w": s.get("analysis_load"),
            }

        site_ids = list(site_meta.keys())
        if not site_ids:
            self.stdout.write(self.style.WARNING("Aucun site trouvé pour les filtres fournis."))
            return

        self.stdout.write(f"Sites éligibles : {len(site_ids)}")

        period = _period_q(ys, ms, ye, me)

        # ─────────────────────────────────────────────────────────────
        # 2) Données locales : Facturation Sénélec
        # ─────────────────────────────────────────────────────────────
        billing_index: dict[tuple[str, int, int], dict[str, Any]] = {}

        billing_rows = (
            BillingMonthlySynthesis.objects
            .filter(
                period,
                source__site__site_id__in=site_ids,
                source__site__invoice_payment__iexact="Aktivco",
                source__site__grid_fee=True,
                source__payment_status__in=["PAID", "UNPAID"],
            )
            .values("source__site__site_id", "year", "month")
            .annotate(
                conso_kwh=Sum("conso"),
                montant_ht=Sum("montant_hors_tva"),
                montant_energie=Sum("montant_energie"),
                nb_jours=Sum("days_covered"),
            )
            .order_by("source__site__site_id", "year", "month")
        )

        for r in billing_rows.iterator(chunk_size=5000):
            key = (r["source__site__site_id"], int(r["year"]), int(r["month"]))
            billing_index[key] = r

        self.stdout.write(f"Lignes facturation agrégées : {len(billing_index)}")

        # ─────────────────────────────────────────────────────────────
        # 3) Données locales : Estimation
        # ─────────────────────────────────────────────────────────────
        estimation_index: dict[tuple[str, int, int], dict[str, Any]] = {}

        estimation_period = _period_q(
            ys,
            ms,
            ye,
            me,
            year_field="batch__year",
            month_field="batch__month",
        )

        estimation_rows = (
            EstimationResult.objects
            .filter(
                estimation_period,
                batch__status="DONE",
                site__site_id__in=site_ids,
                site__invoice_payment__iexact="Aktivco",
                site__grid_fee=True,
            )
            .values("site__site_id", "batch__year", "batch__month", "source_utilisee")
            .annotate(
                conso_estimee_kwh=Sum("conso_estimee_kwh"),
                montant_estime=Sum("montant_estime"),
                acm_conso_kwh=Sum("acm_conso_kwh"),
                grid_conso_kwh=Sum("grid_conso_kwh"),
                target_conso_kwh=Sum("target_conso_kwh"),
                theorique_conso_kwh=Sum("theorique_conso_kwh"),
                nb_jours_mois=Max("nb_jours_mois"),
            )
            .order_by("site__site_id", "batch__year", "batch__month")
        )

        for r in estimation_rows.iterator(chunk_size=5000):
            key = (r["site__site_id"], int(r["batch__year"]), int(r["batch__month"]))
            estimation_index[key] = r

        self.stdout.write(f"Lignes estimation agrégées : {len(estimation_index)}")

        # ─────────────────────────────────────────────────────────────
        # 4) FinancialEvaluation : typologie / load / marge
        # ─────────────────────────────────────────────────────────────
        eval_index: dict[tuple[str, int, int], dict[str, Any]] = {}

        eval_rows = (
            FinancialEvaluation.objects
            .filter(period, site__site_id__in=site_ids)
            .values(
                "site__site_id",
                "year",
                "month",
                "typology",
                "load_w",
                "redevance",
                "marge",
                "marge_statut",
                "recurrence_type",
                "hors_catalogue",
            )
            .order_by("site__site_id", "year", "month")
        )

        for r in eval_rows.iterator(chunk_size=5000):
            key = (r["site__site_id"], int(r["year"]), int(r["month"]))
            eval_index[key] = r

            # Améliore les meta site si FinancialEvaluation est plus précis
            meta = site_meta.get(r["site__site_id"])
            if meta:
                meta["typology"] = _first_not_empty(r.get("typology"), meta.get("typology"))
                meta["load_w"] = _first_not_empty(r.get("load_w"), meta.get("load_w"))

        self.stdout.write(f"Lignes évaluation financière : {len(eval_index)}")

        # ─────────────────────────────────────────────────────────────
        # 5) Remote SQL Server : eFMS/ACM/Solar en mode SAFE
        # ─────────────────────────────────────────────────────────────
        remote_index: dict[tuple[str, int, int], dict[str, Any]] = {}

        if remote_mode == "safe":
            self.stdout.write(
                self.style.NOTICE(
                    f"Remote SQL Server activé en mode safe : chunk={remote_chunk_size}, sleep={remote_sleep}s"
                )
            )

            for year in years:
                y_ms, y_me = _year_bounds(year, all_months)

                for start in range(0, len(site_ids), remote_chunk_size):
                    chunk = site_ids[start:start + remote_chunk_size]

                    try:
                        data = FinancialConsoService.fetch_bulk_for_list(
                            site_ids=chunk,
                            year_start=year,
                            month_start=y_ms,
                            year_end=year,
                            month_end=y_me,
                        )
                        remote_index.update(data)
                    except Exception as exc:
                        self.stdout.write(
                            self.style.WARNING(
                                f"Remote ignoré pour année={year}, lot={start // remote_chunk_size + 1} : {exc}"
                            )
                        )

                    if remote_sleep:
                        time.sleep(remote_sleep)

                self.stdout.write(f"Remote année {year} terminé.")

        else:
            self.stdout.write(self.style.WARNING("Remote SQL Server désactivé : export basé sur les données locales."))

        # ─────────────────────────────────────────────────────────────
        # 6) Préparation lignes par année
        # ─────────────────────────────────────────────────────────────
        rows_by_year: dict[int, list[dict[str, Any]]] = defaultdict(list)
        summary_by_year: dict[int, dict[str, Any]] = {}

        def choose_suivi_value(key: tuple[str, int, int]) -> tuple[float, str]:
            """
            Règle de calcul du suivi conso :
              1. Conso facturée Sénélec si présente
              2. Conso estimée si présente
              3. ACM remote si présent
              4. Grid remote si présent
              5. 0 si donnée manquante

            Comme demandé : une donnée manquante reste à 0,
            donc les formules Excel SUM/AVERAGE prennent bien le manque en compte.
            """
            bill = billing_index.get(key, {})
            est = estimation_index.get(key, {})
            rem = remote_index.get(key, {})

            if bill.get("conso_kwh") is not None:
                return _to_float(bill.get("conso_kwh")), "FACTUREE"

            if est.get("conso_estimee_kwh") is not None:
                return _to_float(est.get("conso_estimee_kwh")), "ESTIMATION"

            if rem.get("fms_acm_kwh") is not None:
                return _to_float(rem.get("fms_acm_kwh")), "ACM_REMOTE"

            if rem.get("fms_grid_kwh") is not None:
                return _to_float(rem.get("fms_grid_kwh")), "GRID_REMOTE"

            return 0.0, "MISSING"

        for year in years:
            months = [m for y, m in all_months if y == year]
            year_total = 0.0
            year_with_data = 0
            year_possible_points = len(site_ids) * len(months)

            for sid in site_ids:
                meta = site_meta[sid]
                monthly_values: list[float] = []
                monthly_sources: list[str] = []

                for month in months:
                    value, source = choose_suivi_value((sid, year, month))
                    monthly_values.append(value)
                    monthly_sources.append(source)

                if only_sites_with_data and all(src == "MISSING" for src in monthly_sources):
                    continue

                source_counter = Counter([s for s in monthly_sources if s != "MISSING"])
                dominant_source = source_counter.most_common(1)[0][0] if source_counter else "MISSING"

                mois_avec_donnee = sum(1 for s in monthly_sources if s != "MISSING")
                mois_manquants = len(months) - mois_avec_donnee
                total = sum(monthly_values)

                rows_by_year[year].append({
                    "zone": meta.get("zone"),
                    "site_id": sid,
                    "site_name": meta.get("site_name"),
                    "typology": meta.get("typology"),
                    "load_w": meta.get("load_w"),
                    "months": months,
                    "values": monthly_values,
                    "sources": monthly_sources,
                    "mois_avec_donnee": mois_avec_donnee,
                    "mois_manquants": mois_manquants,
                    "source_dominante": dominant_source,
                })

                year_total += total
                year_with_data += mois_avec_donnee

            summary_by_year[year] = {
                "months_count": len(months),
                "sites_count": len(rows_by_year[year]),
                "total_kwh": year_total,
                "avg_kwh": year_total / year_possible_points if year_possible_points else 0,
                "points_with_data": year_with_data,
                "possible_points": year_possible_points,
                "coverage": year_with_data / year_possible_points if year_possible_points else 0,
            }

        # ─────────────────────────────────────────────────────────────
        # 7) Génération Excel
        # ─────────────────────────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(
            wb=wb,
            summary_by_year=summary_by_year,
            period_label=period_label,
            remote_mode=remote_mode,
            remote_chunk_size=remote_chunk_size,
            remote_sleep=remote_sleep,
        )

        for year in years:
            self._create_year_sheet(wb, year, rows_by_year[year])

        wb.save(output)
        self.stdout.write(self.style.SUCCESS(f"Export généré : {output}"))

    # ─────────────────────────────────────────────────────────────
    # Excel helpers
    # ─────────────────────────────────────────────────────────────

    def _styles(self):
        return {
            "title_fill": PatternFill("solid", fgColor="0F172A"),
            "header_fill": PatternFill("solid", fgColor="1D4ED8"),
            "subheader_fill": PatternFill("solid", fgColor="E0F2FE"),
            "zero_fill": PatternFill("solid", fgColor="F1F5F9"),
            "good_fill": PatternFill("solid", fgColor="DCFCE7"),
            "warn_fill": PatternFill("solid", fgColor="FEF3C7"),
            "bad_fill": PatternFill("solid", fgColor="FEE2E2"),
            "white_font": Font(color="FFFFFF", bold=True),
            "title_font": Font(color="FFFFFF", bold=True, size=16),
            "header_font": Font(color="FFFFFF", bold=True),
            "normal_bold": Font(bold=True),
            "thin_border": Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            ),
        }

    def _create_summary_sheet(
        self,
        wb: Workbook,
        summary_by_year: dict[int, dict[str, Any]],
        period_label: str,
        remote_mode: str,
        remote_chunk_size: int,
        remote_sleep: float,
    ) -> None:
        st = self._styles()
        ws = wb.create_sheet("Synthèse")

        ws.merge_cells("A1:H1")
        ws["A1"] = "Synthèse annuelle — Suivi des consommations"
        ws["A1"].fill = st["title_fill"]
        ws["A1"].font = st["title_font"]
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws["A2"] = "Période"
        ws["B2"] = period_label
        ws["D2"] = "Généré le"
        ws["E2"] = timezone.now().strftime("%Y-%m-%d %H:%M")
        ws["A3"] = "Remote SQL Server"
        ws["B3"] = f"{remote_mode} | chunk={remote_chunk_size} | sleep={remote_sleep}s"
        ws["D3"] = "Règle suivi conso"
        ws["E3"] = "Facturée > Estimation > ACM remote > Grid remote > 0"

        headers = [
            "Année",
            "Nb sites",
            "Nb mois",
            "Total suivi conso kWh",
            "Moyenne suivi conso kWh",
            "Points avec donnée",
            "Points attendus",
            "Couverture",
        ]
        ws.append([])
        ws.append(headers)

        header_row = 5
        for cell in ws[header_row]:
            cell.fill = st["header_fill"]
            cell.font = st["header_font"]
            cell.alignment = Alignment(horizontal="center")
            cell.border = st["thin_border"]

        for year in sorted(summary_by_year):
            s = summary_by_year[year]
            ws.append([
                year,
                s["sites_count"],
                s["months_count"],
                round(s["total_kwh"], 3),
                round(s["avg_kwh"], 3),
                s["points_with_data"],
                s["possible_points"],
                s["coverage"],
            ])

        last_row = ws.max_row
        if last_row >= 6:
            for row in ws.iter_rows(min_row=6, max_row=last_row, min_col=1, max_col=8):
                for cell in row:
                    cell.border = st["thin_border"]
                    cell.alignment = Alignment(vertical="center")
            ws[f"D6:D{last_row}"][0][0].number_format = '#,##0.000'
            for row in range(6, last_row + 1):
                ws[f"D{row}"].number_format = '#,##0.000'
                ws[f"E{row}"].number_format = '#,##0.000'
                ws[f"H{row}"].number_format = "0.00%"

            tab = Table(displayName="TableSyntheseConso", ref=f"A5:H{last_row}")
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)

            chart = BarChart()
            chart.title = "Total suivi conso par année"
            chart.y_axis.title = "kWh"
            chart.x_axis.title = "Année"
            data = Reference(ws, min_col=4, min_row=5, max_row=last_row)
            cats = Reference(ws, min_col=1, min_row=6, max_row=last_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, "J5")

        widths = {
            "A": 12, "B": 14, "C": 12, "D": 24, "E": 26,
            "F": 20, "G": 18, "H": 14, "J": 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A6"

    def _create_year_sheet(self, wb: Workbook, year: int, rows: list[dict[str, Any]]) -> None:
        st = self._styles()
        ws = wb.create_sheet(str(year))

        if rows:
            months = rows[0]["months"]
        else:
            months = list(range(1, 13))

        month_headers = [
            f"{calendar.month_abbr[m].capitalize()} {year}" for m in months
        ]

        headers = [
            "Zone",
            "Site ID",
            "Nom site",
            "Typologie",
            "Load (W)",
            *month_headers,
            "Total annuel kWh",
            "Moyenne suivi conso kWh",
            "Mois avec donnée",
            "Mois manquants",
            "Taux couverture",
            "Source dominante",
        ]

        last_col = get_column_letter(len(headers))

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"] = f"Suivi des consommations — {year}"
        ws["A1"].fill = st["title_fill"]
        ws["A1"].font = st["title_font"]
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws["A2"] = "Calcul"
        ws["B2"] = "Moyenne = AVERAGE des mois affichés ; mois sans donnée = 0"
        ws["A2"].font = st["normal_bold"]

        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col_idx, header)
            cell.fill = st["header_fill"]
            cell.font = st["header_font"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = st["thin_border"]

        first_month_col = 6
        last_month_col = first_month_col + len(months) - 1
        total_col = last_month_col + 1
        avg_col = total_col + 1
        data_start_row = header_row + 1

        for row_idx, item in enumerate(rows, start=data_start_row):
            base = [
                item.get("zone") or "",
                item.get("site_id") or "",
                item.get("site_name") or "",
                item.get("typology") or "",
                item.get("load_w") or "",
            ]
            for col_idx, value in enumerate(base, start=1):
                ws.cell(row_idx, col_idx, value)

            for offset, value in enumerate(item["values"]):
                c = ws.cell(row_idx, first_month_col + offset, round(value, 3))
                c.number_format = '#,##0.000'

                # Les mois manquants sont explicitement à 0 et marqués visuellement.
                if item["sources"][offset] == "MISSING":
                    c.fill = st["zero_fill"]

            first_month_letter = get_column_letter(first_month_col)
            last_month_letter = get_column_letter(last_month_col)
            total_letter = get_column_letter(total_col)
            avg_letter = get_column_letter(avg_col)

            ws.cell(row_idx, total_col, f"=SUM({first_month_letter}{row_idx}:{last_month_letter}{row_idx})")
            ws.cell(row_idx, avg_col, f"=AVERAGE({first_month_letter}{row_idx}:{last_month_letter}{row_idx})")
            ws.cell(row_idx, avg_col + 1, item["mois_avec_donnee"])
            ws.cell(row_idx, avg_col + 2, item["mois_manquants"])
            ws.cell(row_idx, avg_col + 3, f"={get_column_letter(avg_col + 1)}{row_idx}/{len(months)}")
            ws.cell(row_idx, avg_col + 4, item["source_dominante"])

        last_row = max(ws.max_row, data_start_row)

        # Formats data
        for row in ws.iter_rows(min_row=data_start_row, max_row=last_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = st["thin_border"]
                cell.alignment = Alignment(vertical="center")

        for col in range(first_month_col, avg_col + 1):
            letter = get_column_letter(col)
            for row in range(data_start_row, last_row + 1):
                ws[f"{letter}{row}"].number_format = '#,##0.000'

        coverage_col = avg_col + 3
        coverage_letter = get_column_letter(coverage_col)
        for row in range(data_start_row, last_row + 1):
            ws[f"{coverage_letter}{row}"].number_format = "0.00%"

        # Mise en évidence : couverture faible
        ws.conditional_formatting.add(
            f"{coverage_letter}{data_start_row}:{coverage_letter}{last_row}",
            CellIsRule(operator="lessThan", formula=["0.8"], fill=st["bad_fill"]),
        )

        # Table Excel
        if rows:
            safe_year = str(year)
            tab = Table(displayName=f"TableSuiviConso{safe_year}", ref=f"A4:{last_col}{last_row}")
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)

        # Largeurs
        widths = {
            1: 12, 2: 14, 3: 28, 4: 22, 5: 12,
        }
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width

        for idx in range(first_month_col, last_month_col + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 13

        ws.column_dimensions[get_column_letter(total_col)].width = 18
        ws.column_dimensions[get_column_letter(avg_col)].width = 24
        ws.column_dimensions[get_column_letter(avg_col + 1)].width = 18
        ws.column_dimensions[get_column_letter(avg_col + 2)].width = 16
        ws.column_dimensions[get_column_letter(avg_col + 3)].width = 16
        ws.column_dimensions[get_column_letter(avg_col + 4)].width = 18

        ws.freeze_panes = "F5"
        ws.auto_filter.ref = f"A4:{last_col}{last_row}"
