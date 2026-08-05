# fuel_tracking/services/suivi_commande_import.py
"""
Import de la feuille "Suivis commande" du fichier Excel mensuel "Commande
FUEL ESCO SENEGAL <mois>" dans FuelSuiviCommandeSite — un snapshot par site
(≈3260 sites), limité aux colonnes mises en évidence en bleu (fond "Accent 1"
du thème, RGB #4472C4) dans le fichier source, sur les 139 colonnes que
compte cette feuille. C'est le même fichier et le même mois que
FuelCommandeSynthese (voir commande_synthese_import.py) — importé en même
temps, depuis le même upload.

Les colonnes bleues ont été identifiées une fois manuellement (conversion
temporaire du .xlsb en .xlsx via LibreOffice pour lire les couleurs de
cellule — pyxlsb ne lit que les valeurs, pas la mise en forme — puis
repérage des cellules dont le remplissage est le thème "Accent 1"). Leurs
positions de colonne (0-indexées) sont fixes ci-dessous : si la structure de
la feuille change d'un mois à l'autre, il faudra les mettre à jour.

Colonnes (0-indexées, telles que renvoyées par pyxlsb/openpyxl) :
  2=ID du site  3=Nom du site  4=Typologie contractuelle  5=Load commandé
  6=Indoor/Outdoor  8=Longitude  9=Batch  12=Typologie facturée
  75=Conso Moy par jour  85=Commande proposée sans marge
  87=Commande proposée avec marge  105=Estimation stock final  132=Typo opérations

Formats supportés : .xlsb (pyxlsb) et .xlsx/.xlsm (openpyxl).
"""
from django.utils import timezone

from fuel_tracking.models import FuelSuiviCommandeSite


SHEET_NAME = "Suivis commande"
HEADER_ROW = 12  # 0-indexé — ligne des libellés de colonnes
DATA_START_ROW = 13  # 0-indexé — première ligne de données (1 site)

COL_SITE_ID = 2
COL_SITE_NAME = 3
COL_TYPOLOGIE_CONTRACTUELLE = 4
COL_LOAD_COMMANDE = 5
COL_INDOOR_OUTDOOR = 6
COL_LONGITUDE = 8
COL_BATCH = 9
COL_TYPOLOGIE_FACTUREE = 12
COL_CONSO_MOY_JOUR = 75
COL_COMMANDE_SANS_MARGE = 85
COL_COMMANDE_AVEC_MARGE = 87
COL_ESTIMATION_STOCK_FINAL = 105
COL_TYPO_OPERATIONS = 132


class SuiviCommandeImportError(Exception):
    pass


def _read_sheet_grid(path: str) -> dict:
    """Retourne {(row, col): value} pour la feuille SHEET_NAME, quel que soit
    le format (.xlsb via pyxlsb, .xlsx/.xlsm via openpyxl)."""
    if path.lower().endswith(".xlsb"):
        from pyxlsb import open_workbook

        with open_workbook(path) as wb:
            if SHEET_NAME not in wb.sheets:
                raise SuiviCommandeImportError(
                    f"Feuille '{SHEET_NAME}' introuvable. Feuilles disponibles : {wb.sheets}"
                )
            sheet = wb.get_sheet(SHEET_NAME)
            grid = {}
            for row in sheet.rows():
                for cell in row:
                    if cell.v is not None:
                        grid[(cell.r, cell.c)] = cell.v
            return grid

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SuiviCommandeImportError(
            f"Feuille '{SHEET_NAME}' introuvable. Feuilles disponibles : {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    grid = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid[(cell.row - 1, cell.column - 1)] = cell.value
    wb.close()
    return grid


def _str_or_none(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_suivi_commande(path: str, month_year: str, log=None):
    """
    Parse le fichier à `path` et retourne une liste d'instances
    FuelSuiviCommandeSite non sauvegardées (une par site). `log` est un
    callable optionnel pour tracer la progression.
    """
    log = log or (lambda msg: None)

    grid = _read_sheet_grid(path)
    if not grid:
        raise SuiviCommandeImportError("La feuille Suivis commande est vide.")

    max_row = max(r for r, _ in grid.keys())
    filename = path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]

    objects = []
    row = DATA_START_ROW
    while row <= max_row:
        site_id = grid.get((row, COL_SITE_ID))
        if site_id is None or str(site_id).strip() == "":
            row += 1
            continue

        objects.append(FuelSuiviCommandeSite(
            month_year=month_year,
            site_id=str(site_id).strip(),
            site_name=_str_or_none(grid.get((row, COL_SITE_NAME))),
            typologie_contractuelle=_str_or_none(grid.get((row, COL_TYPOLOGIE_CONTRACTUELLE))),
            load_commande=grid.get((row, COL_LOAD_COMMANDE)) or 0,
            indoor_outdoor=_str_or_none(grid.get((row, COL_INDOOR_OUTDOOR))),
            longitude=grid.get((row, COL_LONGITUDE)),
            batch=_str_or_none(grid.get((row, COL_BATCH))),
            typologie_facturee=_str_or_none(grid.get((row, COL_TYPOLOGIE_FACTUREE))),
            conso_moy_jour_l=grid.get((row, COL_CONSO_MOY_JOUR)) or 0,
            commande_sans_marge_l=grid.get((row, COL_COMMANDE_SANS_MARGE)) or 0,
            commande_avec_marge_l=grid.get((row, COL_COMMANDE_AVEC_MARGE)) or 0,
            estimation_stock_final_l=grid.get((row, COL_ESTIMATION_STOCK_FINAL)) or 0,
            typo_operations=_str_or_none(grid.get((row, COL_TYPO_OPERATIONS))),
            source_filename=filename,
            imported_at=timezone.now(),
        ))
        row += 1

    log(f"  Suivis commande : {len(objects)} sites détectés")

    if not objects:
        raise SuiviCommandeImportError("Aucun site détecté dans la feuille Suivis commande.")

    return objects


def import_suivi_commande_file(path: str, month_year: str, log=None) -> int:
    """Parse + écrit en base. Ne remplace que les lignes du mois donné."""
    objects = parse_suivi_commande(path, month_year=month_year, log=log)

    FuelSuiviCommandeSite.objects.filter(month_year=month_year).delete()
    FuelSuiviCommandeSite.objects.bulk_create(objects, batch_size=1000)

    return len(objects)
