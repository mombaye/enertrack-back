# fuel_tracking/services/commande_synthese_import.py
"""
Logique partagée d'import de la feuille "Synthèse Commande" du fichier Excel
mensuel "Commande FUEL ESCO SENEGAL <mois> <année>" — utilisée à la fois par
la commande de gestion import_commande_synthese et par l'endpoint d'upload
FuelCommandeSyntheseImportView (bouton "Importer" du frontend).

Import brut, sans recalcul : chaque ligne du tableau source (par catégorie/
batch, puis par typologie facturée) est reprise telle quelle, colonnes du
mois courant / mois précédent / écart déjà calculées dans le fichier.

La feuille est organisée en deux blocs verticaux détectés par libellé
(robuste à l'ajout/suppression de lignes intermédiaires d'un mois à l'autre) :
  - bloc "CATEGORIE"          : 2 lignes de sous-en-tête, puis données
  - bloc "Typologie facturée" : 2 lignes de sous-en-tête, puis données
Chaque bloc s'arrête à la ligne "TOTAL COMMANDE" (incluse). Les lignes sans
libellé (col B vide, simples séparateurs visuels) sont ignorées.

Colonnes (0-indexées, telles que renvoyées par pyxlsb/openpyxl) :
  B(1)=libellé  C(2)=nb sites courant   D(3)=cmde normale courant
  E(4)=cmde hivernale courant           F(5)=total courant
  G(6)=nb sites précédent               H(7)=cmde normale précédent
  I(8)=cmde hivernale précédent         J(9)=total précédent
  K(10)=écart sites   L(11)=écart qté   N(13)=commentaires

Le mois courant et le mois précédent sont fournis explicitement par
l'appelant (obligatoire côté API — voir FuelCommandeSyntheseImportView) :
la détection automatique depuis les libellés du fichier ("Août" / "Juillet")
s'est révélée peu fiable (feuille pas toujours structurée pareil d'un mois
à l'autre), d'où le choix de laisser l'utilisateur les préciser lui-même au
moment de l'upload plutôt que de deviner. La commande de gestion garde une
détection automatique en repli, pratique pour les imports en ligne de
commande, mais accepte aussi --month-year/--prev-month-year en override.

Formats supportés : .xlsb (pyxlsb) et .xlsx/.xlsm (openpyxl) — le fichier
client est habituellement un .xlsb, mais on reste robuste à un export .xlsx.
"""
import re

from django.utils import timezone

from fuel_tracking.models import FuelCommandeSynthese


SHEET_NAME = "Synthèse Commande"

BLOCK_HEADERS = {
    "CATEGORIE": FuelCommandeSynthese.GroupType.CATEGORIE,
    "TYPOLOGIE FACTUREE": FuelCommandeSynthese.GroupType.TYPOLOGIE,
    "TYPOLOGIE FACTURÉE": FuelCommandeSynthese.GroupType.TYPOLOGIE,
}

COL_LABEL = 1
COL_NB_SITES = 2
COL_CMD_NORMALE = 3
COL_CMD_HIVERNALE = 4
COL_TOTAL = 5
COL_NB_SITES_PREV = 6
COL_CMD_NORMALE_PREV = 7
COL_CMD_HIVERNALE_PREV = 8
COL_TOTAL_PREV = 9
COL_ECART_SITES = 10
COL_ECART_QTE = 11
COL_COMMENTAIRES = 13

FRENCH_MONTHS = {
    "janvier": 1, "fevrier": 2, "février": 2, "mars": 3, "avril": 4,
    "mai": 5, "juin": 6, "juillet": 7, "aout": 8, "août": 8,
    "septembre": 9, "octobre": 10, "novembre": 11, "decembre": 12, "décembre": 12,
}


class CommandeSyntheseImportError(Exception):
    pass


def month_label_to_yyyymm(label, year):
    key = str(label).strip().lower()
    month = FRENCH_MONTHS.get(key)
    if not month:
        return None
    return f"{year:04d}-{month:02d}"


def derive_prev_month_year(yyyymm):
    """Mois précédent par défaut (repli CLI uniquement) — yyyymm - 1 mois."""
    year, month = (int(x) for x in yyyymm.split("-"))
    if month == 1:
        return f"{year - 1:04d}-12"
    return f"{year:04d}-{month - 1:02d}"


MONTH_YEAR_RE = re.compile(r"^\d{4}-\d{2}$")


def validate_month_year(value: str, field_label: str) -> str:
    if not value or not MONTH_YEAR_RE.match(value):
        raise CommandeSyntheseImportError(f"{field_label} invalide (attendu YYYY-MM) : {value!r}")
    return value


def _read_sheet_grid(path: str) -> dict:
    """Retourne {(row, col): value} pour la feuille SHEET_NAME, quel que soit
    le format (.xlsb via pyxlsb, .xlsx/.xlsm via openpyxl)."""
    if path.lower().endswith(".xlsb"):
        from pyxlsb import open_workbook

        with open_workbook(path) as wb:
            if SHEET_NAME not in wb.sheets:
                raise CommandeSyntheseImportError(
                    f"Feuille '{SHEET_NAME}' introuvable. Feuilles disponibles : {wb.sheets}"
                )
            sheet = wb.get_sheet(SHEET_NAME)
            grid = {}
            for row in sheet.rows():
                for cell in row:
                    if cell.v is not None:
                        grid[(cell.r, cell.c)] = cell.v
            return grid

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise CommandeSyntheseImportError(
            f"Feuille '{SHEET_NAME}' introuvable. Feuilles disponibles : {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    grid = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid[(cell.row - 1, cell.column - 1)] = cell.value
    wb.close()
    return grid


def parse_commande_synthese(path: str, month_year: str | None = None, prev_month_year: str | None = None, log=None):
    """
    Parse le fichier à `path` et retourne (objects, month_year, prev_month_year)
    — `objects` est une liste d'instances FuelCommandeSynthese non sauvegardées.
    `log` est un callable optionnel (ex: self.stdout.write) pour tracer la
    progression ligne par ligne (utilisé par la commande de gestion).

    `month_year`/`prev_month_year` doivent normalement être fournis par
    l'appelant (voir docstring module) ; à défaut, on retente une détection
    automatique depuis le fichier (repli CLI).
    """
    log = log or (lambda msg: None)

    grid = _read_sheet_grid(path)
    if not grid:
        raise CommandeSyntheseImportError("La feuille Synthèse Commande est vide.")

    max_row = max(r for r, _ in grid.keys())
    filename = path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]

    if month_year:
        validate_month_year(month_year, "Mois courant")
    else:
        year_match = re.search(r"(20\d{2})", filename)
        year = int(year_match.group(1)) if year_match else timezone.now().year
        month_label = grid.get((2, COL_NB_SITES))  # ligne sous-en-tête ("Août")
        month_year = month_label_to_yyyymm(month_label, year)
        if not month_year:
            raise CommandeSyntheseImportError(
                f"Impossible de déduire le mois courant (libellé trouvé : {month_label!r})."
            )

    if prev_month_year:
        prev_ym = validate_month_year(prev_month_year, "Mois précédent")
    else:
        prev_ym = derive_prev_month_year(month_year)
    log(f"  Mois courant : {month_year}  |  Mois précédent : {prev_ym}")

    objects = []
    row = 1
    while row <= max_row:
        label = grid.get((row, COL_LABEL))
        key = str(label).strip().upper() if label else None
        if key in BLOCK_HEADERS:
            group_type = BLOCK_HEADERS[key]
            log(f"\n  Bloc détecté : {label!r} (ligne {row}) -> {group_type}")
            data_row = row + 3  # saute l'en-tête + les 2 lignes de sous-en-tête
            order_index = 0
            while data_row <= max_row:
                row_label = grid.get((data_row, COL_LABEL))
                if row_label is None or str(row_label).strip() == "":
                    data_row += 1
                    continue

                clean_label = str(row_label).strip()
                is_total = "total" in clean_label.lower()
                is_block_end = clean_label == "TOTAL COMMANDE"

                objects.append(FuelCommandeSynthese(
                    month_year=month_year,
                    prev_month_year=prev_ym,
                    group_type=group_type,
                    order_index=order_index,
                    label=clean_label,
                    is_total_row=is_total,
                    nb_sites=grid.get((data_row, COL_NB_SITES)) or 0,
                    commande_normale_l=grid.get((data_row, COL_CMD_NORMALE)) or 0,
                    commande_hivernale_l=grid.get((data_row, COL_CMD_HIVERNALE)) or 0,
                    total_l=grid.get((data_row, COL_TOTAL)) or 0,
                    nb_sites_prev=grid.get((data_row, COL_NB_SITES_PREV)) or 0,
                    commande_normale_prev_l=grid.get((data_row, COL_CMD_NORMALE_PREV)) or 0,
                    commande_hivernale_prev_l=grid.get((data_row, COL_CMD_HIVERNALE_PREV)) or 0,
                    total_prev_l=grid.get((data_row, COL_TOTAL_PREV)) or 0,
                    ecart_sites=grid.get((data_row, COL_ECART_SITES)) or 0,
                    ecart_qte_l=grid.get((data_row, COL_ECART_QTE)) or 0,
                    commentaires=(str(grid.get((data_row, COL_COMMENTAIRES))).strip() or None)
                    if grid.get((data_row, COL_COMMENTAIRES)) is not None else None,
                    source_filename=filename,
                    imported_at=timezone.now(),
                ))
                order_index += 1
                log(f"    {'[TOTAL] ' if is_total else ''}{row_label}")

                if is_block_end:
                    row = data_row
                    break
                data_row += 1
        row += 1

    if not objects:
        raise CommandeSyntheseImportError("Aucune ligne détectée — vérifie la structure du fichier.")

    return objects, month_year, prev_ym


def import_commande_synthese_file(
    path: str, month_year: str | None = None, prev_month_year: str | None = None, log=None
) -> tuple[int, str]:
    """
    Parse + écrit en base. Ne remplace QUE les lignes du mois détecté dans ce
    fichier (resolved_month_year) — les autres mois déjà importés restent
    intacts, chacun reste consultable via ?month=YYYY-MM. Réimporter le même
    mois (ex: nouveau brouillon) remplace proprement ses lignes, sans toucher
    aux autres mois stockés.
    """
    objects, resolved_month_year, _ = parse_commande_synthese(
        path, month_year=month_year, prev_month_year=prev_month_year, log=log
    )

    FuelCommandeSynthese.objects.filter(month_year=resolved_month_year).delete()
    FuelCommandeSynthese.objects.bulk_create(objects, batch_size=500)

    return len(objects), resolved_month_year
