# fuel_tracking/services/xlsb_utils.py
"""
Lecture partagée d'un classeur Excel mensuel (.xlsb/.xlsx/.xlsm) — ouvre le
fichier UNE SEULE FOIS pour en extraire plusieurs feuilles, au lieu de le
rouvrir/redécompresser une fois par feuille. L'upload combine "Synthèse
Commande" + "Suivis commande" depuis le même fichier .xlsb (peut peser
plusieurs Mo) : lire une seule fois raccourcit d'autant le traitement d'un
import déjà pénalisé par un upload volumineux sur un réseau instable.
"""


def read_workbook_grids(path: str, sheet_names: list[str]) -> dict[str, dict]:
    """Retourne {sheet_name: {(row, col): value}} pour chaque feuille demandée
    présente dans le classeur (les feuilles absentes sont simplement omises —
    à l'appelant de gérer ce cas, voir parse_commande_synthese/parse_suivi_commande)."""
    grids: dict[str, dict] = {}

    if path.lower().endswith(".xlsb"):
        from pyxlsb import open_workbook

        with open_workbook(path) as wb:
            for name in sheet_names:
                if name not in wb.sheets:
                    continue
                sheet = wb.get_sheet(name)
                grid = {}
                for row in sheet.rows():
                    for cell in row:
                        if cell.v is not None:
                            grid[(cell.r, cell.c)] = cell.v
                grids[name] = grid
        return grids

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        grid = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    grid[(cell.row - 1, cell.column - 1)] = cell.value
        grids[name] = grid
    wb.close()
    return grids
