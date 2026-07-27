# fuel_tracking/export.py
"""
Construction du classeur Excel "Suivi Carburant", au format du modèle client
(Suivi_Carburant_AMELIORE.xlsx) : mêmes noms/ordre de feuilles, en-têtes
fusionnés multi-lignes avec bandes de section colorées, volets figés, mise en
forme conditionnelle sur les colonnes d'écart.

Les cellules contiennent les valeurs déjà calculées par l'app (pas de formules
Excel live) — tout le calcul est fait côté serveur, reproduire des formules
serait redondant. Les colonnes de référentiel site (typologie, GE, cuves)
reprennent uniquement les champs simples déjà disponibles côté serveur — les
enrichissements avancés (courbe de conso théorique, cible CPH interpolée,
facteur de charge) restent un calcul d'affichage propre au frontend et ne sont
pas reproduits ici.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Palette — harmonisée bleu-900/blanc Camusat
# ─────────────────────────────────────────────────────────────────────────────

NAVY = "0B1F4D"
BLUE_700 = "123C8C"
BLUE_600 = "1A56C4"
BLUE_500 = "2464D6"
SLATE = "475569"
WHITE = "FFFFFF"
BLUE_LIGHT = "E4EFFE"
SLATE_LIGHT = "F1F5F9"
RED = "DC2626"
RED_LIGHT = "FEE2E2"
ORANGE = "D97706"
ORANGE_LIGHT = "FEF3C7"
GREEN = "059669"
GREEN_LIGHT = "D1FAE5"

THIN = Side(style="thin", color="D6DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _title_row(ws, ncols: int, text: str):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    cell.fill = _fill(NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _write_grouped_header(ws, groups: list[dict], start_row: int = 2):
    """
    groups: [{"label": str, "color": hex, "columns": [str, ...]}, ...]
    Écrit la ligne de bandes de section (fusionnées) puis la ligne d'en-têtes
    de colonnes juste en dessous, avec la même couleur de fond par groupe.
    """
    col = 1
    band_row = start_row
    header_row = start_row + 1

    for group in groups:
        span = len(group["columns"])
        if span > 1:
            ws.merge_cells(start_row=band_row, start_column=col, end_row=band_row, end_column=col + span - 1)
        band_cell = ws.cell(row=band_row, column=col, value=group["label"])
        band_cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        band_cell.fill = _fill(group["color"])
        band_cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, header in enumerate(group["columns"]):
            c = ws.cell(row=header_row, column=col + i, value=header)
            c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
            c.fill = _fill(group["color"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER

        col += span

    ws.row_dimensions[band_row].height = 20
    ws.row_dimensions[header_row].height = 34
    return header_row


def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, w / 7)


def _write_rows(ws, start_row: int, rows: list[list], num_cols: int, zebra: str = SLATE_LIGHT):
    for r, row in enumerate(rows):
        excel_row = start_row + r
        fill = _fill(zebra) if r % 2 else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=excel_row, column=c, value=row[c - 1] if c - 1 < len(row) else None)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            if fill:
                cell.fill = fill
            if isinstance(row[c - 1] if c - 1 < len(row) else None, (int, float)):
                cell.number_format = "#,##0"
    return start_row + len(rows)


def _pct_conditional_formatting(ws, col_letter: str, first_row: int, last_row: int):
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThan", formula=["10"], fill=_fill(RED_LIGHT))
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="lessThan", formula=["-10"], fill=_fill(RED_LIGHT))
    )


# ─────────────────────────────────────────────────────────────────────────────
# Helpers d'extraction (référentiel site — sous-ensemble simple, sans les
# enrichissements calculés côté frontend uniquement)
# ─────────────────────────────────────────────────────────────────────────────

def _g(d: dict | None, *keys, default=None):
    for k in keys:
        if d and d.get(k) not in (None, ""):
            return d.get(k)
    return default


def _site_typology(row: dict):
    site_ref = row.get("site_ref") or {}
    enoc_ref = row.get("enoc_site_ref") or {}
    return _g(site_ref, "billing_typology") or _g(enoc_ref, "typology_contractual", "new_typo", "typo_simple") or "—"


def _region(row: dict):
    enoc_ref = row.get("enoc_site_ref") or {}
    return row.get("zone_label") or row.get("zone") or enoc_ref.get("region") or "—"


def _batch(row: dict):
    site_ref = row.get("site_ref") or {}
    enoc_ref = row.get("enoc_site_ref") or {}
    return _g(site_ref, "batch_operational") or _g(enoc_ref, "batch_operational", "batch") or "—"


def _conf(row: dict):
    site_ref = row.get("site_ref") or {}
    enoc_ref = row.get("enoc_site_ref") or {}
    return _g(site_ref, "configuration") or _g(enoc_ref, "ongrid_offgrid", "indoor_outdoor_after_passive") or "—"


def _priority(row: dict):
    return _g(row.get("enoc_site_ref"), "priority", default="—")


def _load(row: dict):
    site_ref = row.get("site_ref") or {}
    enoc_ref = row.get("enoc_site_ref") or {}
    return site_ref.get("analysis_load") or enoc_ref.get("new_load_contract_v2") or enoc_ref.get("new_load") or enoc_ref.get("load")


# ─────────────────────────────────────────────────────────────────────────────
# Feuilles
# ─────────────────────────────────────────────────────────────────────────────

def _build_dashboard_sheet(wb, month: str, kpis: dict):
    ws = wb.create_sheet("DASHBOARD")
    _title_row(ws, 6, f"DASHBOARD CARBURANT — Synthèse Mensuelle Globale ({month})")

    labels = [
        ("Sites totaux", kpis.get("total_sites")),
        ("Sites eFMS", kpis.get("efms_sites")),
        ("Sites ENOC", kpis.get("enoc_sites")),
        ("Fuel commandé (L)", kpis.get("fuel_order_l")),
        ("Fuel livré (L)", kpis.get("fuel_deli_l")),
        ("Conso réelle (L)", kpis.get("fuel_conso_l")),
        ("Quantité ajoutée ENOC (L)", kpis.get("enoc_quantity_added_liters")),
        ("Mouvements ENOC", kpis.get("movements_count")),
        ("Sites OK", kpis.get("ok")),
        ("Sites à suivre", kpis.get("warning")),
        ("Sites NOK", kpis.get("nok")),
        ("Écart livré vs ENOC (L)", kpis.get("gap_deli_vs_enoc_l")),
    ]
    row = 3
    for label, value in labels:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(name="Arial", size=10, bold=True, color=NAVY)
        lc.fill = _fill(BLUE_LIGHT)
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(name="Calibri", size=11, bold=True)
        vc.fill = _fill(SLATE_LIGHT)
        if isinstance(value, (int, float)):
            vc.number_format = "#,##0"
        row += 1

    _set_widths(ws, [260, 140])


JOURNAL_GROUPS = [
    {"label": "IDENTIFICATION", "color": NAVY, "columns": ["Site ID", "N° Ticket FMS", "Type d'action"]},
    {"label": "TRAÇABILITÉ", "color": BLUE_700, "columns": ["Date", "Responsable", "Source (Site/Dépôt)"]},
    {"label": "MESURES PHYSIQUES", "color": BLUE_600, "columns": ["Qté initiale site (L)", "Qté transférée (L)", "Qté finale site (L)", "Méthode Jaugeage"]},
    {"label": "CONTRÔLEUR / RMS", "color": BLUE_500, "columns": ["DG RH lu contrôleur (h)", "Qté init. RMS (L)", "Qté fin. RMS (L)", "RMS DG RH (h)"]},
    {"label": "LIVRAISON BL", "color": SLATE, "columns": ["N° Bon de Livraison", "Qté BL (L)"]},
    {"label": "VALIDATION", "color": NAVY, "columns": ["Validé Par", "Statut"]},
    {"label": "CONTRÔLE AUTO", "color": RED, "columns": ["Écart BL/Mesuré (L)", "Écart BL (%)", "Balance Check"]},
]


def _build_journal_sheet(wb, journal_rows: list[dict]):
    ws = wb.create_sheet("JOURNAL_RAVITAILLEMENT")
    ncols = sum(len(g["columns"]) for g in JOURNAL_GROUPS)
    _title_row(ws, ncols, "JOURNAL DES MOUVEMENTS CARBURANT — Dépotage / Transfert / Ajout")
    header_row = _write_grouped_header(ws, JOURNAL_GROUPS)

    rows = []
    for r in journal_rows:
        level_before = r.get("level_before")
        level_after = r.get("level_after")
        qty = r.get("quantity_added_liters")
        bl_qty = r.get("delivery_note_quantity_liters")
        ecart_bl = None
        ecart_bl_pct = None
        if bl_qty not in (None, "") and level_before not in (None, ""):
            try:
                ecart_bl = float(bl_qty) - float(level_before)
                ecart_bl_pct = (ecart_bl / float(bl_qty) * 100) if float(bl_qty) else None
            except (TypeError, ValueError):
                pass
        rows.append([
            r.get("site_id"), r.get("request_code"), r.get("operation_type"),
            r.get("operation_date"), r.get("done_by") or r.get("created_by") or r.get("technician_name"), r.get("supplier") or "—",
            level_before, qty, level_after, r.get("gauging_method"),
            r.get("hour_meter_after"), r.get("rms_level_before"), r.get("rms_level_after"), None,
            r.get("delivery_note_number"), bl_qty,
            r.get("validated_by"), r.get("status"),
            ecart_bl, round(ecart_bl_pct, 1) if ecart_bl_pct is not None else None, ("OK" if ecart_bl is not None and abs(ecart_bl) <= 1 else ("Écart" if ecart_bl is not None else "—")),
        ])

    data_start = header_row + 1
    last_row = _write_rows(ws, data_start, rows, ncols)
    if rows:
        pct_col = get_column_letter(ncols - 1)  # "Écart BL (%)"
        _pct_conditional_formatting(ws, pct_col, data_start, last_row - 1)

    _set_widths(ws, [90] + [130] * (ncols - 1))
    ws.freeze_panes = f"A{data_start}"


CONSO_GROUPS = [
    {"label": "SITE", "color": NAVY, "columns": ["Site ID", "Site Name"]},
    {"label": "RÉFÉRENTIEL SITE", "color": BLUE_700, "columns": ["Région", "Batch", "Typo Facturée", "Conf", "Priorité", "Puissance (W)"]},
    {"label": "CIBLES", "color": BLUE_600, "columns": ["Target Aktivco (L/mois)"]},
    {"label": "DONNÉES MOIS PRÉCÉDENT", "color": BLUE_500, "columns": ["RH Mois Précédent (h)", "Ravitaillement (L)", "Ponction (L)"]},
    {"label": "DONNÉES MOIS EN COURS", "color": SLATE, "columns": ["RH Final (h)", "RH Delta (h)"]},
    {"label": "CONSOMMATION CALCULÉE", "color": BLUE_700, "columns": ["Conso Réelle (L)", "Conso Théorique (L)", "CPH Réel (L/h)"]},
    {"label": "ÉCARTS & ALERTES", "color": RED, "columns": ["Écart vs Target (L)", "Écart vs Target (%)", "Statut NOK/OK"]},
]


def _build_conso_sheet(wb, monthly_rows: list[dict]):
    ws = wb.create_sheet("CONSO_MENSUELLE")
    ncols = sum(len(g["columns"]) for g in CONSO_GROUPS)
    _title_row(ws, ncols, "SUIVI CONSOMMATION MENSUELLE PAR SITE — avec contrôles automatiques")
    header_row = _write_grouped_header(ws, CONSO_GROUPS)

    rows = []
    for r in monthly_rows:
        efms = r.get("efms") or {}
        enoc = r.get("enoc") or {}
        gaps = r.get("gaps") or {}
        ravitaillement = (enoc.get("refueling_liters") or 0) + (enoc.get("ajout_in_liters") or 0)
        ponction = enoc.get("prelevement_out_liters") or 0
        gap_pct = gaps.get("deli_vs_enoc_pct")
        rows.append([
            r.get("site_id"), r.get("site_name"),
            _region(r), _batch(r), _site_typology(r), _conf(r), _priority(r), _load(r),
            enoc.get("monthly_target_liters"),
            efms.get("rh_initial_hours"), ravitaillement or None, ponction or None,
            efms.get("rh_hours"), efms.get("rh_delta_hours"),
            efms.get("fuel_conso_l"), efms.get("fuel_deli_l"), efms.get("cph_l_per_hour"),
            gaps.get("deli_vs_enoc_l"), round(gap_pct, 1) if gap_pct is not None else None,
            (gaps.get("status") or {}).get("label"),
        ])

    data_start = header_row + 1
    last_row = _write_rows(ws, data_start, rows, ncols)
    if rows:
        pct_col_index = ncols - 1  # "Écart vs Target (%)" is second-to-last column
        _pct_conditional_formatting(ws, get_column_letter(pct_col_index), data_start, last_row - 1)

    _set_widths(ws, [90, 170] + [120] * (ncols - 2))
    ws.freeze_panes = f"A{data_start}"


def _build_stock_depot_sheet(wb):
    ws = wb.create_sheet("STOCK_DÉPÔT")
    headers = ["Date", "Fournisseur", "N° BL", "Entrée (L)", "Sortie vers site", "Site Destinataire", "Quantité sortie (L)", "Solde Dépôt (L)", "Responsable", "Commentaire"]
    _title_row(ws, len(headers), "SUIVI STOCK DÉPÔT CENTRAL — Entrées Fournisseurs / Sorties vers Sites")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        c.fill = _fill(BLUE_700)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _set_widths(ws, [110] * len(headers))
    ws.freeze_panes = "A3"


def _build_cph_sheet(wb, cph_data: list[dict]):
    ws = wb.create_sheet("CPH")
    _title_row(ws, 12, "CPH — Matrice de consommation par famille moteur")
    row = 3
    for engine in cph_data:
        c = ws.cell(row=row, column=1, value=f"Moteur {engine.get('engine_family')}")
        c.font = Font(name="Arial", size=11, bold=True, color=NAVY)
        row += 1
        pct_labels = sorted({k for r in engine.get("rows", []) for k in r.get("values", {}).keys()}, key=lambda x: float(x))
        ws.cell(row=row, column=1, value="KVA")
        for j, pct in enumerate(pct_labels, start=2):
            hc = ws.cell(row=row, column=j, value=f"{float(pct) * 100:.0f}%")
            hc.font = Font(bold=True)
        row += 1
        for r in engine.get("rows", []):
            ws.cell(row=row, column=1, value=r.get("dg_capacity_kva"))
            for j, pct in enumerate(pct_labels, start=2):
                ws.cell(row=row, column=j, value=r.get("values", {}).get(pct))
            row += 1
        row += 1
    _set_widths(ws, [80] * 12)


def _build_ref_sites_sheet(wb, monthly_rows: list[dict]):
    ws = wb.create_sheet("REF_SITES")
    headers = ["Site ID", "Site Name", "Région", "Batch", "Typo Facturée", "Conf", "Priorité", "Puissance (W)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        [r.get("site_id"), r.get("site_name"), _region(r), _batch(r), _site_typology(r), _conf(r), _priority(r), _load(r)]
        for r in monthly_rows
    ]
    _write_rows(ws, 2, rows, len(headers))
    _set_widths(ws, [90, 170, 110, 130, 130, 70, 90, 110])
    ws.freeze_panes = "A2"


LISTES = {
    "type_action": ["Dépotage", "Transfert inter-sites", "Ajout manuel", "Prélèvement", "Correction"],
    "statut": ["En cours", "Soldé", "Partiel", "Annulé"],
    "methode_jauge": ["Jauge manuelle", "RMS", "Estimation"],
    "priorite": ["P1", "P2", "P3", "P4", "P5"],
}


def _build_listes_sheet(wb):
    ws = wb.create_sheet("LISTES")
    for col, (key, values) in enumerate(LISTES.items(), start=1):
        c = ws.cell(row=1, column=col, value=key)
        c.font = Font(bold=True, size=11)
        for i, v in enumerate(values, start=2):
            ws.cell(row=i, column=col, value=v)
    _set_widths(ws, [140] * len(LISTES))


# ─────────────────────────────────────────────────────────────────────────────
# Entrée principale
# ─────────────────────────────────────────────────────────────────────────────

def build_fuel_tracking_workbook(month: str, monthly_rows: list[dict], journal_rows: list[dict], kpis: dict, cph_data: list[dict]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # feuille par défaut

    _build_dashboard_sheet(wb, month, kpis)
    _build_journal_sheet(wb, journal_rows)
    _build_conso_sheet(wb, monthly_rows)
    _build_stock_depot_sheet(wb)
    _build_cph_sheet(wb, cph_data)
    _build_ref_sites_sheet(wb, monthly_rows)
    _build_listes_sheet(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
