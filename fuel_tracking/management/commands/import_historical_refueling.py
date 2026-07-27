# fuel_tracking/management/commands/import_historical_refueling.py
"""
Importe l'historique de ravitaillement depuis le fichier de suivi opérationnel
"SUIVI REFUELING MENSUEL - ok.xlsx" (feuilles TRACKER REFUELING <mois>) dans
FuelEnocMovement — PAS dans MongoDB ENOC (leur schéma fuel_operations exige
un fuel_requests lié, on ne fabrique pas de faux workflow de validation).

Structure source : une ligne par site, avec jusqu'à 16 "emplacements" de
ravitaillement répétés en blocs de 9 colonnes (Refueling date, Quantity
before [en réalité une hauteur cm, libellé trompeur], Quantity refueled [L],
Current Height, Index generator, CPH, Technicien, Origine Fuel, Comment).

Règles de mapping (validées avec l'utilisateur) :
  - Origine Fuel = "Total"  → operation_type = TOTAL_CARD (carte carburant)
  - Origine Fuel = "Ponction", quantité > 0  → ce site est destinataire (PONCTION)
  - Origine Fuel = "Ponction", quantité < 0  → ce site est SOURCE d'une ponction
    sortante. On tente de retrouver le(s) site(s) destinataire(s) cité(s) en
    texte libre dans le commentaire (ex: "ravitailler KEUR_SAMBA_GAYE" ou
    "THILLA_GARANG 100L, MEDINA_SABAKH 100L") et on crée le mouvement sous le
    site DESTINATAIRE (avec ponction.source_site_id = ce site), exactement
    comme le fait le vrai schéma ENOC. Si aucun site n'est reconnu avec
    confiance, on importe quand même sous le site source en quantité négative,
    marqué pour revue manuelle (rien n'est perdu, rien n'est deviné à tort).

Usage:
    docker compose exec web python manage.py import_historical_refueling --dry-run
    docker compose exec web python manage.py import_historical_refueling --sheets "TRACKER REFUELING JUIN"
    docker compose exec web python manage.py import_historical_refueling
"""
import re
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from fuel_tracking.models import FuelEnocMovement

FILE_PATH = "data_imports/SUIVI REFUELING MENSUEL - ok.xlsx"

SHEETS = {
    "TRACKER REFUELING MARS": 3,
    "TRACKER REFUELING AVRIL": 4,
    "TRACKEUR REFUELING MAI": 5,
    "TRACKER REFUELING JUIN": 6,
}
YEAR = 2026
HEADER_ROW = 4
SOURCE_SYSTEM = "ENOC_HISTORICAL"


def _normalize(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", "", text)
    return text


def _parse_cm(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).lower().replace("cm", "").replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _parse_hour_meter(value) -> Decimal | None:
    """
    'Index generator' est hétérogène : un nombre pur (724.3, 17181.53 — index
    compteur en heures), un format "5287h47"/"2466H"/"1306h 58m" (heures +
    minutes), ou du texte non exploitable ("Panel HS", "Non renseigné").
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    s = str(value).strip()
    m = re.match(r"^(\d+)\s*[hH]\s*(\d+)?\s*m?", s)
    if m:
        hours = int(m.group(1))
        minutes = int(m.group(2)) if m.group(2) else 0
        return Decimal(hours) + Decimal(minutes) / Decimal(60)

    try:
        return Decimal(s.replace(",", "."))
    except InvalidOperation:
        return None


def _dec(value) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _to_aware(dt) -> datetime | None:
    if not dt or not isinstance(dt, datetime):
        return None
    return timezone.make_aware(dt, timezone.get_current_timezone()) if timezone.is_naive(dt) else dt


class SiteResolver:
    """Résout un nom de site (texte libre) vers un site_id connu."""

    def __init__(self):
        from core.models import Site

        self.by_norm_name = {}
        self.by_norm_id = {}
        for site_id, name in Site.objects.values_list("site_id", "name"):
            if name:
                self.by_norm_name[_normalize(name)] = site_id
            if site_id:
                self.by_norm_id[_normalize(site_id)] = site_id

    def resolve(self, token: str) -> str | None:
        norm = _normalize(token)
        if not norm or len(norm) < 4:
            return None
        if norm in self.by_norm_id:
            return self.by_norm_id[norm]
        if norm in self.by_norm_name:
            return self.by_norm_name[norm]

        # Match approximatif (orthographe phonétique dans les commentaires libres) —
        # on n'accepte que si un seul nom dépasse nettement le seuil de confiance.
        import difflib

        candidates = list(self.by_norm_name.keys())
        close = difflib.get_close_matches(norm, candidates, n=2, cutoff=0.82)
        if len(close) == 1:
            return self.by_norm_name[close[0]]
        return None

    def parse_destinations(self, comment: str) -> list[tuple[str, float | None]]:
        """
        Extrait les (site_id, quantité|None) cités dans un commentaire de
        ponction sortante. Retourne une liste vide si rien de fiable trouvé.
        """
        if not comment or not isinstance(comment, str):
            return []

        m = re.search(r"ravitailler\s+(.+)", comment, re.IGNORECASE)
        tail = m.group(1) if m else comment

        tail = tail.replace(" et ", ",")
        tokens = [t.strip() for t in tail.split(",") if t.strip()]

        results = []
        for token in tokens:
            qty_match = re.search(r"(\d+(?:[.,]\d+)?)\s*L\b", token, re.IGNORECASE)
            qty = float(qty_match.group(1).replace(",", ".")) if qty_match else None
            name_part = token[: qty_match.start()].strip() if qty_match else token
            site_id = self.resolve(name_part)
            if site_id:
                results.append((site_id, qty))
        return results


class Command(BaseCommand):
    help = "Importe l'historique refueling (xlsx) dans FuelEnocMovement"

    def add_arguments(self, parser):
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--sheets", type=str, default=None, help="Nom d'une feuille précise à traiter")

    def handle(self, *args, **options):
        import openpyxl

        dry_run = options["dry_run"]
        only_sheet = options.get("sheets")

        self.stdout.write("\n" + "═" * 80)
        self.stdout.write("  IMPORT HISTORIQUE REFUELING (xlsx → FuelEnocMovement)")
        self.stdout.write("═" * 80)

        resolver = SiteResolver()
        self.stdout.write(f"  Référentiel sites chargé : {len(resolver.by_norm_id)} sites\n")

        wb = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)

        payloads = []
        stats = {
            "total_cells": 0,
            "total_card": 0,
            "ponction_in": 0,
            "ponction_out_matched": 0,
            "ponction_out_unmatched": 0,
            "skipped_no_qty": 0,
            "hour_meter_parsed": 0,
        }

        sheets_to_process = {only_sheet: SHEETS[only_sheet]} if only_sheet else SHEETS

        for sheet_name, month in sheets_to_process.items():
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"  Feuille introuvable, ignorée : {sheet_name}"))
                continue

            ws = wb[sheet_name]
            headers = {c.column: c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW)) if c.value}
            block_starts = sorted(col for col, v in headers.items() if isinstance(v, str) and "refueling date" in v.lower())

            # Chaque feuille a une largeur/composition de bloc différente (8 ou 9
            # colonnes, présence ou non de "Technicien"/"Origine Fuel") — on repère
            # donc chaque champ par son en-tête réel plutôt qu'un décalage fixe.
            block_ends = block_starts[1:] + [block_starts[-1] + 10]
            blocks = []
            for b, next_b in zip(block_starts, block_ends):
                field_col = {}
                for c in range(b, min(next_b, b + 10)):
                    h = str(headers.get(c) or "").strip().lower()
                    if "quantity refueled" in h:
                        field_col["qty"] = c
                    elif "current height" in h:
                        field_col["height_after"] = c
                    elif "quantity before" in h:
                        field_col["height_before"] = c
                    elif h == "technicien":
                        field_col["technicien"] = c
                    elif "index generator" in h:
                        field_col["index_generator"] = c
                    elif "origine" in h:
                        field_col["origine"] = c
                    elif h == "comment":
                        field_col["comment"] = c
                blocks.append((b, field_col))

            self.stdout.write(f"  {sheet_name} : {len(blocks)} emplacements/site")

            for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=ws.max_row):
                site_id = row[0].value
                if not site_id or not isinstance(site_id, str):
                    continue
                site_id = site_id.strip()

                for block_idx, (b, field_col) in enumerate(blocks):
                    date_v = row[b - 1].value
                    height_before = row[field_col["height_before"] - 1].value if "height_before" in field_col else None
                    qty = row[field_col["qty"] - 1].value if "qty" in field_col else None
                    height_after = row[field_col["height_after"] - 1].value if "height_after" in field_col else None
                    technicien = row[field_col["technicien"] - 1].value if "technicien" in field_col else None
                    origine = row[field_col["origine"] - 1].value if "origine" in field_col else None
                    comment = row[field_col["comment"] - 1].value if "comment" in field_col else None
                    index_gen = row[field_col["index_generator"] - 1].value if "index_generator" in field_col else None
                    comment = str(comment) if comment is not None else None
                    technicien = str(technicien) if technicien is not None else None
                    hour_meter = _parse_hour_meter(index_gen)
                    if hour_meter is not None:
                        stats["hour_meter_parsed"] += 1

                    if not date_v or qty in (None, 0) or not isinstance(date_v, datetime):
                        continue

                    stats["total_cells"] += 1
                    origine_norm = (str(origine or "")).strip().lower()

                    # MARS n'a pas de colonne "Origine Fuel" — on se rabat sur le
                    # signe de la quantité (négatif = forcément une sortie/ponction).
                    if not origine and qty < 0:
                        origine_norm = "ponction"
                    op_date = _to_aware(date_v)

                    if "ponction" in origine_norm and qty < 0:
                        destinations = resolver.parse_destinations(comment)
                        if destinations:
                            stats["ponction_out_matched"] += 1
                            remaining = abs(qty)
                            for i, (dest_site_id, dest_qty) in enumerate(destinations):
                                amount = dest_qty if dest_qty is not None else (remaining / len(destinations))
                                payloads.append({
                                    "site_id": dest_site_id,
                                    "operation_type": "PONCTION",
                                    "quantity_added_liters": amount,
                                    "operation_date": op_date,
                                    "technician_name": technicien,
                                    "comment": comment,
                                    "ponction": {
                                        "source_site_id": site_id,
                                        "source_site_name": None,
                                        "comment": comment,
                                    },
                                    "source_id": f"{sheet_name}|{site_id}|{block_idx}|{i}|{op_date.date().isoformat()}",
                                })
                        else:
                            stats["ponction_out_unmatched"] += 1
                            payloads.append({
                                "site_id": site_id,
                                "operation_type": "PONCTION",
                                # Quantité en positif : c'est le montant de la ponction sortante,
                                # pas un "ajout" pour ce site. La vue exclut ce cas de Ajout In
                                # via ponction.source_site_id == site_id (auto-référence = non résolu).
                                "quantity_added_liters": abs(qty),
                                "operation_date": op_date,
                                "technician_name": technicien,
                                "comment": f"[NON RÉSOLU — destinataire non identifié] {comment or ''}".strip(),
                                "ponction": {"source_site_id": site_id, "source_site_name": None, "comment": comment},
                                "source_id": f"{sheet_name}|{site_id}|{block_idx}|{op_date.date().isoformat()}",
                            })
                        continue

                    if "ponction" in origine_norm:
                        stats["ponction_in"] += 1
                        op_type = "PONCTION"
                    else:
                        stats["total_card"] += 1
                        op_type = "TOTAL_CARD"

                    payloads.append({
                        "site_id": site_id,
                        "operation_type": op_type,
                        "quantity_added_liters": qty,
                        "level_before": _parse_cm(height_before),
                        "level_before_unit": "CM",
                        "level_after": _parse_cm(height_after),
                        "level_after_unit": "CM",
                        # Relevé compteur horaire GE au moment du passage — un seul point
                        # de mesure par visite, réutilisé pour before ET after (le delta
                        # sur la période se calcule entre deux visites successives, cf.
                        # RhCalculationService._apply_enoc_fallback).
                        "hour_meter_before": hour_meter,
                        "hour_meter_after": hour_meter,
                        "operation_date": op_date,
                        "technician_name": technicien,
                        "comment": comment,
                        "gauging_method": "Hauteur cuve",
                        "source_id": f"{sheet_name}|{site_id}|{block_idx}|{op_date.date().isoformat()}",
                    })

        self.stdout.write("\n  ── Statistiques de parsing ──────────────────────────")
        for k, v in stats.items():
            self.stdout.write(f"    {k}: {v}")
        self.stdout.write(f"    total mouvements à importer: {len(payloads)}\n")

        if dry_run:
            self.stdout.write(self.style.WARNING("  DRY RUN — aperçu des 10 premiers mouvements :\n"))
            for p in payloads[:10]:
                self.stdout.write(f"    {p['operation_date'].date()} | {p['site_id']} | {p['operation_type']} | {p['quantity_added_liters']} L | {p.get('comment') or ''}")
            self.stdout.write(self.style.WARNING("\n  Aucune donnée insérée (dry-run).\n"))
            return

        now = timezone.now()
        objects = []
        seen_keys = set()
        for p in payloads:
            key = (SOURCE_SYSTEM, p["source_id"])
            if key in seen_keys:
                continue
            seen_keys.add(key)
            objects.append(FuelEnocMovement(
                source_system=SOURCE_SYSTEM,
                source_id=p["source_id"],
                site_id=p["site_id"],
                operation_type=p["operation_type"],
                operation_date=p["operation_date"],
                quantity_added_liters=_dec(p["quantity_added_liters"]),
                level_before=_dec(p.get("level_before")),
                level_before_unit=p.get("level_before_unit"),
                level_after=_dec(p.get("level_after")),
                level_after_unit=p.get("level_after_unit"),
                hour_meter_before=_dec(p.get("hour_meter_before")),
                hour_meter_after=_dec(p.get("hour_meter_after")),
                gauging_method=p.get("gauging_method"),
                technician_name=p.get("technician_name"),
                comment=p.get("comment"),
                ponction=p.get("ponction"),
                import_source="historical_xlsx_import",
                import_key=p["source_id"],
                status="done",
                raw_payload={
                    **{k: v for k, v in p.items() if k not in ("operation_date", "hour_meter_before", "hour_meter_after")},
                    "operation_date": p["operation_date"].isoformat(),
                    "hour_meter_before": float(p["hour_meter_before"]) if p.get("hour_meter_before") is not None else None,
                    "hour_meter_after": float(p["hour_meter_after"]) if p.get("hour_meter_after") is not None else None,
                },
                synced_at=now,
                updated_at=now,
            ))

        with transaction.atomic():
            FuelEnocMovement.objects.bulk_create(
                objects,
                batch_size=1000,
                update_conflicts=True,
                unique_fields=["source_system", "source_id"],
                update_fields=[
                    "site_id", "operation_type", "operation_date", "quantity_added_liters",
                    "level_before", "level_before_unit", "level_after", "level_after_unit",
                    "hour_meter_before", "hour_meter_after",
                    "gauging_method", "technician_name", "comment", "ponction",
                    "raw_payload", "updated_at",
                ],
            )

        self.stdout.write(self.style.SUCCESS(f"\n  Import terminé — {len(objects)} mouvements insérés/mis à jour.\n"))
