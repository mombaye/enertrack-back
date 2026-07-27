# prediction/views.py
from __future__ import annotations

import calendar
import os
import pickle
from datetime import date
from io import BytesIO
from typing import Any

import numpy as np
import pandas as pd
from django.conf import settings
from django.db.models import Q, Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from billing.models import ContractSiteLink, MonthlySynthesis
from core.models import Site
from estimation.models import EstimationResult
from financial.models import FinancialEvaluation
from ml.features.events import event_features_for_month, events_in_month, get_event_dates
from ml.features.meteo import fetch_zone_meteo_cached
from ml.features.zones import normalize_zone


SOURCE_SCORE = {
    "GRID": 4,
    "ACM": 3,
    "HISTO": 2,
    "THEORIQUE": 1,
    "TARGET": 1,
    "NC": 0,
    "HORS_SCOPE": 0,
}

DEFAULT_FEATURES = [
    "month",
    "year",
    "is_hivernage",
    "is_hivernage_meteo",
    "load_w",
    "hors_catalogue",
    "source_score",
    "estimation_available",
    "conso_estimee_kwh",
    "acm_conso_kwh",
    "grid_conso_kwh",
    "histo_conso_30j",
    "conso_lag_1m",
    "conso_lag_2m",
    "conso_lag_3m",
    "conso_lag_6m",
    "conso_lag_12m",
    "conso_lag_24m",
    "same_month_last_year",
    "same_month_2y_ago",
    "same_month_avg",
    "rolling_3m_conso",
    "rolling_6m_conso",
    "trend_3m_vs_12m",
    "recurrence_score",
    "nb_jours_factures",
    "observed_month_ratio",
    "is_partial_billing",
    "temp_max_mean",
    "temp_min_mean",
    "precip_total",
    "humidity_max",
    "et0_mean",
    "event_magal_pressure",
    "event_gamou_pressure",
    "event_tabaski_pressure",
    "event_korite_pressure",
    "event_tamkharit_pressure",
    "event_magal_darou_pressure",
    "event_layene_pressure",
    "total_event_pressure",
]


def _month_add(year: int, month: int, delta: int) -> tuple[int, int]:
    m0 = year * 12 + (month - 1) + delta
    return m0 // 12, (m0 % 12) + 1


def _month_range(
    year_start: int,
    month_start: int,
    year_end: int | None = None,
    month_end: int | None = None,
    horizon: int | None = None,
) -> list[tuple[int, int]]:
    if year_end and month_end:
        if (year_end, month_end) < (year_start, month_start):
            raise ValidationError({"detail": "La période de fin doit être supérieure ou égale à la période de début."})

        out: list[tuple[int, int]] = []
        y, m = year_start, month_start
        while (y, m) <= (year_end, month_end):
            out.append((y, m))
            y, m = _month_add(y, m, 1)
        return out

    h = max(int(horizon or 6), 1)
    return [_month_add(year_start, month_start, i) for i in range(h)]


def _month_days(year: int, month: int) -> int:
    return calendar.monthrange(int(year), int(month))[1]


def _safe_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        value = float(v)
        if np.isfinite(value):
            return value
    except Exception:
        pass
    return default


def _safe_avg(values: list[Any]) -> float:
    vals = [_safe_float(v, 0) for v in values if v is not None and _safe_float(v, 0) > 0]
    return sum(vals) / len(vals) if vals else 0.0


def _project_to_full_month(conso: float, days: float, year: int, month: int) -> float:
    if days <= 0:
        return conso
    return conso / max(days, 1) * _month_days(year, month)


def _build_history_maps(site: Site, link: ContractSiteLink, min_year: int, max_year: int) -> dict[tuple[int, int], dict]:
    qs = (
        MonthlySynthesis.objects.filter(
            numero_compte_contrat=link.numero_compte_contrat,
            source__payment_status__in=["PAID", "UNPAID"],
            year__gte=min_year,
            year__lte=max_year,
        )
        .values("year", "month")
        .annotate(
            conso=Sum("conso"),
            montant_hors_tva=Sum("montant_hors_tva"),
            days_covered=Sum("days_covered"),
        )
        .order_by("year", "month")
    )

    by_key: dict[tuple[int, int], dict] = {}
    for h in qs:
        y, m = int(h["year"]), int(h["month"])
        conso = _safe_float(h.get("conso"), 0)
        ht = _safe_float(h.get("montant_hors_tva"), 0)
        days = _safe_float(h.get("days_covered"), 0) or _month_days(y, m)
        conso_full = _project_to_full_month(conso, days, y, m)
        month_days = _month_days(y, m)

        by_key[(y, m)] = {
            "year": y,
            "month": m,
            "conso": conso,
            "conso_full": conso_full,
            "montant_hors_tva": ht,
            "days": days,
            "month_days": month_days,
            "is_partial": days < month_days * 0.85,
        }

    return by_key


def _load_estimation_map(site: Site, min_year: int, max_year: int) -> dict[tuple[int, int], dict]:
    qs = (
        EstimationResult.objects.filter(
            site=site,
            batch__status="DONE",
            batch__year__gte=min_year,
            batch__year__lte=max_year,
        )
        .values(
            "batch__year",
            "batch__month",
            "conso_estimee_kwh",
            "source_utilisee",
            "fiabilite_grid",
            "acm_conso_kwh",
            "grid_conso_kwh",
            "histo_conso_30j",
            "histo_nb_mois",
        )
    )

    out: dict[tuple[int, int], dict] = {}
    for r in qs:
        y, m = int(r["batch__year"]), int(r["batch__month"])
        src = (r.get("source_utilisee") or "NC").upper()
        out[(y, m)] = {
            "conso_estimee_kwh": _safe_float(r.get("conso_estimee_kwh"), 0),
            "source_utilisee": src,
            "source_score": SOURCE_SCORE.get(src, 0),
            "acm_conso_kwh": _safe_float(r.get("acm_conso_kwh"), 0),
            "grid_conso_kwh": _safe_float(r.get("grid_conso_kwh"), 0),
            "histo_conso_30j": _safe_float(r.get("histo_conso_30j"), 0),
            "fiabilite_grid": r.get("fiabilite_grid"),
            "histo_nb_mois": r.get("histo_nb_mois"),
        }

    return out


def _load_financial_map(site: Site, min_year: int, max_year: int) -> tuple[dict[tuple[int, int], dict], dict]:
    qs = FinancialEvaluation.objects.filter(
        site=site,
        year__gte=min_year,
        year__lte=max_year,
    ).values(
        "year",
        "month",
        "load_w",
        "redevance",
        "recurrence_mois_nok",
        "hors_catalogue",
        "marge_statut",
        "marge",
    )

    out: dict[tuple[int, int], dict] = {}
    latest: tuple[tuple[int, int], dict] | None = None

    for r in qs:
        y, m = int(r["year"]), int(r["month"])
        item = {
            "load_w": _safe_float(r.get("load_w"), 0),
            "redevance": _safe_float(r.get("redevance"), 0),
            "recurrence_mois_nok": int(r.get("recurrence_mois_nok") or 0),
            "hors_catalogue": bool(r.get("hors_catalogue")),
            "marge_statut": r.get("marge_statut"),
            "marge": _safe_float(r.get("marge"), 0),
        }
        out[(y, m)] = item

        if latest is None or (y, m) > latest[0]:
            latest = ((y, m), item)

    return out, latest[1] if latest else {}


def _series_value(series: dict[tuple[int, int], float], year: int, month: int, lag: int) -> float:
    y, m = _month_add(year, month, -lag)
    return _safe_float(series.get((y, m)), 0)


def _rolling(series: dict[tuple[int, int], float], year: int, month: int, window: int) -> float:
    return _safe_avg([_series_value(series, year, month, lag) for lag in range(1, window + 1)])


def _rule_based_forecast(feature_row: dict) -> float:
    candidates = [
        (feature_row.get("conso_lag_1m"), 0.17),
        (feature_row.get("conso_lag_2m"), 0.08),
        (feature_row.get("conso_lag_3m"), 0.08),
        (feature_row.get("conso_lag_6m"), 0.08),
        (feature_row.get("conso_lag_12m"), 0.18),
        (feature_row.get("same_month_avg"), 0.17),
        (feature_row.get("rolling_3m_conso"), 0.16),
        (feature_row.get("rolling_6m_conso"), 0.08),
    ]

    weight_sum = sum(w for v, w in candidates if _safe_float(v, 0) > 0)
    base = sum(_safe_float(v, 0) * w for v, w in candidates if _safe_float(v, 0) > 0) / max(weight_sum, 0.01)

    est = _safe_float(feature_row.get("conso_estimee_kwh"), 0)
    source_score = _safe_float(feature_row.get("source_score"), 0)
    if est > 0 and source_score >= 2:
        alpha = 0.22 if source_score >= 3 else 0.14
        base = base * (1 - alpha) + est * alpha

    event_pressure = _safe_float(feature_row.get("total_event_pressure"), 0)
    precip = _safe_float(feature_row.get("precip_total"), 0)
    temp = _safe_float(feature_row.get("temp_max_mean"), 32)

    event_mult = 1 + min(event_pressure * 0.012, 0.22)
    weather_mult = 1.0

    if temp >= 36:
        weather_mult += 0.04
    if precip > 80:
        weather_mult += 0.03
    if feature_row.get("is_hivernage") and precip > 120:
        weather_mult += 0.02

    return max(base * event_mult * weather_mult, 0)


def _partial_month_final(observed_kwh: float, observed_days: float, full_month_pred: float, year: int, month: int):
    month_days = _month_days(year, month)

    if observed_kwh <= 0 or observed_days <= 0:
        return full_month_pred, 0.0, full_month_pred, False, 0, month_days

    remaining_days = max(month_days - int(round(observed_days)), 0)
    if remaining_days <= 0 or observed_days >= month_days * 0.95:
        return observed_kwh, observed_kwh, 0.0, False, int(round(observed_days)), 0

    projected_from_observed = observed_kwh / max(observed_days, 1) * month_days
    if observed_days >= 5:
        adjusted_full = full_month_pred * 0.60 + projected_from_observed * 0.40
    else:
        adjusted_full = full_month_pred * 0.82 + projected_from_observed * 0.18

    adjusted_full = max(adjusted_full, observed_kwh)
    remaining_pred = max(adjusted_full - observed_kwh, 0)

    return observed_kwh + remaining_pred, observed_kwh, remaining_pred, True, int(round(observed_days)), remaining_days


def _confidence(history_count: int, feature_row: dict, is_partial: bool, model_used: bool) -> float:
    conf = 0.78 if model_used else 0.66

    if history_count >= 18:
        conf += 0.08
    elif history_count < 6:
        conf -= 0.12

    if _safe_float(feature_row.get("total_event_pressure"), 0) > 5:
        conf -= 0.05
    if is_partial:
        conf += 0.03
    if _safe_float(feature_row.get("same_month_avg"), 0) <= 0:
        conf -= 0.05

    return round(min(max(conf, 0.45), 0.92), 2)


def _confidence_interval(pred: float, conf: float, event_pressure: float, is_partial: bool) -> tuple[int, int]:
    width = 0.18 - (conf - 0.45) * 0.12

    if event_pressure > 5:
        width += 0.04
    if is_partial:
        width -= 0.03

    width = min(max(width, 0.07), 0.26)
    return round(pred * (1 - width)), round(pred * (1 + width))


def _predict_fnp_score(recurrence_nok: int, marge_pred: float, redevance: float) -> float:
    score = 0.05

    if recurrence_nok >= 6:
        score += 0.45
    elif recurrence_nok >= 3:
        score += 0.20

    if marge_pred < 0:
        score += 0.25
    elif redevance > 0 and marge_pred < redevance * 0.1:
        score += 0.10

    return min(score, 0.99)


def _explain(feature_row: dict, events: list[dict], is_partial: bool, zone_norm: str, source_est: str | None) -> list[str]:
    out: list[str] = []

    if _safe_float(feature_row.get("same_month_avg"), 0) > 0:
        out.append("Saisonnalité annuelle prise en compte : même mois des années précédentes.")

    if _safe_float(feature_row.get("rolling_3m_conso"), 0) > 0:
        out.append("Tendance récente intégrée via moyenne mobile 3 mois.")

    if events:
        names = ", ".join(sorted({e.get("name", "") for e in events if e.get("name")}))
        out.append(f"Événement détecté sur la zone {zone_norm} : {names}.")

    if feature_row.get("is_hivernage"):
        out.append("Mois d’hivernage pris en compte.")

    if source_est and source_est not in ["NC", "HORS_SCOPE"]:
        out.append(f"Estimation existante utilisée comme signal complémentaire : {source_est}.")

    if is_partial:
        out.append("Mois partiel détecté : consommation déjà observée + reste du mois prédit.")

    return out


# ─────────────────────────────────────────────────────────────────────────────
# Excel export helpers
# ─────────────────────────────────────────────────────────────────────────────


def _excel_title(ws, title: str, end_column: int = 8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=15, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="021A40")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def _style_table(ws, header_row: int = 3):
    header_fill = PatternFill("solid", fgColor="032566")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))

    if ws.max_row < header_row:
        return

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 12
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 45))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len


def _append_rows(ws, headers: list[str], rows: list[dict], start_row: int = 3):
    for col, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=header)

    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(header))

    _style_table(ws, start_row)


def _export_bulk_predictions_xlsx(payload: dict) -> HttpResponse:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Synthèse"

    summary = payload.get("summary", {})
    results = payload.get("results", [])
    errors = payload.get("errors", [])

    _excel_title(ws_summary, "Rapport prévisionnel global parc", end_column=8)

    summary_rows = [
        {"Indicateur": "Sites demandés", "Valeur": summary.get("sites_requested")},
        {"Indicateur": "Sites traités", "Valeur": summary.get("sites_processed")},
        {"Indicateur": "Sites en erreur", "Valeur": summary.get("sites_error")},
        {"Indicateur": "Période début", "Valeur": summary.get("period_start")},
        {"Indicateur": "Période fin", "Valeur": summary.get("period_end")},
        {"Indicateur": "Mois prédits", "Valeur": summary.get("months_predicted")},
        {"Indicateur": "Conso prévue totale kWh", "Valeur": summary.get("total_conso_pred")},
        {"Indicateur": "HT prédit total", "Valeur": summary.get("total_ht_pred")},
        {"Indicateur": "Marge prévue totale", "Valeur": summary.get("total_marge_pred")},
        {"Indicateur": "Mois marge NOK", "Valeur": summary.get("months_marge_nok")},
        {"Indicateur": "Confiance moyenne", "Valeur": summary.get("avg_confidence")},
    ]
    _append_rows(ws_summary, ["Indicateur", "Valeur"], summary_rows)

    detail_headers = [
        "Site ID",
        "Nom site",
        "Zone brute",
        "Zone normalisée",
        "Période",
        "Conso prédite kWh",
        "Observé kWh",
        "Reste estimé kWh",
        "Mois partiel",
        "Jours observés",
        "Jours restants",
        "IC bas",
        "IC haut",
        "HT prédit",
        "Redevance",
        "Marge prédite",
        "Marge OK",
        "Confiance",
        "Risque FNP",
        "Pression événement",
        "Temp max moyenne",
        "Pluie totale",
        "Source estimation",
        "Conso estimation kWh",
        "Explications",
    ]
    detail_rows: list[dict] = []
    for site_result in results:
        for p in site_result.get("predictions", []):
            detail_rows.append(
                {
                    "Site ID": site_result.get("site_id"),
                    "Nom site": site_result.get("site_name"),
                    "Zone brute": site_result.get("zone_raw"),
                    "Zone normalisée": site_result.get("zone_normalized"),
                    "Période": p.get("period"),
                    "Conso prédite kWh": p.get("conso_pred"),
                    "Observé kWh": p.get("observed_kwh"),
                    "Reste estimé kWh": p.get("remaining_pred_kwh"),
                    "Mois partiel": "Oui" if p.get("is_partial_month") else "Non",
                    "Jours observés": p.get("observed_days"),
                    "Jours restants": p.get("remaining_days"),
                    "IC bas": p.get("ic_lo"),
                    "IC haut": p.get("ic_hi"),
                    "HT prédit": p.get("ht_pred"),
                    "Redevance": p.get("redevance"),
                    "Marge prédite": p.get("marge_pred"),
                    "Marge OK": "Oui" if p.get("marge_ok") else "Non",
                    "Confiance": p.get("confidence"),
                    "Risque FNP": p.get("fnp_score"),
                    "Pression événement": p.get("event_pressure"),
                    "Temp max moyenne": (p.get("meteo") or {}).get("temp_max_mean"),
                    "Pluie totale": (p.get("meteo") or {}).get("precip_total"),
                    "Source estimation": (p.get("estimation") or {}).get("source"),
                    "Conso estimation kWh": (p.get("estimation") or {}).get("conso_estimee_kwh"),
                    "Explications": " | ".join(p.get("explanation") or []),
                }
            )

    ws_detail = wb.create_sheet("Détail mensuel")
    _excel_title(ws_detail, "Détail mensuel par site", end_column=len(detail_headers))
    _append_rows(ws_detail, detail_headers, detail_rows)

    event_headers = [
        "Site ID",
        "Zone normalisée",
        "Période",
        "Événement",
        "Date",
        "Début fenêtre",
        "Fin fenêtre",
        "Jours dans le mois",
        "Poids zone",
        "Pression",
    ]
    event_rows: list[dict] = []
    for site_result in results:
        for p in site_result.get("predictions", []):
            for e in p.get("events") or []:
                event_rows.append(
                    {
                        "Site ID": site_result.get("site_id"),
                        "Zone normalisée": site_result.get("zone_normalized"),
                        "Période": p.get("period"),
                        "Événement": e.get("name"),
                        "Date": e.get("date"),
                        "Début fenêtre": e.get("window_start"),
                        "Fin fenêtre": e.get("window_end"),
                        "Jours dans le mois": e.get("days_in_month"),
                        "Poids zone": e.get("zone_weight"),
                        "Pression": e.get("pressure"),
                    }
                )

    ws_events = wb.create_sheet("Événements")
    _excel_title(ws_events, "Événements détectés", end_column=len(event_headers))
    _append_rows(ws_events, event_headers, event_rows)

    hist_headers = [
        "Site ID",
        "Nom site",
        "Zone normalisée",
        "Période",
        "Conso brute kWh",
        "Conso normalisée kWh",
        "Montant HT",
        "Jours",
        "Partiel",
    ]
    hist_rows: list[dict] = []
    for site_result in results:
        for h in site_result.get("historic", []):
            hist_rows.append(
                {
                    "Site ID": site_result.get("site_id"),
                    "Nom site": site_result.get("site_name"),
                    "Zone normalisée": site_result.get("zone_normalized"),
                    "Période": h.get("period"),
                    "Conso brute kWh": h.get("conso"),
                    "Conso normalisée kWh": h.get("conso_full"),
                    "Montant HT": h.get("ht"),
                    "Jours": h.get("days"),
                    "Partiel": "Oui" if h.get("is_partial") else "Non",
                }
            )

    ws_hist = wb.create_sheet("Historique")
    _excel_title(ws_hist, "Historique utilisé", end_column=len(hist_headers))
    _append_rows(ws_hist, hist_headers, hist_rows)

    ws_errors = wb.create_sheet("Erreurs")
    _excel_title(ws_errors, "Sites non traités", end_column=2)
    _append_rows(ws_errors, ["Site ID", "Erreur"], errors)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"prediction_parc_{summary.get('period_start', '')}_{summary.get('period_end', '')}.xlsx".replace("/", "-")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# API views
# ─────────────────────────────────────────────────────────────────────────────


class PredictionView(APIView):
    permission_classes = [IsAuthenticated]

    MODEL_PATH = os.path.join(settings.BASE_DIR, "ml", "models", "lgbm_conso.pkl")
    FEATURES_PATH = os.path.join(settings.BASE_DIR, "ml", "models", "features.pkl")
    METADATA_PATH = os.path.join(settings.BASE_DIR, "ml", "models", "metadata.pkl")

    def _load_model_features(self):
        model = None
        features = DEFAULT_FEATURES
        metadata = {"model_version": "rule_based_hybrid_v2"}

        try:
            with open(self.MODEL_PATH, "rb") as f:
                model = pickle.load(f)
        except FileNotFoundError:
            model = None

        try:
            with open(self.FEATURES_PATH, "rb") as f:
                features = pickle.load(f)
        except FileNotFoundError:
            pass

        try:
            with open(self.METADATA_PATH, "rb") as f:
                metadata = pickle.load(f)
        except FileNotFoundError:
            pass

        return model, features, metadata

    def _forecast_one_site(
        self,
        site_id: str,
        year_start: int,
        month_start: int,
        year_end_i: int | None = None,
        month_end_i: int | None = None,
        horizon: int = 6,
    ) -> dict:
        try:
            site = Site.objects.get(site_id=site_id)
        except Site.DoesNotExist:
            raise ValidationError({"site": f"Site {site_id!r} introuvable"})

        link = ContractSiteLink.objects.filter(site=site).first()
        if not link:
            raise ValidationError({"site": f"Aucun contrat trouvé pour le site {site_id}"})

        periods = _month_range(
            year_start=year_start,
            month_start=month_start,
            year_end=year_end_i,
            month_end=month_end_i,
            horizon=horizon,
        )

        if not periods:
            raise ValidationError({"detail": "Période de prédiction invalide."})

        min_year = min(y for y, _ in periods) - 3
        max_year = max(y for y, _ in periods) + 1

        history_map = _build_history_maps(site, link, min_year, max_year)
        if len(history_map) < 3:
            raise ValidationError({"detail": f"Historique insuffisant pour {site_id} (< 3 mois)"})

        estimation_map = _load_estimation_map(site, min_year, max_year)
        financial_map, latest_financial = _load_financial_map(site, min_year, max_year)

        zone_raw = site.zone or "DKR"
        zone_norm = normalize_zone(zone_raw, getattr(site, "name", None), site.site_id)

        events_calendar = get_event_dates(min_year, max_year)
        model, features, metadata = self._load_model_features()

        series: dict[tuple[int, int], float] = {
            (y, m): h["conso_full"]
            for (y, m), h in history_map.items()
            if h.get("conso_full", 0) > 0
        }

        predictions: list[dict] = []

        for y, m in periods:
            hist_current = history_map.get((y, m), {})
            est = estimation_map.get((y, m), {})
            fin = financial_map.get((y, m), latest_financial or {})

            meteo = fetch_zone_meteo_cached(zone_norm, y, m)
            event_row = event_features_for_month(
                events_calendar,
                y,
                m,
                zone_norm,
                getattr(site, "name", None),
                site.site_id,
            )
            month_events = events_in_month(
                events_calendar,
                y,
                m,
                zone_norm,
                getattr(site, "name", None),
                site.site_id,
            )

            lag1 = _series_value(series, y, m, 1)
            lag2 = _series_value(series, y, m, 2)
            lag3 = _series_value(series, y, m, 3)
            lag6 = _series_value(series, y, m, 6)
            lag12 = _series_value(series, y, m, 12)
            lag24 = _series_value(series, y, m, 24)

            rolling3 = _rolling(series, y, m, 3)
            rolling6 = _rolling(series, y, m, 6)
            same_avg = _safe_avg([lag12, lag24])

            observed_kwh = _safe_float(hist_current.get("conso"), 0)
            observed_days = _safe_float(hist_current.get("days"), 0)

            month_days = _month_days(y, m)
            observed_ratio = min(observed_days / month_days, 1) if observed_days > 0 else 0
            is_partial = bool(observed_kwh > 0 and observed_ratio < 0.95)

            source_est = (est.get("source_utilisee") or "NC").upper()

            feature_row = {
                "month": m,
                "year": y,
                "is_hivernage": int(m in [6, 7, 8, 9, 10]),
                "is_hivernage_meteo": int(_safe_float(meteo.get("precip_total"), 0) > 20),
                "load_w": _safe_float(fin.get("load_w"), _safe_float(getattr(site, "analysis_load", 0), 0)),
                "hors_catalogue": int(bool(fin.get("hors_catalogue", False))),
                "source_score": SOURCE_SCORE.get(source_est, 0),
                "estimation_available": int(_safe_float(est.get("conso_estimee_kwh"), 0) > 0),
                "conso_estimee_kwh": _safe_float(est.get("conso_estimee_kwh"), 0),
                "acm_conso_kwh": _safe_float(est.get("acm_conso_kwh"), 0),
                "grid_conso_kwh": _safe_float(est.get("grid_conso_kwh"), 0),
                "histo_conso_30j": _safe_float(est.get("histo_conso_30j"), 0),
                "conso_lag_1m": lag1,
                "conso_lag_2m": lag2,
                "conso_lag_3m": lag3,
                "conso_lag_6m": lag6,
                "conso_lag_12m": lag12,
                "conso_lag_24m": lag24,
                "same_month_last_year": lag12,
                "same_month_2y_ago": lag24,
                "same_month_avg": same_avg,
                "rolling_3m_conso": rolling3,
                "rolling_6m_conso": rolling6,
                "trend_3m_vs_12m": (rolling3 / lag12 - 1) if lag12 > 0 else 0,
                "recurrence_score": int(fin.get("recurrence_mois_nok") or 0),
                "nb_jours_factures": observed_days or 30,
                "observed_month_ratio": observed_ratio or 1,
                "is_partial_billing": int(is_partial),
                "temp_max_mean": _safe_float(meteo.get("temp_max_mean"), 32),
                "temp_min_mean": _safe_float(meteo.get("temp_min_mean"), 22),
                "precip_total": _safe_float(meteo.get("precip_total"), 0),
                "humidity_max": _safe_float(meteo.get("humidity_max"), 65),
                "et0_mean": _safe_float(meteo.get("et0_mean"), 4.5),
                **event_row,
            }

            rule_pred = _rule_based_forecast(feature_row)
            model_used = model is not None

            if model is not None:
                row_for_model = {col: feature_row.get(col, 0) for col in features}
                X = pd.DataFrame([row_for_model])[features]
                X = X.replace([np.inf, -np.inf], np.nan).fillna(0)

                ml_pred = float(model.predict(X)[0])
                full_month_pred = ml_pred * 0.70 + rule_pred * 0.30
            else:
                ml_pred = None
                full_month_pred = rule_pred

            est_conso = _safe_float(est.get("conso_estimee_kwh"), 0)
            if est_conso > 0:
                alpha = 0.22 if SOURCE_SCORE.get(source_est, 0) >= 3 else 0.12
                full_month_pred = full_month_pred * (1 - alpha) + est_conso * alpha

            (
                final_pred,
                observed_part,
                remaining_pred,
                partial_flag,
                observed_days_out,
                remaining_days,
            ) = _partial_month_final(observed_kwh, observed_days, full_month_pred, y, m)

            past_ratios = []
            for lag in range(1, 7):
                py, pm = _month_add(y, m, -lag)
                h = history_map.get((py, pm))
                if h and h.get("conso", 0) > 0 and h.get("montant_hors_tva", 0) > 0:
                    past_ratios.append(h["montant_hors_tva"] / max(h["conso"], 1))

            avg_ratio = _safe_avg(past_ratios) or 150.0
            ht_pred = final_pred * avg_ratio
            redevance = _safe_float(fin.get("redevance"), 0)
            marge_pred = redevance - ht_pred

            conf = _confidence(len(history_map), feature_row, partial_flag, model_used)
            ic_lo, ic_hi = _confidence_interval(
                final_pred,
                conf,
                _safe_float(event_row.get("total_event_pressure"), 0),
                partial_flag,
            )
            fnp_score = _predict_fnp_score(
                int(fin.get("recurrence_mois_nok") or 0),
                marge_pred,
                redevance,
            )

            series[(y, m)] = final_pred

            predictions.append(
                {
                    "period": f"{y}-{m:02d}",
                    "year": y,
                    "month": m,
                    "conso_pred": round(final_pred, 1),
                    "conso_full_month_model": round(full_month_pred, 1),
                    "conso_rule_based": round(rule_pred, 1),
                    "conso_ml": round(ml_pred, 1) if ml_pred is not None else None,
                    "observed_kwh": round(observed_part, 1),
                    "remaining_pred_kwh": round(remaining_pred, 1),
                    "is_partial_month": partial_flag,
                    "observed_days": observed_days_out,
                    "remaining_days": remaining_days,
                    "ht_pred": round(ht_pred),
                    "redevance": round(redevance),
                    "marge_pred": round(marge_pred),
                    "marge_ok": marge_pred >= 0,
                    "ic_lo": ic_lo,
                    "ic_hi": ic_hi,
                    "confidence": conf,
                    "fnp_score": round(fnp_score, 3),
                    "baseline": {
                        "lag_1m": round(lag1, 1),
                        "lag_3m": round(lag3, 1),
                        "lag_6m": round(lag6, 1),
                        "lag_12m": round(lag12, 1),
                        "same_month_avg": round(same_avg, 1),
                        "rolling_3m": round(rolling3, 1),
                        "rolling_6m": round(rolling6, 1),
                        "trend_3m_vs_12m": round(feature_row["trend_3m_vs_12m"], 4),
                    },
                    "events": month_events,
                    "event_pressure": round(_safe_float(event_row.get("total_event_pressure"), 0), 3),
                    "meteo": meteo,
                    "estimation": {
                        "available": bool(est_conso > 0),
                        "source": source_est,
                        "conso_estimee_kwh": round(est_conso, 1) if est_conso else None,
                        "source_score": SOURCE_SCORE.get(source_est, 0),
                    },
                    "explanation": _explain(feature_row, month_events, partial_flag, zone_norm, source_est),
                }
            )

        historic: list[dict] = []
        for (y, m), h in sorted(history_map.items())[-12:]:
            historic.append(
                {
                    "period": f"{y}-{m:02d}",
                    "conso": round(h["conso"], 1),
                    "conso_full": round(h["conso_full"], 1),
                    "ht": round(h["montant_hors_tva"], 1),
                    "days": h["days"],
                    "is_partial": h["is_partial"],
                }
            )

        return {
            "site_id": site.site_id,
            "site_name": getattr(site, "name", None),
            "zone_raw": zone_raw,
            "zone_normalized": zone_norm,
            "period_start": f"{periods[0][0]}-{periods[0][1]:02d}",
            "period_end": f"{periods[-1][0]}-{periods[-1][1]:02d}",
            "horizon": len(periods),
            "generated_at": date.today().isoformat(),
            "model_version": metadata.get("model_version") if model is not None else "rule_based_hybrid_v2",
            "model_used": model is not None,
            "historic": historic,
            "predictions": predictions,
        }

    def get(self, request):
        site_id = request.query_params.get("site")
        if not site_id:
            raise ValidationError({"site": "Paramètre requis"})

        horizon = int(request.query_params.get("horizon", 6))
        today = date.today()
        year_start = int(request.query_params.get("year_start", today.year))
        month_start = int(request.query_params.get("month_start", today.month))
        year_end = request.query_params.get("year_end")
        month_end = request.query_params.get("month_end")

        payload = self._forecast_one_site(
            site_id=site_id,
            year_start=year_start,
            month_start=month_start,
            year_end_i=int(year_end) if year_end else None,
            month_end_i=int(month_end) if month_end else None,
            horizon=horizon,
        )
        return Response(payload)


class PredictionBulkForecastView(PredictionView):
    """Prévision globale parc ou par zone.

    JSON :
    GET /api/prediction/forecast-bulk/?year_start=2026&month_start=4&year_end=2026&month_end=9

    Excel :
    GET /api/prediction/forecast-bulk/?year_start=2026&month_start=4&year_end=2026&month_end=9&export=xlsx
    """

    def get(self, request):
        today = date.today()
        horizon = int(request.query_params.get("horizon", 6))
        year_start = int(request.query_params.get("year_start", today.year))
        month_start = int(request.query_params.get("month_start", today.month))
        year_end = request.query_params.get("year_end")
        month_end = request.query_params.get("month_end")
        year_end_i = int(year_end) if year_end else None
        month_end_i = int(month_end) if month_end else None

        zone = (request.query_params.get("zone") or "").strip()
        zone_norm_filter = normalize_zone(zone) if zone else ""
        search = (request.query_params.get("search") or "").strip()
        limit = int(request.query_params.get("limit", 0) or 0)
        export = (request.query_params.get("export") or "").lower()

        links = ContractSiteLink.objects.select_related("site").all().order_by("site__site_id")

        if search:
            links = links.filter(
                Q(site__site_id__icontains=search)
                | Q(site__name__icontains=search)
                | Q(numero_compte_contrat__icontains=search)
            )

        results: list[dict] = []
        errors: list[dict] = []
        sites_requested = 0
        seen_sites: set[str] = set()

        for link in links:
            site = link.site
            if not site or not getattr(site, "site_id", None):
                continue

            if site.site_id in seen_sites:
                continue
            seen_sites.add(site.site_id)

            if zone_norm_filter:
                current_zone_norm = normalize_zone(site.zone, getattr(site, "name", None), site.site_id)
                if current_zone_norm != zone_norm_filter:
                    continue

            sites_requested += 1

            if limit > 0 and len(results) + len(errors) >= limit:
                break

            try:
                payload = self._forecast_one_site(
                    site_id=site.site_id,
                    year_start=year_start,
                    month_start=month_start,
                    year_end_i=year_end_i,
                    month_end_i=month_end_i,
                    horizon=horizon,
                )
                results.append(payload)
            except Exception as exc:
                errors.append(
                    {
                        "Site ID": getattr(site, "site_id", None),
                        "Erreur": str(exc),
                    }
                )

        all_predictions = []
        for site_result in results:
            for p in site_result.get("predictions", []):
                all_predictions.append(p)

        total_conso = sum(float(p.get("conso_pred") or 0) for p in all_predictions)
        total_ht = sum(float(p.get("ht_pred") or 0) for p in all_predictions)
        total_marge = sum(float(p.get("marge_pred") or 0) for p in all_predictions)
        months_nok = sum(1 for p in all_predictions if not p.get("marge_ok"))
        avg_conf = (
            sum(float(p.get("confidence") or 0) for p in all_predictions) / len(all_predictions)
            if all_predictions
            else 0
        )

        period_start = f"{year_start}-{month_start:02d}"
        if year_end_i and month_end_i:
            period_end = f"{year_end_i}-{month_end_i:02d}"
        else:
            last_year, last_month = _month_add(year_start, month_start, horizon - 1)
            period_end = f"{last_year}-{last_month:02d}"

        payload = {
            "mode": "bulk",
            "generated_at": date.today().isoformat(),
            "summary": {
                "period_start": period_start,
                "period_end": period_end,
                "zone": zone_norm_filter or None,
                "sites_requested": sites_requested,
                "sites_processed": len(results),
                "sites_error": len(errors),
                "months_predicted": len(all_predictions),
                "total_conso_pred": round(total_conso, 1),
                "total_ht_pred": round(total_ht),
                "total_marge_pred": round(total_marge),
                "months_marge_nok": months_nok,
                "avg_confidence": round(avg_conf, 3),
            },
            "results": results,
            "errors": errors,
        }

        if export == "xlsx":
            return _export_bulk_predictions_xlsx(payload)

        return Response(payload)
