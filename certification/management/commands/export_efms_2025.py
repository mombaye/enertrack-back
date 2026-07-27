# certification/management/commands/export_efms_2025.py
"""
Commande Django pour extraire les données eFMS (ACM + Grid) de l'année 2025 vers Excel.

Usage:
    python manage.py export_efms_2025
    python manage.py export_efms_2025 --output /chemin/vers/fichier.xlsx
    python manage.py export_efms_2025 --site-id DKR_0028
    python manage.py export_efms_2025 --start 2025-01-01 --end 2025-06-30
"""
import sys
from datetime import date
from pathlib import Path

import pandas as pd
import pyodbc
from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class Command(BaseCommand):
    help = "Exporte les données eFMS (tables ACM et Grid) pour 2025 vers Excel"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", "-o",
            type=str,
            default="efms_export_2025.xlsx",
            help="Chemin du fichier Excel de sortie (défaut: efms_export_2025.xlsx)",
        )
        parser.add_argument(
            "--site-id",
            type=str,
            default=None,
            help="Filtrer sur un site_id spécifique (optionnel)",
        )
        parser.add_argument(
            "--start",
            type=str,
            default="2025-01-01",
            help="Date de début (défaut: 2025-01-01)",
        )
        parser.add_argument(
            "--end",
            type=str,
            default="2025-12-31",
            help="Date de fin (défaut: 2025-12-31)",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=None,
            help="Limiter le nombre de lignes extraites (pour test)",
        )

    def _get_connection_string(self) -> str:
        host = getattr(settings, "EFMS_SQL_HOST", "172.30.0.149")
        port = int(getattr(settings, "EFMS_SQL_PORT", 1433))
        db = getattr(settings, "EFMS_SQL_DB", "SQL1-ProdDB")
        user = getattr(settings, "EFMS_SQL_USER", "")
        password = getattr(settings, "EFMS_SQL_PASSWORD", "")
        driver = getattr(settings, "EFMS_SQL_DRIVER", "ODBC Driver 17 for SQL Server")
        timeout = int(getattr(settings, "EFMS_SQL_TIMEOUT", 30))

        return (
            f"DRIVER={{{driver}}};"
            f"SERVER={host},{port};"
            f"DATABASE={db};"
            f"UID={user};"
            f"PWD={password};"
            f"Connection Timeout={timeout};"
            "TrustServerCertificate=yes;"
        )

    def _get_table_columns(self, cursor, table_name: str) -> list[str]:
        """Récupère la liste des colonnes d'une table."""
        try:
            cursor.execute(f"SELECT TOP 0 * FROM {table_name}")
            return [col[0] for col in cursor.description]
        except Exception:
            return []

    def _style_header(self, ws):
        """Applique le style aux en-têtes (première ligne)."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def _auto_column_width(self, ws):
        """Ajuste automatiquement la largeur des colonnes."""
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_len)
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def handle(self, *args, **options):
        output_path = options["output"]
        site_filter = options["site_id"]
        start_date = options["start"]
        end_date = options["end"]
        limit = options["limit"]

        self.stdout.write("\n" + "═" * 60)
        self.stdout.write("  EXPORT eFMS → EXCEL")
        self.stdout.write("═" * 60 + "\n")

        self.stdout.write(f"  Période     : {start_date} → {end_date}")
        self.stdout.write(f"  Fichier     : {output_path}")
        if site_filter:
            self.stdout.write(f"  Site filtre : {site_filter}")
        self.stdout.write("")

        # ── Connexion SQL Server ──────────────────────────────────────────────
        self.stdout.write("  Connexion à eFMS SQL Server...")
        try:
            conn = pyodbc.connect(self._get_connection_string(), timeout=30)
            cursor = conn.cursor()
            self.stdout.write(self.style.SUCCESS("  ✓ Connecté"))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"  ✗ Erreur connexion: {e}"))
            sys.exit(1)

        # ── Tables eFMS ───────────────────────────────────────────────────────
        TABLE_ACM = "[SQL1-ProdDB].[dbo].[silver.gfms_AC_Meter_Report_day]"
        TABLE_GRID = "[SQL1-ProdDB].[dbo].[silver.gfms_Grid_Report_day]"

        limit_clause = f"TOP {limit}" if limit else ""

        # ── Découverte des colonnes ACM ───────────────────────────────────────
        self.stdout.write("\n  Découverte colonnes ACM...")
        acm_columns = self._get_table_columns(cursor, TABLE_ACM)
        if acm_columns:
            self.stdout.write(f"    Colonnes disponibles: {', '.join(acm_columns[:10])}...")
        else:
            self.stdout.write(self.style.WARNING("    Table ACM inaccessible"))

        # ── Extraction ACM (toutes colonnes) ──────────────────────────────────
        self.stdout.write("\n  Extraction table ACM...")
        df_acm = pd.DataFrame()

        if acm_columns:
            where_acm = f"WHERE [Date] >= '{start_date}' AND [Date] <= '{end_date}'"
            if site_filter:
                where_acm += f" AND site_id = '{site_filter}'"

            query_acm = f"""
                SELECT {limit_clause} *
                FROM {TABLE_ACM}
                {where_acm}
                ORDER BY site_id, [Date]
            """

            try:
                df_acm = pd.read_sql(query_acm, conn)
                self.stdout.write(self.style.SUCCESS(f"  ✓ ACM : {len(df_acm):,} lignes, {len(df_acm.columns)} colonnes"))
            except Exception as e:
                self.stdout.write(self.style.WARNING(f"  ⚠ Erreur ACM: {e}"))

        # ── Découverte des colonnes Grid ──────────────────────────────────────
        self.stdout.write("\n  Découverte colonnes Grid...")
        grid_columns = self._get_table_columns(cursor, TABLE_GRID)
        if grid_columns:
            self.stdout.write(f"    Colonnes disponibles: {', '.join(grid_columns[:10])}...")
        else:
            self.stdout.write(self.style.WARNING("    Table Grid inaccessible"))

        # ── Extraction Grid (toutes colonnes) ─────────────────────────────────
        self.stdout.write("\n  Extraction table Grid...")
        df_grid = pd.DataFrame()

        if grid_columns:
            where_grid = f"WHERE [Date] >= '{start_date}' AND [Date] <= '{end_date}'"
            if site_filter:
                where_grid += f" AND site_id = '{site_filter}'"

            query_grid = f"""
                SELECT {limit_clause} *
                FROM {TABLE_GRID}
                {where_grid}
                ORDER BY site_id, [Date]
            """

            try:
                df_grid = pd.read_sql(query_grid, conn)
                self.stdout.write(self.style.SUCCESS(f"  ✓ Grid : {len(df_grid):,} lignes, {len(df_grid.columns)} colonnes"))
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"  ✗ Erreur Grid: {e}"))

        conn.close()

        # ── Vérification données ──────────────────────────────────────────────
        if df_acm.empty and df_grid.empty:
            self.stdout.write(self.style.ERROR("\n  ✗ Aucune donnée extraite. Vérifiez la période et la connexion."))
            sys.exit(1)

        # ── Export Excel ──────────────────────────────────────────────────────
        self.stdout.write("\n  Génération fichier Excel...")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            if not df_acm.empty:
                df_acm.to_excel(writer, sheet_name="ACM_2025", index=False)

            if not df_grid.empty:
                df_grid.to_excel(writer, sheet_name="Grid_2025", index=False)

            # Feuille Résumé
            summary_data = {
                "Métrique": [
                    "Période extraction",
                    "Date export",
                    "Sites ACM distincts",
                    "Sites Grid distincts",
                    "Lignes ACM",
                    "Lignes Grid",
                    "Colonnes ACM",
                    "Colonnes Grid",
                ],
                "Valeur": [
                    f"{start_date} → {end_date}",
                    str(date.today()),
                    df_acm["site_id"].nunique() if not df_acm.empty and "site_id" in df_acm.columns else 0,
                    df_grid["site_id"].nunique() if not df_grid.empty and "site_id" in df_grid.columns else 0,
                    len(df_acm),
                    len(df_grid),
                    len(df_acm.columns) if not df_acm.empty else 0,
                    len(df_grid.columns) if not df_grid.empty else 0,
                ],
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="Résumé", index=False)

            # Feuille liste des colonnes
            col_data = {"Table": [], "Colonne": []}
            for col in (acm_columns or []):
                col_data["Table"].append("ACM")
                col_data["Colonne"].append(col)
            for col in (grid_columns or []):
                col_data["Table"].append("Grid")
                col_data["Colonne"].append(col)
            if col_data["Table"]:
                pd.DataFrame(col_data).to_excel(writer, sheet_name="Colonnes", index=False)

        # ── Mise en forme Excel ───────────────────────────────────────────────
        wb = load_workbook(output_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._style_header(ws)
            self._auto_column_width(ws)
            ws.freeze_panes = "A2"

        wb.save(output_path)

        # ── Résumé final ──────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 60)
        self.stdout.write(self.style.SUCCESS(f"  ✓ Export terminé : {output_path}"))
        self.stdout.write("")
        self.stdout.write("  Feuilles créées :")
        if not df_acm.empty:
            self.stdout.write(f"    • ACM_2025  : {len(df_acm):,} lignes, {df_acm['site_id'].nunique()} sites")
        if not df_grid.empty:
            self.stdout.write(f"    • Grid_2025 : {len(df_grid):,} lignes, {df_grid['site_id'].nunique()} sites")
        self.stdout.write("    • Résumé    : statistiques globales")
        self.stdout.write("    • Colonnes  : liste des colonnes disponibles")
        self.stdout.write("═" * 60 + "\n")

        self.stdout.write(self.style.SUCCESS(f"\n  Fichier prêt à envoyer : {Path(output_path).absolute()}\n"))