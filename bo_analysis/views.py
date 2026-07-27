from __future__ import annotations

import io
import unicodedata

import openpyxl
from openpyxl.utils import column_index_from_string

from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404

from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from core.models import Site

from .models import ActionOwner, BOAnalysis, BOAnalysisRequest, BOMarginSnapshot, CategorieBO
from .tasks import notify_bo_analysis_submitted, notify_bo_bulk_request_created, notify_bo_request_created


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _norm_key(s) -> str:
    """Clé de comparaison robuste aux accents/casse/espaces (pour matcher Categorie BO / Action Owner)."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.replace(" ", "")


def _build_choice_lookup(choices_cls):
    return {_norm_key(label): value for value, label in choices_cls.choices}


_CATEGORIE_BO_LOOKUP = _build_choice_lookup(CategorieBO)
_ACTION_OWNER_LOOKUP = _build_choice_lookup(ActionOwner)


def _match_choice(raw_value, lookup: dict, autre_value: str) -> tuple[str, str]:
    """Retourne (choice_value, choice_autre_text). Fallback sur 'autre' + texte brut si aucun match."""
    if raw_value is None or str(raw_value).strip() == "":
        return "", ""
    key = _norm_key(raw_value)
    matched = lookup.get(key)
    if matched:
        return matched, ""
    return autre_value, str(raw_value).strip()


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _bool_done(v) -> bool:
    return str(v).strip().lower() in ("done", "oui", "yes", "true", "1")


# ─────────────────────────────────────────────────────────────────────────────
# Import du fichier historique Analyse Marge.xlsx (instantané unique, lecture seule)
# ─────────────────────────────────────────────────────────────────────────────

# Colonnes du fichier "Analyse Marge.xlsx", feuille "Energy Analysis" — header ligne 13, données dès ligne 14.
_COLS = {
    "site_id":                 "A",
    "site_name":               "B",
    "zone":                    "E",
    "typologie_reelle":        "G",
    "puissance_facturee_a":    "N",
    "puissance_facturee_b":    "O",
    "redevance_grid_a":        "R",
    "estimation_conso_kwh_a":  "S",
    "estimation_conso_xof_a":  "T",
    "redevance_vs_estimation_a": "U",
    "redevance_grid_b":        "V",
    "estimation_conso_kwh_b":  "W",
    "estimation_conso_xof_b":  "X",
    "redevance_vs_estimation_b": "Y",
    "statut_marge":            "Z",
    "grid_action_plan_ops":    "AE",
    "categorie_ops":           "AF",
    "commentaire_bo":          "AG",
    "categorie_bo":            "AH",
    "check":                   "AI",
    "action_owner":            "AJ",
    "commentaire":             "AK",
}
_COL_IDX = {k: column_index_from_string(v) - 1 for k, v in _COLS.items()}

_HEADER_ROW = 13
_DATA_START_ROW = 14
_SHEET_NAME = "Energy Analysis"
_MONTH_A_LABEL = "Mai"
_MONTH_B_LABEL = "Juin"


def _cell(row, key):
    idx = _COL_IDX[key]
    return row[idx] if idx < len(row) else None


def _parse_bo_margin_xlsx(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_name = _SHEET_NAME if _SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    records = []
    for i, row in enumerate(ws.iter_rows(min_row=_DATA_START_ROW, values_only=True), start=_DATA_START_ROW):
        site_id = _cell(row, "site_id")
        if not site_id or str(site_id).strip() in ("", "nan", "None"):
            continue

        categorie_bo, categorie_bo_autre = _match_choice(
            _cell(row, "categorie_bo"), _CATEGORIE_BO_LOOKUP, CategorieBO.AUTRE
        )
        action_owner, action_owner_autre = _match_choice(
            _cell(row, "action_owner"), _ACTION_OWNER_LOOKUP, ActionOwner.AUTRE
        )

        records.append({
            "site_id_raw": str(site_id).strip(),
            "site_name_raw": str(_cell(row, "site_name") or "").strip(),
            "zone": str(_cell(row, "zone") or "").strip(),
            "typologie_reelle": str(_cell(row, "typologie_reelle") or "").strip(),
            "month_a_label": _MONTH_A_LABEL,
            "month_b_label": _MONTH_B_LABEL,
            "puissance_facturee_a": _num(_cell(row, "puissance_facturee_a")),
            "puissance_facturee_b": _num(_cell(row, "puissance_facturee_b")),
            "redevance_grid_a": _num(_cell(row, "redevance_grid_a")),
            "redevance_grid_b": _num(_cell(row, "redevance_grid_b")),
            "estimation_conso_kwh_a": _num(_cell(row, "estimation_conso_kwh_a")),
            "estimation_conso_kwh_b": _num(_cell(row, "estimation_conso_kwh_b")),
            "estimation_conso_xof_a": _num(_cell(row, "estimation_conso_xof_a")),
            "estimation_conso_xof_b": _num(_cell(row, "estimation_conso_xof_b")),
            "redevance_vs_estimation_a": _num(_cell(row, "redevance_vs_estimation_a")),
            "redevance_vs_estimation_b": _num(_cell(row, "redevance_vs_estimation_b")),
            "statut_marge": str(_cell(row, "statut_marge") or "").strip(),
            "grid_action_plan_ops": str(_cell(row, "grid_action_plan_ops") or "").strip(),
            "categorie_ops": str(_cell(row, "categorie_ops") or "").strip(),
            "commentaire_bo": str(_cell(row, "commentaire_bo") or "").strip(),
            "categorie_bo": categorie_bo,
            "categorie_bo_autre": categorie_bo_autre,
            "check_done": _bool_done(_cell(row, "check")),
            "action_owner": action_owner,
            "action_owner_autre": action_owner_autre,
            "commentaire": str(_cell(row, "commentaire") or "").strip(),
            "row_number": i,
        })

    wb.close()
    return records


class BOMarginSnapshotImportView(APIView):
    """
    POST /api/bo-analysis/snapshots/import/
    Fichier : Analyse Marge.xlsx — import figé, historique de référence (pas de pipeline récurrent).
    """
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Aucun fichier fourni."}, status=400)

        try:
            records = _parse_bo_margin_xlsx(f.read())
        except Exception as e:
            return Response({"detail": f"Erreur lecture fichier : {e}"}, status=400)

        if not records:
            return Response({"detail": "Aucune ligne valide trouvée."}, status=400)

        site_map = dict(Site.objects.values_list("site_id", "pk"))

        created = 0
        errors = []
        with transaction.atomic():
            for rec in records:
                try:
                    site_pk = site_map.get(rec["site_id_raw"])
                    BOMarginSnapshot.objects.create(
                        site_id=site_pk,
                        source_filename=f.name,
                        **rec,
                    )
                    created += 1
                except Exception as e:
                    errors.append({"site_id": rec["site_id_raw"], "error": str(e)})

        return Response(
            {
                "message": "Import du snapshot historique BO terminé.",
                "total_parsed": len(records),
                "created": created,
                "skipped": len(errors),
                "errors_sample": errors[:20],
            },
            status=status.HTTP_201_CREATED,
        )


class BOMarginSnapshotListView(APIView):
    """GET /api/bo-analysis/snapshots/?site=&categorie_bo=&zone=&search=&page=&page_size="""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = BOMarginSnapshot.objects.select_related("site").order_by("site_id_raw")

        site = request.query_params.get("site")
        categorie = request.query_params.get("categorie_bo")
        zone = request.query_params.get("zone")
        search = request.query_params.get("search")

        if site:
            qs = qs.filter(site_id_raw=site)
        if categorie:
            qs = qs.filter(categorie_bo=categorie)
        if zone:
            qs = qs.filter(zone__iexact=zone)
        if search:
            qs = qs.filter(Q(site_id_raw__icontains=search) | Q(site_name_raw__icontains=search))

        total = qs.count()
        try:
            page = max(1, int(request.query_params.get("page", 1)))
            page_size = min(200, max(10, int(request.query_params.get("page_size", 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50

        offset = (page - 1) * page_size
        qs = qs[offset: offset + page_size]

        data = [
            {
                "id": o.id,
                "site_id": o.site_id_raw,
                "site_name": o.site_name_raw,
                "zone": o.zone,
                "typologie_reelle": o.typologie_reelle,
                "statut_marge": o.statut_marge,
                "categorie_bo": o.categorie_bo,
                "categorie_bo_display": o.get_categorie_bo_display() if o.categorie_bo else None,
                "categorie_bo_autre": o.categorie_bo_autre,
                "commentaire_bo": o.commentaire_bo,
                "action_owner": o.action_owner,
                "action_owner_display": o.get_action_owner_display() if o.action_owner else None,
                "action_owner_autre": o.action_owner_autre,
                "commentaire": o.commentaire,
                "check_done": o.check_done,
                "redevance_vs_estimation_a": o.redevance_vs_estimation_a,
                "redevance_vs_estimation_b": o.redevance_vs_estimation_b,
                "month_a_label": o.month_a_label,
                "month_b_label": o.month_b_label,
            }
            for o in qs
        ]
        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size if page_size else 1,
            "results": data,
        })


# ─────────────────────────────────────────────────────────────────────────────
# Workflow BO in-app : activation → alerte → analyse → alerte
# ─────────────────────────────────────────────────────────────────────────────

def _serialize_request(req: BOAnalysisRequest) -> dict:
    analysis = getattr(req, "analysis", None)
    targeted = list(req.targeted_bos.all())
    return {
        "id": req.id,
        "site_id": req.site.site_id,
        "site_name": req.site.name,
        "year": req.year,
        "month": req.month,
        "status": req.status,
        "notes": req.notes,
        "requested_by": req.requested_by_id,
        "requested_by_username": req.requested_by.username,
        "assigned_bo": req.assigned_bo_id,
        "assigned_bo_username": req.assigned_bo.username if req.assigned_bo else None,
        "targeted_bos": [u.id for u in targeted],
        "targeted_bos_usernames": [u.username for u in targeted],
        "requested_at": req.requested_at,
        "updated_at": req.updated_at,
        "analysis": None if not analysis else {
            "categorie_bo": analysis.categorie_bo,
            "categorie_bo_display": analysis.get_categorie_bo_display(),
            "categorie_bo_autre": analysis.categorie_bo_autre,
            "commentaire_bo": analysis.commentaire_bo,
            "action_owner": analysis.action_owner,
            "action_owner_display": analysis.get_action_owner_display(),
            "action_owner_autre": analysis.action_owner_autre,
            "commentaire": analysis.commentaire,
            "check_done": analysis.check_done,
            "submitted_by_username": analysis.submitted_by.username,
            "submitted_at": analysis.submitted_at,
        },
    }


class BOAnalysisRequestListCreateView(APIView):
    """
    GET/POST /api/bo-analysis/requests/

    - Un user 'bo' ne voit que ses demandes assignées + le pool 'pending' non assigné.
    - Un user admin/analyst voit tout (option ?mine=true pour ne voir que ses propres demandes créées).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = (
            BOAnalysisRequest.objects
            .select_related("site", "requested_by", "assigned_bo", "analysis")
            .prefetch_related("targeted_bos")
        )

        if request.user.role == "bo":
            qs = qs.filter(
                Q(assigned_bo=request.user)
                | Q(assigned_bo__isnull=True, status="pending", targeted_bos__isnull=True)
                | Q(assigned_bo__isnull=True, status="pending", targeted_bos=request.user)
            ).distinct()
        elif request.query_params.get("mine") == "true":
            qs = qs.filter(requested_by=request.user)

        site = request.query_params.get("site")
        status_filter = request.query_params.get("status")
        year = request.query_params.get("year")
        month = request.query_params.get("month")

        if site:
            qs = qs.filter(site__site_id=site)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if year:
            qs = qs.filter(year=int(year))
        if month:
            qs = qs.filter(month=int(month))

        total = qs.count()
        try:
            page = max(1, int(request.query_params.get("page", 1)))
            page_size = min(200, max(10, int(request.query_params.get("page_size", 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50
        offset = (page - 1) * page_size
        qs = qs[offset: offset + page_size]

        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size if page_size else 1,
            "results": [_serialize_request(r) for r in qs],
        })

    def post(self, request):
        site_id = request.data.get("site_id")
        year = request.data.get("year")
        month = request.data.get("month")

        if not site_id or not year or not month:
            return Response({"detail": "site_id, year et month requis."}, status=400)

        site = get_object_or_404(Site, site_id=site_id)

        try:
            year, month = int(year), int(month)
            if not (1 <= month <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({"detail": "year/month invalides."}, status=400)

        targeted_ids = request.data.get("targeted_bos") or []
        targeted_users = []
        if targeted_ids:
            from users.models import CustomUser
            targeted_users = list(CustomUser.objects.filter(pk__in=targeted_ids, role="bo"))

        req = BOAnalysisRequest.objects.create(
            site=site, year=year, month=month,
            requested_by=request.user,
            notes=request.data.get("notes", ""),
        )
        if targeted_users:
            req.targeted_bos.set(targeted_users)

        notify_bo_request_created.delay(req.id)

        return Response(_serialize_request(req), status=status.HTTP_201_CREATED)


class BOAnalysisRequestBulkCreateView(APIView):
    """
    POST /api/bo-analysis/requests/bulk/
    Body : {items: [{site_id, year, month}], targeted_bos?: [id], notes?: str}

    Active une demande d'analyse BO sur plusieurs sites en une fois (ex: tous
    les sites en marge NOK d'une période filtrée). Déclenche UNE seule
    notification groupée par BO destinataire (pas une par site).
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        items = request.data.get("items") or []
        if not items or not isinstance(items, list):
            return Response({"detail": "items requis (liste de {site_id, year, month})."}, status=400)

        targeted_ids = request.data.get("targeted_bos") or []
        targeted_users = []
        if targeted_ids:
            from users.models import CustomUser
            targeted_users = list(CustomUser.objects.filter(pk__in=targeted_ids, role="bo"))

        notes = request.data.get("notes", "")

        site_ids = {str(it.get("site_id")) for it in items if it.get("site_id")}
        site_map = dict(Site.objects.filter(site_id__in=site_ids).values_list("site_id", "pk"))

        created_ids = []
        errors = []

        with transaction.atomic():
            for it in items:
                sid = str(it.get("site_id") or "")
                site_pk = site_map.get(sid)
                if not site_pk:
                    errors.append({"site_id": sid, "error": "Site introuvable."})
                    continue
                try:
                    year, month = int(it["year"]), int(it["month"])
                except (KeyError, TypeError, ValueError):
                    errors.append({"site_id": sid, "error": "year/month invalides."})
                    continue

                req = BOAnalysisRequest.objects.create(
                    site_id=site_pk, year=year, month=month,
                    requested_by=request.user, notes=notes,
                )
                if targeted_users:
                    req.targeted_bos.set(targeted_users)
                created_ids.append(req.id)

        if created_ids:
            notify_bo_bulk_request_created.delay(created_ids)

        return Response(
            {
                "message": f"{len(created_ids)} demande(s) d'analyse BO créée(s).",
                "created": len(created_ids),
                "request_ids": created_ids,
                "errors": errors,
            },
            status=status.HTTP_201_CREATED,
        )


class BOAnalysisRequestDetailView(APIView):
    """GET/PATCH /api/bo-analysis/requests/<id>/"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        req = get_object_or_404(
            BOAnalysisRequest.objects.select_related("site", "requested_by", "assigned_bo", "analysis"), pk=pk
        )
        return Response(_serialize_request(req))

    def patch(self, request, pk):
        req = get_object_or_404(BOAnalysisRequest, pk=pk)

        assigned_bo_id = request.data.get("assigned_bo")
        if assigned_bo_id is not None:
            from users.models import CustomUser
            req.assigned_bo = get_object_or_404(CustomUser, pk=assigned_bo_id, role="bo") if assigned_bo_id else None

        new_status = request.data.get("status")
        if new_status:
            if new_status not in BOAnalysisRequest.Status.values:
                return Response({"detail": "Statut invalide."}, status=400)
            req.status = new_status

        req.save()
        return Response(_serialize_request(req))


class BOAnalysisSubmitView(APIView):
    """
    POST /api/bo-analysis/requests/<id>/submit/
    Body: {categorie_bo, categorie_bo_autre?, commentaire_bo?, action_owner, action_owner_autre?, commentaire?, check_done?}
    Seul le BO assigné (ou tout BO si la demande est encore non assignée) peut soumettre.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        req = get_object_or_404(BOAnalysisRequest.objects.select_related("site", "requested_by"), pk=pk)

        if request.user.role != "bo":
            return Response({"detail": "Seul un utilisateur Back Office peut soumettre une analyse."}, status=403)

        if req.assigned_bo_id and req.assigned_bo_id != request.user.id:
            return Response({"detail": "Cette demande est assignée à un autre BO."}, status=403)

        if hasattr(req, "analysis"):
            return Response({"detail": "Une analyse a déjà été soumise pour cette demande."}, status=400)

        categorie_bo = request.data.get("categorie_bo")
        action_owner = request.data.get("action_owner")
        if not categorie_bo or not action_owner:
            return Response({"detail": "categorie_bo et action_owner requis."}, status=400)
        if categorie_bo not in CategorieBO.values or action_owner not in ActionOwner.values:
            return Response({"detail": "categorie_bo/action_owner invalide."}, status=400)

        with transaction.atomic():
            if not req.assigned_bo_id:
                req.assigned_bo = request.user

            BOAnalysis.objects.create(
                request=req,
                categorie_bo=categorie_bo,
                categorie_bo_autre=request.data.get("categorie_bo_autre", ""),
                commentaire_bo=request.data.get("commentaire_bo", ""),
                action_owner=action_owner,
                action_owner_autre=request.data.get("action_owner_autre", ""),
                commentaire=request.data.get("commentaire", ""),
                check_done=bool(request.data.get("check_done", False)),
                submitted_by=request.user,
            )
            req.status = BOAnalysisRequest.Status.DONE
            req.save(update_fields=["status", "assigned_bo", "updated_at"])

        notify_bo_analysis_submitted.delay(req.id)

        return Response(_serialize_request(req), status=status.HTTP_201_CREATED)
